from flask import Flask, render_template, request, redirect, url_for, flash, get_flashed_messages, session, send_file, jsonify, Response
from datetime import datetime, timedelta, time, date
from dotenv import load_dotenv
load_dotenv()
from flask_socketio import SocketIO, emit, join_room
from io import BytesIO
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf.csrf import CSRFProtect, CSRFError
from utils.db import users, leaves, salaries, timetable, init_db, db, leave_class_allocations, faculty_notifications, timetable_history, leave_drafts, leave_types, hod_requests, department_hods, permissions, broadcast_notifications, staff_conversations, staff_messages, staff_socket_sessions, password_resets
from bson.objectid import ObjectId
import os
import secrets
import pandas as pd
from chatbot_engine import get_hrms_response_stream
from utils.auth import (
    admin_required,
    lecturer_required,
    salary_access_required,
    is_salary_unlocked,
    verify_salary_password,
)
from utils.salary_pdf import build_salary_pdf_bytes
from utils.salary_email import (
    is_valid_email,
    smtp_configured,
    smtp_status_message,
    send_salary_slip_email,
    send_password_reset_email,
    send_lockout_security_alert_email,
    send_test_warning_email,
    get_payroll_smtp_for_admin,
    save_payroll_smtp,
    clear_payroll_smtp,
    test_smtp_login,
    _looks_like_gmail_app_password,
)
from utils.security import (
    get_lockout_status,
    record_failed_attempt,
    record_successful_login,
    validate_password_policy,
    is_strong_password,
    is_valid_phone,
)

app = Flask(__name__)
app.jinja_env.add_extension('jinja2.ext.do')


@app.template_filter('salary_display')
def salary_display_amount(v):
    """Show blank in slip/view when amount is 0 or empty."""
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    try:
        if float(s.replace(",", "")) == 0:
            return ""
    except Exception:
        pass
    return s


def _salary_display_amt(v):
    return salary_display_amount(v)


@app.template_filter('avatar_initials')
def avatar_initials_filter(name):
    """
    Extracts 2-letter initials ignoring honorific titles (Mr., Mrs., Dr., Prof., Ms., etc.)
    Example: 'Mr. Pranam R Betrabet' -> 'PR'
    """
    if not name:
        return "U"
    titles = {"mr", "mr.", "mrs", "mrs.", "ms", "ms.", "dr", "dr.", "prof", "prof.", "sir", "madam"}
    parts = [p.strip() for p in str(name).split() if p.strip()]
    
    while parts and parts[0].lower() in titles:
        parts.pop(0)
        
    if not parts:
        return "U"
    
    if len(parts) == 1:
        return parts[0][:2].upper()
    
    return (parts[0][0] + parts[1][0]).upper()


@app.template_filter('format_date_dmy')
@app.template_filter('dmy')
def format_date_dmy_filter(value):
    """
    Formats dates into DD-MM-YYYY format across Jinja templates.
    Example: '2026-07-30' -> '30-07-2026'
    """
    if not value:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d-%m-%Y")
    
    val_str = str(value).strip()
    if not val_str:
        return ""
        
    try:
        # Check if it starts with YYYY-MM-DD
        m = re.match(r'^(?P<y>\d{4})-(?P<m>\d{2})-(?P<d>\d{2})(?P<rest>.*)$', val_str)
        if m:
            return f"{m.group('d')}-{m.group('m')}-{m.group('y')}{m.group('rest')}"
    except Exception:
        pass
        
    return val_str


app.secret_key = os.getenv("SECRET_KEY") or os.urandom(24)
app.config.setdefault("WTF_CSRF_TIME_LIMIT", None)
socketio = SocketIO(app, cors_allowed_origins="*")

bcrypt = Bcrypt(app)
csrf = CSRFProtect(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json:
        return jsonify({
            "ok": False,
            "success": False,
            "message": "Security token expired. Refresh the page and try again.",
        }), 400
    flash("Security token expired. Please try again.", "danger")
    return redirect(request.referrer or url_for("index"))


def emit_to_user(event_name, user_id, payload=None):
    """Deliver a Socket.IO event to one logged-in user (user_{id} room)."""
    if not user_id:
        return
    uid = str(user_id)
    data = dict(payload or {})
    data.setdefault("recipient_id", uid)
    socketio.emit(event_name, data, room=f"user_{uid}")

@app.errorhandler(500)
def handle_500(e):
    import traceback
    with open("chatbot_errors.log", "a") as f:
        f.write(f"\n[500 ERROR] {datetime.now()}\n")
        traceback.print_exc(file=f)
    return jsonify({"success": False, "message": "Internal Server Error", "error": str(e)}), 500


class User(UserMixin):
    def __init__(self, user_data):
        self.id = str(user_data['_id'])
        self.username = user_data['username']
        self.role = user_data['role']
        self.name = user_data.get('name', '')
        self.staff_id = user_data.get('staff_id', '')
        self.designation = user_data.get('designation', '')
        self.department = user_data.get('department', '')
        self.category = user_data.get('category', '')
        self.email = user_data.get('email', '')
        self.phone = user_data.get('phone', '')
        self.display_password = user_data.get('display_password', '')

@login_manager.user_loader
def load_user(user_id):
    user_data = users.find_one({"_id": ObjectId(user_id)})
    if user_data:
        return User(user_data)

@app.route('/lecturer/update-phone', methods=['POST'])
@lecturer_required
def lecturer_update_phone():
    new_phone = request.json.get('phone')
    if not new_phone:
        return jsonify({"success": False, "message": "Phone number is required"}), 400
    
    users.update_one({"_id": ObjectId(current_user.id)}, {"$set": {"phone": new_phone}})
    return jsonify({"success": True, "message": "Phone number updated successfully"})
    return None

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('lecturer_dashboard'))
    return render_template('landing.html')

@app.before_request
def check_user_security_and_lockout():
    if current_user and current_user.is_authenticated:
        # 1. Check account lockout status
        lockout = get_lockout_status(current_user.username)
        if lockout["is_locked"]:
            rem_fmt = lockout["formatted_time"]
            logout_user()
            flash(f"Your account ID has been locked for security. Please wait {rem_fmt} before logging in again.", "danger")
            return redirect(url_for('login', username=current_user.username))

        # 2. Enforce mandatory password & profile (email, phone) update
        endpoint = request.endpoint or ''
        allowed_endpoints = ('change_password', 'logout', 'static')
        if endpoint not in allowed_endpoints and not request.path.startswith('/static') and not request.path.startswith('/api'):
            user_doc = users.find_one({"_id": ObjectId(current_user.id)})
            if user_doc:
                disp_pass = user_doc.get("display_password", "")
                must_change = user_doc.get("must_change_password", False)
                user_email = (user_doc.get("email") or "").strip()
                user_phone = (user_doc.get("phone") or "").strip()

                if must_change or (disp_pass and not is_strong_password(disp_pass)) or not is_valid_email(user_email) or not user_phone:
                    flash("Security Setup Required: Please set a strong password, registered email, and phone number to open the dashboard.", "warning")
                    return redirect(url_for('change_password'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    user_doc = users.find_one({"_id": ObjectId(current_user.id)})
    if not user_doc:
        flash("User account not found.", "danger")
        return redirect(url_for('logout'))

    user_email = user_doc.get('email', '')
    user_phone = user_doc.get('phone', '')

    if request.method == 'POST':
        curr_pass = request.form.get('current_password') or ''
        new_pass = request.form.get('new_password') or ''
        confirm_pass = request.form.get('confirm_password') or ''
        email = (request.form.get('email') or '').strip()
        phone = (request.form.get('phone') or '').strip()

        # 1. Verify current password
        if not bcrypt.check_password_hash(user_doc['password'], curr_pass):
            flash("Current password entered is incorrect.", "danger")
            return render_template('change_password.html', user_email=email, user_phone=phone)

        # 2. Check passwords match
        if new_pass != confirm_pass:
            flash("New passwords do not match. Please re-enter carefully.", "danger")
            return render_template('change_password.html', user_email=email, user_phone=phone)

        # 3. Validate new password policy
        valid_pass, pass_msg = validate_password_policy(new_pass)
        if not valid_pass:
            flash(f"Password Policy Error: {pass_msg}", "danger")
            return render_template('change_password.html', user_email=email, user_phone=phone)

        # 4. Validate Email Address
        if not is_valid_email(email):
            flash("Invalid Email Format: Please enter a valid email address (e.g. faculty@college.edu).", "danger")
            return render_template('change_password.html', user_email=email, user_phone=phone)

        # 5. Validate Phone Number (Strict 10-digit mobile number)
        if not is_valid_phone(phone):
            flash("Invalid Mobile Phone Number: Please enter a valid 10-digit mobile number starting with 6, 7, 8, or 9 (e.g. 9876543210).", "danger")
            return render_template('change_password.html', user_email=email, user_phone=phone)

        hashed_pw = bcrypt.generate_password_hash(new_pass).decode('utf-8')

        # 6. Update user document in MongoDB
        users.update_one(
            {"_id": ObjectId(current_user.id)},
            {
                "$set": {
                    "password": hashed_pw,
                    "display_password": new_pass,
                    "email": email,
                    "phone": phone,
                    "must_change_password": False
                }
            }
        )

        flash("Profile security details updated successfully! Welcome to HRMS.", "success")
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('lecturer_dashboard'))

    return render_template('change_password.html', user_email=user_email, user_phone=user_phone)

@app.route('/login', methods=['GET', 'POST'])
def login():
    lockout_info = None
    target_username = ""

    if request.method == 'POST':
        raw_user = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        role_param = request.args.get('role', '')

        # 1. Verify if username/staff_id exists in DB first
        user_data = users.find_one({
            "$or": [
                {"username": raw_user},
                {"username": raw_user.lower()},
                {"staff_id": raw_user},
                {"staff_id": raw_user.upper()}
            ]
        })

        if not user_data:
            # Username is wrong/invalid -> Flash 'Invalid Username' without setting lockout attempts on valid accounts
            flash('Invalid Username', 'danger')
            return redirect(url_for('login', username=raw_user, role=role_param))

        canonical_username = user_data.get('username', raw_user).lower()
        user_role = user_data.get('role', '')

        # 2. Strict Role Portal Validation: Block Faculty on Management Login & Vice Versa
        if role_param == 'admin' and user_role != 'admin':
            flash("Access Denied: Faculty credentials cannot be used for Management Login. Please use Faculty Login.", "danger")
            return redirect(url_for('login', username=raw_user, role=role_param))

        if role_param == 'lecturer' and user_role == 'admin':
            flash("Access Denied: Management credentials cannot be used for Faculty Login. Please use Management Login.", "danger")
            return redirect(url_for('login', username=raw_user, role='admin'))

        # 3. Check if this correct User ID / Username is currently locked out
        lockout_status = get_lockout_status(canonical_username)
        if lockout_status["is_locked"]:
            rem_fmt = lockout_status["formatted_time"]
            flash(f"Account security lockout: Too many failed password attempts on this ID. Please wait {rem_fmt} before trying again.", "danger")
            return redirect(url_for('login', username=raw_user, role=role_param))

        # 3. Username is CORRECT, now check if password is correct
        if not bcrypt.check_password_hash(user_data['password'], password):
            # Username IS CORRECT, but password is WRONG! Record failed attempt for this user ID.
            fail_res = record_failed_attempt(canonical_username)
            if fail_res["is_locked"]:
                rem_fmt = fail_res["formatted_time"]
                email = (user_data.get("email") or "").strip()
                if is_valid_email(email):
                    token = secrets.token_urlsafe(32)
                    now = datetime.utcnow()
                    expires_at = now + timedelta(minutes=30)
                    password_resets.delete_many({"user_id": str(user_data["_id"]), "type": "unlock"})
                    password_resets.insert_one({
                        "token": token,
                        "user_id": str(user_data["_id"]),
                        "username": canonical_username,
                        "email": email,
                        "type": "unlock",
                        "created_at": now,
                        "expires_at": expires_at,
                        "used": False
                    })
                    unlock_url = url_for('unlock_account', token=token, _external=True)
                    send_lockout_security_alert_email(email, user_data.get('name', 'User'), unlock_url, rem_fmt)
                    email_parts = email.split("@")
                    masked_email = f"{email_parts[0][0]}***@{email_parts[1]}" if len(email_parts[0]) > 1 else email
                    flash(f"Account locked! Security alert & instant unlock link sent to: {masked_email}. Use the email link or Reset Password to stop timer immediately.", "danger")
                else:
                    flash(f"Account locked due to multiple incorrect password entries! Security timer set for {rem_fmt}.", "danger")
            else:
                rem = fail_res["remaining_attempts"]
                flash(f"Invalid Password. You have {rem} attempt{'s' if rem != 1 else ''} remaining before account lockout.", "danger")
            return redirect(url_for('login', username=raw_user, role=role_param))
        else:
            # Username and Password both CORRECT -> reset lockout state & log in
            record_successful_login(canonical_username)
            user_obj = User(user_data)
            remember = 'remember' in request.form
            login_user(user_obj, remember=remember)
            if user_obj.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('lecturer_dashboard'))
    else:
        q_user = (request.args.get('username') or '').strip()
        if q_user:
            user_data = users.find_one({
                "$or": [
                    {"username": q_user},
                    {"username": q_user.lower()},
                    {"staff_id": q_user},
                    {"staff_id": q_user.upper()}
                ]
            })
            c_user = user_data.get('username', q_user).lower() if user_data else q_user.lower()
            lockout_info = get_lockout_status(c_user)
            target_username = q_user

    return render_template('login.html', lockout_info=lockout_info, username=target_username)

@app.route('/api/check-lockout', methods=['GET'])
def check_lockout_api():
    raw_user = (request.args.get('username') or '').strip()
    if not raw_user:
        return jsonify({"is_locked": False, "remaining_seconds": 0})
        
    user_data = users.find_one({
        "$or": [
            {"username": raw_user},
            {"username": raw_user.lower()},
            {"staff_id": raw_user},
            {"staff_id": raw_user.upper()}
        ]
    })
    c_user = user_data.get('username', raw_user).lower() if user_data else raw_user.lower()
    status = get_lockout_status(c_user)
    return jsonify(status)

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        identifier = (request.form.get('identifier') or '').strip()
        if not identifier:
            flash("Please enter your Username or Staff ID.", "danger")
            return render_template('forgot_password.html', identifier=identifier)

        user_data = users.find_one({
            "$or": [
                {"username": identifier},
                {"username": identifier.lower()},
                {"staff_id": identifier},
                {"staff_id": identifier.upper()}
            ]
        })

        if not user_data:
            flash("No account found matching that Username or Staff ID.", "danger")
            return render_template('forgot_password.html', identifier=identifier)

        email = (user_data.get('email') or '').strip()
        if not is_valid_email(email):
            flash(f"No valid email address registered for account '{user_data.get('name', identifier)}'. Please contact system administration.", "warning")
            return render_template('forgot_password.html', identifier=identifier)

        # Generate secure random token valid for 30 minutes
        token = secrets.token_urlsafe(32)
        now = datetime.utcnow()
        expires_at = now + timedelta(minutes=30)

        # Remove existing unused tokens for this user
        password_resets.delete_many({"user_id": str(user_data["_id"])})

        password_resets.insert_one({
            "token": token,
            "user_id": str(user_data["_id"]),
            "username": user_data.get("username", identifier).lower(),
            "email": email,
            "created_at": now,
            "expires_at": expires_at,
            "used": False
        })

        reset_url = url_for('reset_password', token=token, _external=True)

        res = send_password_reset_email(email, user_data.get('name', 'User'), reset_url)
        if res.get("ok"):
            email_parts = email.split("@")
            masked_email = f"{email_parts[0][0]}***@{email_parts[1]}" if len(email_parts[0]) > 1 else email
            flash(f"A password reset link has been sent to your registered email address ({masked_email})! Please check your inbox.", "success")
            return redirect(url_for('login', username=user_data.get("username", identifier)))
        else:
            reason = res.get("reason", "Could not send email.")
            flash(f"Email dispatch note: {reason}", "warning")
            return render_template('forgot_password.html', identifier=identifier)

    q_user = (request.args.get('username') or '').strip()
    if not q_user:
        flash("Please enter your Username or Staff ID first before clicking Forgot password.", "warning")
        return redirect(url_for('login'))

    user_data = users.find_one({
        "$or": [
            {"username": q_user},
            {"username": q_user.lower()},
            {"staff_id": q_user},
            {"staff_id": q_user.upper()}
        ]
    })

    if not user_data:
        flash(f"Invalid Username: No registered account found matching '{q_user}'.", "danger")
        return redirect(url_for('login', username=q_user))

    canonical_identifier = user_data.get('username', q_user)
    return render_template('forgot_password.html', identifier=canonical_identifier)


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    reset_doc = password_resets.find_one({"token": token, "used": False})
    if not reset_doc:
        flash("This password reset link is invalid or has already been used. Please request a new one.", "danger")
        return redirect(url_for('forgot_password'))

    now = datetime.utcnow()
    if reset_doc.get("expires_at") and now > reset_doc.get("expires_at"):
        flash("This password reset link has expired. Please request a new link.", "warning")
        return redirect(url_for('forgot_password'))

    user_data = users.find_one({"_id": ObjectId(reset_doc["user_id"])})
    if not user_data:
        flash("User account not found.", "danger")
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_pass = request.form.get('new_password') or ''
        confirm_pass = request.form.get('confirm_password') or ''

        if new_pass != confirm_pass:
            flash("Passwords do not match. Please re-enter passwords carefully.", "danger")
            return render_template('reset_password.html', token=token, username=user_data.get('username'))

        valid, msg = validate_password_policy(new_pass)
        if not valid:
            flash(f"Password Policy Error: {msg}", "danger")
            return render_template('reset_password.html', token=token, username=user_data.get('username'))

        hashed_pw = bcrypt.generate_password_hash(new_pass).decode('utf-8')

        # Update password in database
        users.update_one(
            {"_id": ObjectId(reset_doc["user_id"])},
            {"$set": {"password": hashed_pw, "display_password": new_pass}}
        )

        # Mark token as used
        password_resets.update_one({"token": token}, {"$set": {"used": True, "used_at": now}})

        # Clear any active security lockout metrics
        record_successful_login(user_data.get('username', ''))

        flash("Your password has been reset successfully! You can now sign in with your new password.", "success")
        return redirect(url_for('login', username=user_data.get('username')))

    return render_template('reset_password.html', token=token, username=user_data.get('username'))

@app.route('/unlock-account/<token>', methods=['GET'])
def unlock_account(token):
    # Find token matching unlock type
    reset_doc = password_resets.find_one({"token": token, "type": "unlock"})
    if not reset_doc:
        flash("This security unlock link has expired or has been replaced by a newer alert email. Please check your inbox for the latest security link.", "warning")
        return redirect(url_for('login'))

    now = datetime.utcnow()
    if reset_doc.get("expires_at") and now > reset_doc.get("expires_at"):
        flash("This security unlock link has expired. Please request a password reset or wait for the timer.", "warning")
        return redirect(url_for('forgot_password'))

    # Mark token used & remove all active unlock tokens for this user so old links cannot be reused
    if reset_doc.get("user_id"):
        password_resets.delete_many({"user_id": reset_doc["user_id"], "type": "unlock"})
    else:
        password_resets.update_one({"token": token}, {"$set": {"used": True, "used_at": now}})

    # Immediately cancel & clear lockout metrics in MongoDB, completely restoring Stage 0 (7 initial attempt chances)!
    raw_user = reset_doc.get("username", "")
    record_successful_login(raw_user)

    user_data = None
    if reset_doc.get("user_id"):
        try:
            user_data = users.find_one({"_id": ObjectId(reset_doc["user_id"])})
        except Exception:
            user_data = None

    if user_data:
        if user_data.get("username"):
            record_successful_login(user_data["username"])
        if user_data.get("staff_id"):
            record_successful_login(user_data["staff_id"])

    target_user = user_data.get("username", raw_user) if user_data else raw_user
    flash(f"Account unlocked successfully! Lockout timer stopped and initial 7 attempt chances restored for '{target_user}'. You can now sign in.", "success")
    return redirect(url_for('login', username=target_user))

@app.route('/logout')
@login_required
def logout():
    session.pop('salary_unlocked', None)
    session.pop('salary_unlock_next', None)
    logout_user()
    return redirect(url_for('index'))

# from datetime import datetime, timedelta, time (moved to top)
from utils.timetable_processor import extract_timetable_structure, log_event
from difflib import get_close_matches
import json
import math
import io
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import mm, inch
import re
import csv
import fitz  # PyMuPDF for optional page-level images
from pathlib import Path

def normalize_name(name: str) -> str:
    """
    Make names comparable between OCR, faculty_detail.json and DB.
    """
    name = (name or "").upper()
    name = re.sub(r"^FACULTY\s*[:\-]\s*", "", name)
    name = re.split(r"\b(MENTOR|TOTAL|DEPARTMENT|DEPT)\b", name, maxsplit=1)[0]
    name = re.sub(r"\b(MR|MRS|MS|MISS|DR|PROF|PROFESSOR)\.?\b", "", name)
    name = re.sub(r"[^A-Z\s]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name

def surname_key(norm_name: str) -> str:
    parts = (norm_name or "").split()
    return parts[-1] if parts else ""

def partial_match(norm_small: str, norm_big: str) -> bool:
    if not norm_small or not norm_big:
        return False
    small_tokens = norm_small.split()
    # A single token like "MEGHA" is too ambiguous for substring matching.
    if len(small_tokens) < 2:
        return False
    if len(norm_small) > len(norm_big):
        norm_small, norm_big = norm_big, norm_small
    if norm_small in norm_big:
        return True
    small_tokens = set(small_tokens)
    big_tokens = set(norm_big.split())
    return len(small_tokens & big_tokens) >= min(2, len(small_tokens))

def _safe_filename(base: str, fallback: str) -> str:
    base = (base or "").strip().lower()
    base = base.replace("&", "and")
    base = re.sub(r"\s+", "_", base)
    base = re.sub(r"[^a-z0-9_\-]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("._- ")
    if not base:
        base = fallback
    if base in {"con", "prn", "aux", "nul"} or re.fullmatch(r"com[1-9]|lpt[1-9]", base):
        base = f"{fallback}_{base}"
    return base[:80]

def process_timetable_async(pdf_bytes, known_faculty_names, all_lecturers, lecturers_by_staff_id, lecturers_by_norm_name, lecturers_by_surname, norm_json_name_to_staff_id, upload_folder, json_folder):
    try:
        pages_with_name, pages_without_name = pdf_to_faculty_images(
            pdf_bytes, known_faculty_names=known_faculty_names or None
        )
        
        total_pages = len(pages_with_name)
        matched_count = 0
        unmatched_pages = []

        log_event(f"Starting async processing of {total_pages} pages.", socketio=socketio)

        for idx, page in enumerate(pages_with_name):
            progress = int(((idx + 1) / total_pages) * 100)
            faculty_name_raw = page["faculty_name"]
            
            log_event(f"Processing page {idx+1}/{total_pages}: {faculty_name_raw}", socketio=socketio)
            socketio.emit('timetable_progress', {'progress': progress, 'status': f'Processing {faculty_name_raw}...'})

            if not faculty_name_raw:
                continue

            norm_ocr_name = normalize_name(faculty_name_raw)
            lecturer = None
            ocr_tokens = norm_ocr_name.split()
            is_single_token_name = len(ocr_tokens) == 1

            staff_id = norm_json_name_to_staff_id.get(norm_ocr_name)
            if staff_id:
                lecturer = lecturers_by_staff_id.get(staff_id)
            if not lecturer:
                lecturer = lecturers_by_norm_name.get(norm_ocr_name)
            if not lecturer and lecturers_by_norm_name and not is_single_token_name:
                for norm_name, lect in lecturers_by_norm_name.items():
                    if partial_match(norm_ocr_name, norm_name):
                        lecturer = lect
                        break
            if not lecturer and lecturers_by_norm_name and not is_single_token_name:
                norm_lecturer_names = list(lecturers_by_norm_name.keys())
                best = get_close_matches(norm_ocr_name, norm_lecturer_names, n=1, cutoff=0.6)
                if best:
                    lecturer = lecturers_by_norm_name.get(best[0])
            if not lecturer and is_single_token_name and lecturers_by_norm_name:
                token = ocr_tokens[0]
                token_candidates = []
                for norm_name, lect in lecturers_by_norm_name.items():
                    if token in norm_name.split():
                        token_candidates.append((norm_name, lect))
                # If we need department-based disambiguation, we may pre-extract once.
                pre_structured = None

                if len(token_candidates) == 1:
                    lecturer = token_candidates[0][1]
                elif len(token_candidates) > 1:
                    # If multiple candidates contain the same single OCR token,
                    # prefer the one with a "non-generic" surname (avoid common ambiguous surnames).
                    GENERIC_TOKENS = {"SHETTY", "RAO", "NAYAK", "KUMAR", "SINGH", "DEVI", "SHARMA"}
                    # Prefer the lecturer whose DB name is just the single token (no surname).
                    no_surname = [(n, l) for (n, l) in token_candidates if len((n or "").split()) == 1]
                    if len(no_surname) == 1:
                        lecturer = no_surname[0][1]
                    else:
                        # 1) Department-based disambiguation using the *image* (most reliable).
                        try:
                            log_event(
                                f"Ambiguous single-name OCR '{faculty_name_raw}' - extracting department for disambiguation...",
                                socketio=socketio
                            )
                            pre_structured = extract_timetable_structure(page["image"], faculty_name_hint=None, socketio=socketio)
                        except Exception:
                            pre_structured = None

                        extracted_dept = ((pre_structured or {}).get("department") or "").strip().lower()
                        if extracted_dept:
                            dept_hits = []
                            for _, l in token_candidates:
                                l_dept = (l.get("department") or "").strip().lower()
                                if l_dept and (extracted_dept in l_dept or l_dept in extracted_dept):
                                    dept_hits.append(l)
                            if len(dept_hits) == 1:
                                lecturer = dept_hits[0]
                                log_event(
                                    f"Disambiguated by department '{extracted_dept}': {faculty_name_raw} -> {lecturer.get('name','')}",
                                    socketio=socketio
                                )

                        # 2) Generic-surname penalty tie-break (only if still unmatched).
                        if not lecturer:
                            scored = []
                            for n, l in token_candidates:
                                parts = (n or "").split()
                                extra = [p for p in parts[1:] if p and p != token]
                                generic_count = sum(1 for p in extra if p in GENERIC_TOKENS)
                                non_generic_count = sum(1 for p in extra if p and p not in GENERIC_TOKENS)
                                # Prefer: fewer generic tokens, more non-generic tokens.
                                scored.append(((generic_count, -non_generic_count, len(extra)), l))
                            scored.sort(key=lambda x: x[0])
                            best_score = scored[0][0] if scored else None
                            best = [l for (s, l) in scored if s == best_score]
                            if len(best) == 1:
                                lecturer = best[0]

                        # 3) If still ambiguous, do not guess.
                        if not lecturer:
                            candidate_names = [lect.get("name", "") for _, lect in token_candidates]
                            log_event(
                                f"Ambiguous single-name OCR '{faculty_name_raw}' -> skipping auto-match. Candidates: {candidate_names}",
                                socketio=socketio
                            )
            if not lecturer and not is_single_token_name:
                sk = surname_key(norm_ocr_name)
                if sk:
                    candidates = lecturers_by_surname.get(sk, [])
                    if len(candidates) == 1:
                        lecturer = candidates[0]

            if not lecturer:
                unmatched_pages.append({
                    "faculty_name": faculty_name_raw,
                    "normalized_name": norm_ocr_name,
                    "page_index": page["page_index"],
                })
                log_event(f"FAILED to match lecturer for: {faculty_name_raw}", socketio=socketio)
                continue

            matched_display_name = lecturer.get("name", faculty_name_raw)
            log_event(f"Matched to: {matched_display_name}", socketio=socketio)

            fallback = f"lect_{str(lecturer['_id'])}"
            safe_name = _safe_filename(matched_display_name, fallback=fallback)
            filename = f"{safe_name}.png"
            fs_image_path = os.path.join(upload_folder, filename)
            page["image"].save(fs_image_path, format="PNG")
            url_image_path = f"timetables/{filename}"

            # If we already extracted once for department disambiguation, reuse it.
            if "pre_structured" in locals() and pre_structured:
                structured = pre_structured
                log_event(f"Reusing extracted structure for {matched_display_name}.", socketio=socketio)
            else:
                log_event(f"Extracting structure via Gemini for {matched_display_name}...", socketio=socketio)
                structured = extract_timetable_structure(page["image"], faculty_name_hint=matched_display_name, socketio=socketio)
            
            timetable.update_one(
                {"lecturer_id": str(lecturer["_id"])},
                {
                    "$set": {
                        "lecturer_id": str(lecturer["_id"]),
                        "lecturer_name": matched_display_name,
                        "image_path": url_image_path,
                        "structured": structured or {},
                        "uploaded_at": datetime.now(),
                    }
                },
                upsert=True,
            )
            # Save to separate JSON file by staff_id
            staff_id = lecturer.get("staff_id")
            if staff_id:
                json_filename = f"{staff_id}.json"
                json_path = os.path.join(json_folder, json_filename)
                try:
                    with open(json_path, "w", encoding="utf-8") as jf:
                        json.dump(structured or {}, jf, indent=2, ensure_ascii=False)
                    log_event(f"JSON saved: {json_filename}", socketio=socketio)
                except Exception as je:
                    log_event(f"Failed to save JSON for {staff_id}: {je}", socketio=socketio)

            matched_count += 1
            log_event(f"SUCCESS: Saved timetable for {matched_display_name}", socketio=socketio)

        final_msg = f"Completed. Matched {matched_count}/{total_pages} timetables."
        log_event(final_msg, socketio=socketio)
        socketio.emit('timetable_progress', {'progress': 100, 'status': final_msg, 'done': True})

    except Exception as e:
        err_msg = f"Async Error: {str(e)}"
        log_event(err_msg, socketio=socketio)
        socketio.emit('timetable_progress', {'progress': 0, 'status': err_msg, 'error': True})

# Leave Types Management Routes
@app.route('/admin/api/leave-types', methods=['GET'])
@login_required
@admin_required
def get_leave_types_api():
    types = list(leave_types.find().sort("name", 1))
    # Calculate usage stats for each type
    all_lecturers = list(users.find({"role": "lecturer"}))
    
    result = []
    for t in types:
        name = t['name']
        # Unify to leave_balances
        total_allocated = sum(float(l.get('leave_balances', {}).get(name, 0)) for l in all_lecturers)
        # Also count how many lecturers have this type
        in_use_count = sum(1 for l in all_lecturers if name in l.get('leave_balances', {}))
        
        result.append({
            "id": str(t['_id']),
            "name": name,
            "total_allocated": total_allocated,
            "in_use_count": in_use_count
        })
    return jsonify(result)

@app.route('/admin/api/leave-types', methods=['POST'])
@login_required
@admin_required
def add_leave_type():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"success": False, "message": "Name is required"}), 400
    if leave_types.find_one({"name": name}):
        return jsonify({"success": False, "message": "Type already exists"}), 400
    leave_types.insert_one({"name": name})
    socketio.emit('leave_types_updated')
    return jsonify({"success": True})

@app.route('/admin/api/leave-types/<name>', methods=['DELETE'])
@login_required
@admin_required
def delete_global_leave_type(name):
    # 1. Remove from global list
    res = leave_types.delete_one({"name": name})
    
    # 2. CASCADING DELETE: Remove this type from all lecturers' balances
    if res.deleted_count > 0:
        all_lecturers = users.find({"role": "lecturer", f"leave_balances.{name}": {"$exists": True}})
        for l in all_lecturers:
            balances = l.get("leave_balances", {})
            if name in balances:
                del balances[name]
                users.update_one({"_id": l["_id"]}, {"$set": {"leave_balances": balances}})
    
    socketio.emit('leave_types_updated')
    return jsonify({"success": res.deleted_count > 0})

@app.route('/admin/api/leave-types/cleanup', methods=['POST'])
@login_required
@admin_required
def cleanup_unused_types():
    # Find all types currently in use by any lecturer
    in_use = set()
    for l in users.find({"role": "lecturer"}):
        balances = l.get('leave_balances', {})
        for k, v in balances.items():
            try:
                if float(v) > 0: in_use.add(k)
            except: continue
            
    # Delete global types not in use
    res = leave_types.delete_many({"name": {"$nin": list(in_use)}})
    socketio.emit('leave_types_updated')
    return jsonify({"success": True, "deleted_count": res.deleted_count})

@app.route('/admin/api/leave-types/clear-all', methods=['POST'])
@login_required
@admin_required
def clear_all_leave_config():
    """Wipe all types and all associated balcony allocations (Full Reset)"""
    # 1. Clear global types
    leave_types.delete_many({})
    
    # 2. Clear all leave_balances from all lecturers
    users.update_many(
        {"role": "lecturer"},
        {"$unset": {"leave_balances": "", "leaves_per_month": ""}}
    )
    
    socketio.emit('leave_types_updated')
    return jsonify({"success": True, "message": "All leave categories and allocations cleared."})

@app.route('/admin/api/assign-hod', methods=['POST'])
@login_required
@admin_required
def assign_hod_api():
    data = request.get_json()
    dept = data.get('department')
    hod_id = data.get('hod_id')
    
    if not dept:
        return jsonify({"success": False, "message": "Department is required"}), 400
        
    if hod_id:
        department_hods.update_one(
            {"department": dept},
            {"$set": {"hod_id": hod_id}},
            upsert=True
        )
    else:
        department_hods.delete_one({"department": dept})
        
    socketio.emit('hod_assigned', {"department": dept, "hod_id": hod_id})
    return jsonify({"success": True})

# Admin Routes
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    # Consume stale flashes (e.g. from salary SMTP saves) so they do not pile up on the login page.
    get_flashed_messages(with_categories=True)

    stats = {
        "teaching_faculty_count": users.count_documents({
            "role": "lecturer",
            "staff_id": {"$regex": r"^BBHCF\d+$", "$options": "i"},
        }),
        "non_teaching_faculty_count": users.count_documents({
            "role": "lecturer",
            "staff_id": {"$regex": r"^BBHCFN\d+$", "$options": "i"},
        }),
        "pending_leaves": leaves.count_documents({"status": "Pending"}),
        "pending_permissions": permissions.count_documents({"status": "Pending"}),
    }
    
    # Unified Recent Requests (Leaves + Permissions)
    recent_l = list(leaves.find({"status": "Pending"}).sort("_id", -1).limit(5))
    recent_p = list(permissions.find({"status": "Pending"}).sort("_id", -1).limit(5))
    
    # Merge and Sort
    all_recent = sorted(recent_l + recent_p, key=lambda x: x['_id'], reverse=True)[:5]

    # Pre-serialize for use in inline JS
    recent_serialized = [
        {
            "id": str(doc.get("_id")),
            "lecturer_name": doc.get("lecturer_name", ""),
            "type": doc.get("type", "Permission"),
            "from_date": doc.get("from_date", ""),
            "to_date": doc.get("to_date", ""),
            "status": doc.get("status", ""),
            "half_day": doc.get("half_day", False),
            "session": doc.get("session", ""),
            "mode": doc.get("mode", "full")
        }
        for doc in all_recent
    ]

    # Fetch Recent Broadcasts
    broadcasts = list(broadcast_notifications.find().sort("created_at", -1).limit(5))
    print(f"DEBUG: Broadcast Count = {len(broadcasts)}")
    for b in broadcasts:
        b['_id'] = str(b['_id'])
        if isinstance(b.get('created_at'), datetime):
            b['created_at_fmt'] = b['created_at'].strftime("%Y-%m-%d %I:%M %p")
        else:
            b['created_at_fmt'] = "N/A"

    # Fetch Admin Profile Pic
    admin_doc = users.find_one({"_id": ObjectId(current_user.id)})
    p_pic = admin_doc.get('profile_pic')

    return render_template(
        'admin/dashboard.html',
        stats=stats,
        recent_leaves=all_recent,
        recent_leaves_serialized=recent_serialized,
        broadcasts=broadcasts,
        profile_pic=p_pic,
        salary_unlocked=is_salary_unlocked(),
        smtp_configured=smtp_configured(),
        smtp_status=smtp_status_message(),
        payroll_smtp=get_payroll_smtp_for_admin(),
        smtp_notice=request.args.get("smtp_notice"),
    )

def _amount_to_indian_words(amount) -> str:
    """Convert rupee amount to words (Indian lakh/crore grouping). Returns '' for zero/invalid."""
    try:
        n = int(round(float(str(amount).replace(",", ""))))
    except (TypeError, ValueError):
        return ""
    if n <= 0:
        return ""

    names = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
        "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen",
        "Eighteen", "Nineteen",
    ]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two_digit(x: int) -> str:
        if x < 20:
            return names[x]
        t, r = divmod(x, 10)
        return tens[t] + (f" {names[r]}" if r else "")

    def three_digit(x: int) -> str:
        h, r = divmod(x, 100)
        out = ""
        if h:
            out = f"{two_digit(h)} Hundred"
        if r:
            out = f"{out} {two_digit(r)}".strip() if out else two_digit(r)
        return out

    crore = n // 10000000
    lakh = (n % 10000000) // 100000
    thousand = (n % 100000) // 1000
    rest = n % 1000
    parts = []
    if crore:
        parts.append(f"{three_digit(crore)} Crore")
    if lakh:
        parts.append(f"{three_digit(lakh)} Lakh")
    if thousand:
        parts.append(f"{three_digit(thousand)} Thousand")
    if rest:
        parts.append(three_digit(rest))
    return " ".join(parts)


def _salary_net_from_payload(earnings: dict, deductions: dict) -> float:
    def _sum(d: dict) -> float:
        total = 0.0
        for v in (d or {}).values():
            try:
                total += float(str(v).replace(",", "") or 0)
            except (TypeError, ValueError):
                pass
        return total

    return _sum(earnings) - _sum(deductions)


def _salary_doc_status(doc: dict) -> str:
    """
    Return one of: not_started | partial | complete
    """
    if not doc:
        return "not_started"
    payload = doc.get("payload") or {}
    # Only consider HEADER fields for readiness status.
    # Earnings/Deductions can be legitimately empty/0 and should not affect red/yellow/green.
    required_keys = [
        "month_year",
        "employee_id",
        "employee_name",
        "department",
        "paid_days",
        "bank_ac_no",
    ]
    present = 0
    for k in required_keys:
        v = payload.get(k)
        if v is not None and str(v).strip() != "":
            present += 1

    if present == len(required_keys):
        return "complete"
    if present == 0:
        return "not_started"
    return "partial"


def _parse_month_year(value: str) -> str:
    """Normalize month filter to 'May 2026' style."""
    value = (value or "").strip()
    if not value:
        return datetime.now().strftime("%B %Y")
    if len(value) == 7 and value[4] == "-":
        try:
            return datetime.strptime(value + "-01", "%Y-%m-%d").strftime("%B %Y")
        except ValueError:
            pass
    try:
        return datetime.strptime(value + " 1", "%B %Y %d").strftime("%B %Y")
    except ValueError:
        return value


def _month_year_to_iso(month_year: str) -> str:
    try:
        return datetime.strptime((month_year or "").strip() + " 1", "%B %Y %d").strftime("%Y-%m")
    except ValueError:
        return datetime.now().strftime("%Y-%m")


def _faculty_login_form_context(staff: dict) -> dict:
    """Prefill faculty HRMS login fields on salary slip form."""
    return {
        "email": (staff.get("email") or "").strip(),
        "username": (staff.get("username") or staff.get("staff_id") or "").strip(),
        "display_password": staff.get("display_password") or "",
    }


def _apply_faculty_login_from_salary_form(staff: dict, form) -> str | None:
    """Update lecturer login from salary slip form. Returns error message or None."""
    email = (form.get("faculty_email") or "").strip()
    username = (form.get("faculty_username") or "").strip()
    password = form.get("faculty_password") or ""

    update = {}
    if email:
        update["email"] = email
    if username:
        existing = users.find_one({"username": username, "_id": {"$ne": staff["_id"]}})
        if existing:
            return f"Username '{username}' is already used by another account."
        update["username"] = username
    if password.strip():
        pwd = password.strip()
        update["password"] = bcrypt.generate_password_hash(pwd).decode("utf-8")
        update["display_password"] = pwd

    if update:
        users.update_one({"_id": staff["_id"]}, {"$set": update})
    return None


def _prefill_bank_ac_from_history(lecturer_id: str, month_year: str, payload: dict) -> str | None:
    """If bank account empty, copy from this faculty's most recent saved slip."""
    if (payload.get("bank_ac_no") or "").strip():
        return None
    prev = salaries.find_one(
        {
            "lecturer_id": lecturer_id,
            "month_year": {"$ne": month_year},
            "payload.bank_ac_no": {"$nin": ["", None]},
        },
        sort=[("updated_at", -1)],
    )
    if not prev:
        return None
    ac = ((prev.get("payload") or {}).get("bank_ac_no") or "").strip()
    if not ac:
        return None
    payload["bank_ac_no"] = ac
    return prev.get("month_year")


@app.route('/admin/salary/unlock', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_salary_unlock():
    if is_salary_unlocked():
        dest = session.pop("salary_unlock_next", None)
        return redirect(dest or url_for("admin_salary_list"))

    if request.method == 'POST':
        password = request.form.get('password') or ''
        if verify_salary_password(password):
            session['salary_unlocked'] = True
            dest = session.pop('salary_unlock_next', None) or url_for('admin_salary_list')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'ok': True, 'redirect': dest})
            return redirect(dest)
        msg = 'Incorrect password. Try again.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'message': msg}), 403
        flash(msg, 'error')
        return render_template('admin/salary_unlock.html')

    return render_template('admin/salary_unlock.html')


@app.route('/admin/salary')
@login_required
@admin_required
@salary_access_required
def admin_salary_list():
    month_year = _parse_month_year(request.args.get("month_year") or "")
    month_iso = _month_year_to_iso(month_year)

    all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
    salary_docs = list(salaries.find({"month_year": month_year}))
    by_lecturer_id = {d.get("lecturer_id"): d for d in salary_docs if d.get("lecturer_id")}

    rows = []
    complete_unpublished = 0
    published_count = 0
    email_ready_count = 0
    for s in all_staff:
        lecturer_id = str(s.get("_id"))
        doc = by_lecturer_id.get(lecturer_id)
        status = _salary_doc_status(doc)
        published = bool(doc.get("published")) if doc else False
        faculty_email = (s.get("email") or "").strip()
        has_valid_email = is_valid_email(faculty_email)
        email_hint = _salary_email_hint(status, doc, has_valid_email, faculty_email)
        can_email = smtp_configured() and status == "complete" and doc and has_valid_email
        if can_email:
            email_ready_count += 1
        if status == "complete" and not published:
            complete_unpublished += 1
        if published:
            published_count += 1
        rows.append({
            "lecturer_id": lecturer_id,
            "staff_id": s.get("staff_id", ""),
            "name": s.get("name", ""),
            "department": s.get("department", ""),
            "email": faculty_email,
            "has_valid_email": has_valid_email,
            "email_hint": email_hint,
            "can_email": can_email,
            "status": status,
            "published": published,
            "emailed_at": doc.get("emailed_at") if doc else None,
            "email_sent_to": doc.get("email_sent_to") if doc else None,
            "updated_at": doc.get("updated_at") if doc else None,
        })

    email_report = session.pop("salary_email_report", None)

    return render_template(
        "admin/salary_list.html",
        rows=rows,
        month_year=month_year,
        month_iso=month_iso,
        complete_unpublished=complete_unpublished,
        published_count=published_count,
        email_ready_count=email_ready_count,
        smtp_configured=smtp_configured(),
        smtp_status=smtp_status_message(),
        payroll_smtp=get_payroll_smtp_for_admin(),
        email_report=email_report,
    )


@app.route('/admin/dashboard/sender-email', methods=['POST'])
@login_required
@admin_required
def admin_dashboard_sender_email():
    smtp_user = (request.form.get("smtp_user") or "").strip()
    smtp_password = request.form.get("smtp_password") or ""

    if not smtp_user or not is_valid_email(smtp_user):
        return redirect(url_for("admin_dashboard", smtp_notice="invalid_email"))

    existing = get_payroll_smtp_for_admin()
    prev_user = (existing.get("smtp_user") or "").strip().lower()
    user_changed = prev_user and prev_user != smtp_user.lower()

    if not smtp_password.strip():
        if not existing.get("has_password"):
            return redirect(url_for("admin_dashboard", smtp_notice="password_required"))
        if user_changed:
            return redirect(url_for("admin_dashboard", smtp_notice="password_required_change"))
    elif not _looks_like_gmail_app_password(smtp_password):
        return redirect(url_for("admin_dashboard", smtp_notice="bad_app_password"))

    save_payroll_smtp(smtp_user, smtp_password if smtp_password.strip() else None)
    return redirect(url_for("admin_dashboard", smtp_notice="saved", smtp_user=smtp_user))


@app.route('/admin/dashboard/sender-email/test', methods=['POST'])
@login_required
@admin_required
def admin_dashboard_sender_email_test():
    smtp_user = (request.form.get("smtp_user") or "").strip()
    smtp_password = request.form.get("smtp_password") or ""
    result = test_smtp_login(smtp_user or None, smtp_password or None)
    return jsonify(result)

@app.route('/admin/test-warning-email', methods=['POST'])
@login_required
@admin_required
def admin_test_warning_email():
    target_email = (request.form.get("test_email") or "").strip()
    if not target_email:
        admin_doc = users.find_one({"_id": ObjectId(current_user.id)})
        target_email = (admin_doc.get("email") or "").strip() if admin_doc else ""

    if not is_valid_email(target_email):
        return jsonify({"ok": False, "reason": "Please provide a valid recipient email address to receive the test email."})

    res = send_test_warning_email(target_email)
    return jsonify(res)


@app.route('/admin/dashboard/sender-email/reset-env', methods=['POST'])
@login_required
@admin_required
def admin_dashboard_sender_email_reset_env():
    clear_payroll_smtp()
    return redirect(url_for("admin_dashboard", smtp_notice="reset_env"))


@app.route('/admin/salary/settings', methods=['POST'])
@login_required
@admin_required
@salary_access_required
def admin_salary_settings():
    month_year = _parse_month_year(request.form.get("month_year") or "")
    smtp_user = (request.form.get("smtp_user") or "").strip()
    smtp_password = request.form.get("smtp_password") or ""

    if not smtp_user or not is_valid_email(smtp_user):
        flash("Enter a valid sender email address.", "error")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    existing = get_payroll_smtp_for_admin()
    prev_user = (existing.get("smtp_user") or "").strip().lower()
    user_changed = prev_user and prev_user != smtp_user.lower()

    if not smtp_password.strip():
        if not existing.get("has_password"):
            flash("Gmail App Password is required for first-time setup.", "error")
            return redirect(url_for("admin_salary_list", month_year=month_year))
        if user_changed:
            flash("Enter App Password when changing the sender email.", "error")
            return redirect(url_for("admin_salary_list", month_year=month_year))
    elif not _looks_like_gmail_app_password(smtp_password):
        flash(
            "App Password must be 16 characters (from Google App Passwords). "
            "Do not use your normal Gmail login password.",
            "error",
        )
        return redirect(url_for("admin_salary_list", month_year=month_year))

    save_payroll_smtp(smtp_user, smtp_password if smtp_password.strip() else None)
    flash(f"Sender email saved: {smtp_user}. Salary slips will be sent from this account.", "success")
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/settings/reset-env', methods=['POST'])
@login_required
@admin_required
@salary_access_required
def admin_salary_settings_reset_env():
    month_year = _parse_month_year(request.form.get("month_year") or "")
    clear_payroll_smtp()
    flash("Cleared saved sender login. HRMS will use SMTP_USER and SMTP_PASSWORD from .env.", "success")
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/settings/test', methods=['POST'])
@login_required
@admin_required
@salary_access_required
def admin_salary_settings_test():
    smtp_user = (request.form.get("smtp_user") or "").strip()
    smtp_password = request.form.get("smtp_password") or ""
    result = test_smtp_login(smtp_user or None, smtp_password or None)
    return jsonify(result)


def _salary_email_hint(status, doc, has_valid_email, faculty_email):
    if not smtp_configured():
        return "SMTP not configured"
    if not doc:
        return "Save slip first"
    if status != "complete":
        return "Complete slip first"
    if not has_valid_email:
        if not faculty_email:
            return "No email — add in Manage Staff"
        return "Invalid email — fix in Manage Staff"
    return "Ready to email"


def _send_salary_slip_to_faculty(staff: dict, doc: dict, month_year: str) -> dict:
    """Send one slip. Returns {status, staff_id, name, email?, reason?}."""
    staff_id = staff.get("staff_id", "")
    name = staff.get("name", "")
    faculty_email = (staff.get("email") or "").strip()

    if not smtp_configured():
        return {
            "status": "skipped",
            "staff_id": staff_id,
            "name": name,
            "reason": "Email not configured — set SMTP_USER and SMTP_PASSWORD in .env",
        }
    if not doc:
        return {
            "status": "skipped",
            "staff_id": staff_id,
            "name": name,
            "reason": "Salary slip not saved for this month",
        }
    if _salary_doc_status(doc) != "complete":
        return {
            "status": "skipped",
            "staff_id": staff_id,
            "name": name,
            "reason": "Slip not complete — fill all header fields",
        }
    if not is_valid_email(faculty_email):
        return {
            "status": "skipped",
            "staff_id": staff_id,
            "name": name,
            "reason": "No valid email — update faculty profile in Manage Staff",
        }

    payload = doc.get("payload") or {}
    try:
        pdf_bytes, filename = build_salary_pdf_bytes(payload)
        result = send_salary_slip_email(
            faculty_email,
            name,
            staff_id,
            month_year,
            pdf_bytes,
            filename,
        )
    except Exception as e:
        return {
            "status": "failed",
            "staff_id": staff_id,
            "name": name,
            "email": faculty_email,
            "reason": f"Could not prepare email: {e}",
        }

    if not result.get("ok"):
        return {
            "status": "failed",
            "staff_id": staff_id,
            "name": name,
            "email": faculty_email,
            "reason": result.get("reason", "Send failed"),
        }

    salaries.update_one(
        {"_id": doc["_id"]},
        {"$set": {"emailed_at": datetime.now(), "email_sent_to": faculty_email}},
    )
    return {
        "status": "sent",
        "staff_id": staff_id,
        "name": name,
        "email": faculty_email,
    }


def _store_salary_email_report(month_year: str, results: list) -> None:
    sent = [r for r in results if r["status"] == "sent"]
    skipped = [r for r in results if r["status"] == "skipped"]
    failed = [r for r in results if r["status"] == "failed"]
    session["salary_email_report"] = {
        "month_year": month_year,
        "sent": sent,
        "skipped": skipped,
        "failed": failed,
        "sent_count": len(sent),
        "skipped_count": len(skipped),
        "failed_count": len(failed),
    }


@app.route('/admin/salary/bulk-email', methods=["POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_bulk_email():
    month_year = _parse_month_year(request.form.get("month_year") or "")
    if not smtp_configured():
        flash("Cannot send emails: set SMTP_USER and SMTP_PASSWORD in .env", "danger")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    salary_docs = list(salaries.find({"month_year": month_year}))
    by_lecturer_id = {d.get("lecturer_id"): d for d in salary_docs if d.get("lecturer_id")}
    all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))

    results = []
    for s in all_staff:
        lecturer_id = str(s.get("_id"))
        doc = by_lecturer_id.get(lecturer_id)
        if not doc or _salary_doc_status(doc) != "complete":
            continue
        if not is_valid_email((s.get("email") or "").strip()):
            results.append({
                "status": "skipped",
                "staff_id": s.get("staff_id", ""),
                "name": s.get("name", ""),
                "reason": "No valid email — update Manage Staff",
            })
            continue
        results.append(_send_salary_slip_to_faculty(s, doc, month_year))

    _store_salary_email_report(month_year, results)
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/<lecturer_id>/email', methods=["POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_email(lecturer_id):
    month_year = _parse_month_year(request.form.get("month_year") or "")
    staff = users.find_one({"_id": ObjectId(lecturer_id), "role": "lecturer"})
    if not staff:
        flash("Faculty not found.", "danger")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    doc = salaries.find_one({"lecturer_id": lecturer_id, "month_year": month_year})
    result = _send_salary_slip_to_faculty(staff, doc, month_year)
    _store_salary_email_report(month_year, [result])
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/bulk-publish', methods=["POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_bulk_publish():
    month_year = _parse_month_year(request.form.get("month_year") or "")
    uploaded = 0
    for doc in salaries.find({"month_year": month_year}):
        if _salary_doc_status(doc) != "complete" or doc.get("published"):
            continue
        salaries.update_one(
            {"_id": doc["_id"]},
            {"$set": {"published": True, "published_at": datetime.now()}},
        )
        uploaded += 1
    if uploaded:
        flash(f"Bulk upload complete: {uploaded} slip(s) published to faculty for {month_year}.", "success")
    else:
        flash("No completed slips ready to upload for this month.", "warning")
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/cancel-month', methods=["POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_cancel_month():
    month_year = _parse_month_year(request.form.get("month_year") or "")
    result = salaries.update_many(
        {"month_year": month_year, "published": True},
        {"$set": {"published": False}, "$unset": {"published_at": ""}},
    )
    if result.modified_count:
        flash(
            f"Cancelled uploads for {month_year}. {result.modified_count} slip(s) hidden from faculty.",
            "success",
        )
    else:
        flash(f"No uploaded slips to cancel for {month_year}.", "info")
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/<lecturer_id>/publish', methods=["POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_publish(lecturer_id):
    month_year = _parse_month_year(request.form.get("month_year") or "")

    doc = salaries.find_one({"lecturer_id": lecturer_id, "month_year": month_year})
    if not doc:
        flash("Slip not found. Please save the slip first.", "danger")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    status = _salary_doc_status(doc)
    if status != "complete":
        flash("Slip is not complete. Fill all required fields before uploading.", "warning")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    salaries.update_one(
        {"_id": doc["_id"]},
        {"$set": {"published": True, "published_at": datetime.now()}}
    )
    flash("Slip uploaded (published) to faculty.", "success")
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/<lecturer_id>/unpublish', methods=["POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_unpublish(lecturer_id):
    month_year = _parse_month_year(request.form.get("month_year") or "")
    doc = salaries.find_one({"lecturer_id": lecturer_id, "month_year": month_year})
    if not doc or not doc.get("published"):
        flash("This slip is not uploaded.", "info")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    salaries.update_one(
        {"_id": doc["_id"]},
        {"$set": {"published": False}, "$unset": {"published_at": ""}},
    )
    flash("Upload cancelled. Slip hidden from faculty.", "success")
    return redirect(url_for("admin_salary_list", month_year=month_year))


@app.route('/admin/salary/<lecturer_id>/slip', methods=["GET", "POST"])
@login_required
@admin_required
@salary_access_required
def admin_salary_slip(lecturer_id):
    staff = users.find_one({"_id": ObjectId(lecturer_id), "role": "lecturer"})
    if not staff:
        flash("Faculty not found.", "danger")
        return redirect(url_for("admin_salary_list"))

    month_year = _parse_month_year(request.values.get("month_year") or "")

    existing = salaries.find_one({"lecturer_id": lecturer_id, "month_year": month_year}) or {}
    payload = existing.get("payload") or {}

    if request.method == "POST":
        def _f(key, default=""):
            return (request.form.get(key) or default).strip()

        def _amount(key):
            raw = _f(key)
            if raw == "":
                return "0"
            try:
                # Store a clean numeric string; keep whole numbers without trailing .0
                n = float(raw.replace(",", ""))
                return str(int(n)) if n.is_integer() else f"{n:.2f}"
            except Exception:
                return "0"

        earnings = {
            "basic_pay": _amount("earn_basic_pay"),
            "da": _amount("earn_da"),
            "hra": _amount("earn_hra"),
            "spl_allowance": _amount("earn_spl_allowance"),
            "allow_phd": _amount("earn_allow_phd"),
            "hod_allowance": _amount("earn_hod_allowance"),
            "addl_remuneration": _amount("earn_addl_remuneration"),
        }
        deductions = {
            "pf": _amount("ded_pf"),
            "pt": _amount("ded_pt"),
            "esi": _amount("ded_esi"),
            "lic_premium": _amount("ded_lic"),
            "others": _amount("ded_others"),
        }

        net_pay = _salary_net_from_payload(earnings, deductions)
        words = _amount_to_indian_words(net_pay)
        netpay_words = f"{words} Only" if words else ""

        payload = {
            "month_year": month_year,
            "employee_name": _f("employee_name", staff.get("name", "")),
            "employee_id": _f("employee_id", staff.get("staff_id", "")),
            "department": _f("department", staff.get("department", "")),
            "paid_days": _f("paid_days"),
            "bank_ac_no": _f("bank_ac_no"),
            "earnings": earnings,
            "deductions": deductions,
            "netpay_words": netpay_words,
            "notes": _f("notes"),
        }

        login_err = _apply_faculty_login_from_salary_form(staff, request.form)
        if login_err:
            flash(login_err, "error")
            staff = users.find_one({"_id": ObjectId(lecturer_id), "role": "lecturer"}) or staff
            bank_auto_from = _prefill_bank_ac_from_history(lecturer_id, month_year, payload)
            faculty_login = _faculty_login_form_context(staff)
            faculty_login["email"] = (request.form.get("faculty_email") or faculty_login["email"]).strip()
            faculty_login["username"] = (request.form.get("faculty_username") or faculty_login["username"]).strip()
            return render_template(
                "admin/salary_slip_form.html",
                staff=staff,
                month_year=month_year,
                payload=payload,
                faculty_login=faculty_login,
                bank_auto_from=bank_auto_from,
            )

        salaries.update_one(
            {"lecturer_id": lecturer_id, "month_year": month_year},
            {"$set": {
                "lecturer_id": lecturer_id,
                "staff_id": staff.get("staff_id", ""),
                "lecturer_name": staff.get("name", ""),
                "month_year": month_year,
                "payload": payload,
                "updated_at": datetime.now(),
            }},
            upsert=True
        )

        flash("Salary slip saved.", "success")
        return redirect(url_for("admin_salary_list", month_year=month_year))

    # Default prefill
    if not payload:
        payload = {
            "month_year": month_year,
            "employee_name": staff.get("name", ""),
            "employee_id": staff.get("staff_id", ""),
            "department": staff.get("department", ""),
            "paid_days": "",
            "bank_ac_no": "",
            "earnings": {},
            "deductions": {},
            "notes": "",
        }

    bank_auto_from = _prefill_bank_ac_from_history(lecturer_id, month_year, payload)

    return render_template(
        "admin/salary_slip_form.html",
        staff=staff,
        month_year=month_year,
        payload=payload,
        faculty_login=_faculty_login_form_context(staff),
        bank_auto_from=bank_auto_from,
    )


@app.route('/admin/api/recent-leaves')
@login_required
@admin_required
def admin_api_recent_leaves():
    """
    Unified JSON API for polling both pending leaves and permissions on the dashboard
    """
    recent_l = list(leaves.find({"status": "Pending"}).sort("_id", -1).limit(5))
    recent_p = list(permissions.find({"status": "Pending"}).sort("_id", -1).limit(5))
    
    all_recent = sorted(recent_l + recent_p, key=lambda x: x['_id'], reverse=True)[:5]
    
    items = []
    for doc in all_recent:
        items.append({
            "id": str(doc.get("_id")),
            "lecturer_name": doc.get("lecturer_name", ""),
            "type": doc.get("type", "Permission"),
            "from_date": doc.get("from_date", ""),
            "to_date": doc.get("to_date", ""),
            "status": doc.get("status", ""),
            "mode": doc.get("mode", "full")
        })
    return jsonify(items)

@app.route('/admin/staff')
@login_required
@admin_required
def manage_staff():
    # Always show lecturers sorted by Staff ID (BBHCF001, BBHCF002, ...)
    all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
    
    # Get unique departments from all lecturers
    departments = sorted(list(set(s.get('department', 'N/A') for s in all_staff if s.get('department'))))
    
    # Get current HOD assignments and their signature status
    hod_assignments = {}
    hod_assignments_sig = {}
    for h in department_hods.find():
        dept = h['department']
        hod_id = h['hod_id']
        hod_assignments[dept] = hod_id
        hod_user = users.find_one({"_id": ObjectId(hod_id)})
        if hod_user and hod_user.get('signature_path'):
            hod_assignments_sig[dept] = True
    
    return render_template('admin/manage_staff.html', staff=all_staff, departments=departments, hod_assignments=hod_assignments, hod_assignments_sig=hod_assignments_sig)

@app.route('/admin/staff/new', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_staff_new():
    error = None
    form = {
        "staff_id": "",
        "name": "",
        "designation": "",
        "department": "",
        "category": "Teaching Faculty",
        "email": "",
        "phone": "",
        "username": "",
    }

    if request.method == 'POST':
        staff_id = (request.form.get('staff_id') or '').strip()
        name = (request.form.get('name') or '').strip()
        designation = (request.form.get('designation') or '').strip()
        department = (request.form.get('department') or '').strip()
        category = (request.form.get('category') or '').strip()
        email = (request.form.get('email') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''

        # Auto-set username = staff_id if not provided
        if not username and staff_id:
            username = staff_id.lower()

        # Default password = "123456" if not provided
        if not password:
            password = "123456"

        form.update(
            staff_id=staff_id,
            name=name,
            designation=designation,
            department=department,
            category=category,
            email=email,
            phone=phone,
            username=username,
        )

        if not staff_id or not name or not designation or not department or not category:
            error = "Please fill all required fields."
        elif users.find_one({"staff_id": staff_id}):
            error = "This Staff ID already exists."
        elif username and users.find_one({"username": username}):
            error = "This username already exists."

        if not error:
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            users.insert_one({
                "role": "lecturer",
                "staff_id": staff_id,
                "name": name,
                "designation": designation,
                "department": department,
                "category": category,
                "email": email,
                "phone": phone,
                "username": username,
                "password": password_hash,
                "display_password": password,  # Store for admin display
                "created_date": datetime.now(),  # Store creation date
                "assigned_subjects": "",  # Initialize assigned subjects
            })
            return redirect(url_for('manage_staff'))

    return render_template('admin/staff_form.html', mode="create", form=form, error=error)

@app.route('/admin/staff/<id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_staff_edit(id):
    error = None
    staff_doc = users.find_one({"_id": ObjectId(id), "role": "lecturer"})
    if not staff_doc:
        return redirect(url_for('manage_staff'))

    form = {
        "staff_id": staff_doc.get("staff_id", ""),
        "name": staff_doc.get("name", ""),
        "designation": staff_doc.get("designation", ""),
        "department": staff_doc.get("department", ""),
        "category": staff_doc.get("category", "Teaching Faculty"),
        "email": staff_doc.get("email", ""),
        "phone": staff_doc.get("phone", ""),
        "username": staff_doc.get("username", ""),
        "display_password": staff_doc.get("display_password", ""),
    }

    if request.method == 'POST':
        staff_id = (request.form.get('staff_id') or '').strip()
        name = (request.form.get('name') or '').strip()
        designation = (request.form.get('designation') or '').strip()
        department = (request.form.get('department') or '').strip()
        category = (request.form.get('category') or '').strip()
        email = (request.form.get('email') or '').strip()
        phone = (request.form.get('phone') or '').strip()
        username = (request.form.get('username') or '').strip()
        new_password = request.form.get('password') or ''

        # Auto-set username = staff_id.lower() if not provided
        if not username and staff_id:
            username = staff_id.lower()

        form.update(
            staff_id=staff_id,
            name=name,
            designation=designation,
            department=department,
            category=category,
            email=email,
            phone=phone,
            username=username,
        )

        if not staff_id or not name or not designation or not department or not category:
            error = "Please fill all required fields."
        else:
            existing_staff_id = users.find_one({"staff_id": staff_id, "_id": {"$ne": staff_doc["_id"]}})
            if existing_staff_id:
                error = "This Staff ID already exists."
            else:
                existing_username = users.find_one({"username": username, "_id": {"$ne": staff_doc["_id"]}})
                if existing_username:
                    error = "This username already exists."

        if not error:
            update = {
                "staff_id": staff_id,
                "name": name,
                "designation": designation,
                "department": department,
                "category": category,
                "email": email,
                "phone": phone,
                "username": username,
            }
            if new_password.strip():
                update["password"] = bcrypt.generate_password_hash(new_password).decode('utf-8')
                update["display_password"] = new_password  # Store for admin display

            users.update_one({"_id": staff_doc["_id"]}, {"$set": update})
            return redirect(url_for('manage_staff'))

    return render_template('admin/staff_form.html', mode="edit", form=form, error=error, staff_id=str(staff_doc["_id"]))

@app.route('/admin/staff/<id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_staff_delete(id):
    users.delete_one({"_id": ObjectId(id), "role": "lecturer"})
    return redirect(url_for('manage_staff'))

@app.route('/admin/broadcast', methods=['POST'])
@login_required
@admin_required
def admin_broadcast():
    message = request.form.get('message', '').strip()
    if not message:
        return jsonify({"success": False, "message": "Message is required"}), 400
    
    image_url = None
    if 'image' in request.files:
        file = request.files['image']
        if file and file.filename != '':
            filename = f"broadcast_{int(datetime.now().timestamp())}.png"
            save_dir = os.path.join(os.getcwd(), 'static', 'img', 'broadcasts')
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)
            file.save(os.path.join(save_dir, filename))
            image_url = f"/static/img/broadcasts/{filename}"
    
    broadcast_notifications.insert_one({
        "message": message,
        "image_url": image_url,
        "created_at": datetime.now(),
        "sender": current_user.name
    })
    
    socketio.emit('new_broadcast', {
        "message": message,
        "image_url": image_url,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sender": current_user.name
    })
    
    return jsonify({"success": True, "message": "Broadcast sent successfully"})

@app.route('/api/broadcasts', methods=['GET'])
@login_required
def get_notifications():
    notifs = list(broadcast_notifications.find().sort("created_at", -1).limit(20))
    for n in notifs:
        n['_id'] = str(n['_id'])
        n['created_at'] = n['created_at'].strftime("%Y-%m-%d %H:%M:%S")
    return jsonify(notifs)

@app.route('/admin/broadcast/delete/<id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_broadcast(id):
    broadcast = broadcast_notifications.find_one({"_id": ObjectId(id)})
    if broadcast and broadcast.get('image_url'):
        # Extract relative path from URL (e.g. /static/img/broadcasts/xxx.png)
        img_path = broadcast['image_url'].lstrip('/')
        full_path = os.path.join(app.root_path, img_path)
        if os.path.exists(full_path):
            try:
                os.remove(full_path)
            except Exception as e:
                print(f"Error deleting file {full_path}: {e}")
                
    broadcast_notifications.delete_one({"_id": ObjectId(id)})
    return jsonify({"success": True, "message": "Broadcast deleted successfully"})
@login_required
@admin_required
def admin_upload_signature(id):
    if 'signature' not in request.files:
        return jsonify({"success": False, "message": "No file part"}), 400
    
    file = request.files['signature']
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400
    
    # Simple validation: must be an image
    if not file.content_type.startswith('image/'):
        return jsonify({"success": False, "message": "File must be an image"}), 400
        
    filename = f"sig_{id}.png"
    # Ensure directory exists (though we already created it)
    save_dir = os.path.join(os.getcwd(), 'static', 'img', 'signatures')
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
        
    save_path = os.path.join(save_dir, filename)
    file.save(save_path)
    
    users.update_one(
        {"_id": ObjectId(id)},
        {"$set": {"signature_path": f"img/signatures/{filename}"}}
    )
    return jsonify({"success": True, "path": f"img/signatures/{filename}"})


@app.route('/admin/staff/delete-all', methods=['POST'])
@login_required
@admin_required
def admin_staff_delete_all():
    result = users.delete_many({"role": "lecturer"})
    count = result.deleted_count
    return redirect(url_for('manage_staff', delete_all_success=count))

@app.route('/admin/staff/<id>/change-password', methods=['POST'])
@login_required
@admin_required
def admin_staff_change_password(id):
    staff_doc = users.find_one({"_id": ObjectId(id), "role": "lecturer"})
    if not staff_doc:
        return redirect(url_for('manage_staff'))
    
    new_password = request.form.get('new_password', '').strip()
    if not new_password:
        return redirect(url_for('manage_staff', password_error="Password cannot be empty."))
    
    password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
    # Store display_password for admin view (not secure but for display purposes)
    users.update_one({"_id": ObjectId(id)}, {"$set": {
        "password": password_hash,
        "display_password": new_password  # Store plain text for display only
    }})
    return redirect(url_for('manage_staff', password_updated='1', updated_name=staff_doc.get('name', 'lecturer')))

@app.route('/admin/staff/export-excel')
@login_required
@admin_required
def admin_staff_export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturers"
        
        # Headers - only essential adding information
        headers = ["Lecturer ID", "Name", "Username", "Password"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data rows - only essential information
        for staff in all_staff:
            lecturer_id = staff.get('staff_id', '')
            name = staff.get('name', '')
            username = staff.get('username', staff.get('staff_id', '').lower())
            password = staff.get('display_password', '123456')
            
            ws.append([
                lecturer_id,
                name,
                username,
                password
            ])
        
        # Auto-adjust column widths
        column_widths = {
            'A': 15,  # Lecturer ID
            'B': 35,  # Name
            'C': 15,  # Username
            'D': 15   # Password
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Create in-memory file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        filename = f"Lecturer_Management_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        flash('openpyxl library not installed. Please install it: pip install openpyxl', 'error')
        return redirect(url_for('manage_staff'))
    except Exception as e:
        flash(f'Error exporting Excel: {str(e)}', 'error')
        return redirect(url_for('manage_staff'))

@app.route('/admin/staff/bulk-upload', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_staff_bulk_upload():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return render_template('admin/bulk_upload.html', error="No file selected.")
        
        file = request.files['excel_file']
        if file.filename == '':
            return render_template('admin/bulk_upload.html', error="No file selected.")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return render_template('admin/bulk_upload.html', error="Please upload a valid Excel file (.xlsx or .xls).")
        
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            workbook = load_workbook(file)
            sheet = workbook.active
            
            # Expected columns: Staff ID, Name, Designation, Department, Category, Email, Username, Password
            headers = [cell.value for cell in sheet[1]]
            
            # Find column indices (accept Staff ID, Name/Faculty Name/Faculty for display name)
            col_map = {}
            for idx, header in enumerate(headers, start=1):
                header_lower = str(header).lower().strip() if header else ""
                if 'staff' in header_lower and 'id' in header_lower:
                    col_map['staff_id'] = idx
                elif header_lower in ('faculty name', 'name') or header_lower == 'faculty':
                    col_map['name'] = idx  # Faculty Name / Name for full display (e.g. Mr.Umesh)
                elif 'name' in header_lower and 'user' not in header_lower:
                    col_map['name'] = idx  # e.g. Staff Name, but not Username
                elif 'designation' in header_lower:
                    col_map['designation'] = idx
                elif 'department' in header_lower:
                    col_map['department'] = idx
                elif 'category' in header_lower:
                    col_map['category'] = idx
                elif 'email' in header_lower:
                    col_map['email'] = idx
                elif 'username' in header_lower:
                    col_map['username'] = idx
                elif 'password' in header_lower:
                    col_map['password'] = idx
            
            if 'staff_id' not in col_map or 'name' not in col_map:
                return render_template('admin/bulk_upload.html', error="Excel file must have 'Staff ID' and 'Name' (or 'Faculty Name') columns.")
            
            success_count = 0
            error_rows = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    staff_id = str(row[col_map['staff_id'] - 1].value or '').strip()
                    name = str(row[col_map['name'] - 1].value or '').strip()
                    
                    if not staff_id or not name:
                        continue
                    
                    designation = str(row[col_map.get('designation', 0) - 1].value or '').strip() if col_map.get('designation') else ''
                    department = str(row[col_map.get('department', 0) - 1].value or '').strip() if col_map.get('department') else ''
                    category = str(row[col_map.get('category', 0) - 1].value or '').strip() if col_map.get('category') else 'Teaching Faculty'
                    email = str(row[col_map.get('email', 0) - 1].value or '').strip() if col_map.get('email') else ''
                    username = str(row[col_map.get('username', 0) - 1].value or '').strip() if col_map.get('username') else staff_id.lower()
                    password = str(row[col_map.get('password', 0) - 1].value or '').strip() if col_map.get('password') else '123456'
                    
                    # Check if staff_id already exists
                    if users.find_one({"staff_id": staff_id}):
                        error_rows.append(f"Row {row_num}: Staff ID {staff_id} already exists")
                        continue
                    
                    # Check if username already exists
                    if username and users.find_one({"username": username}):
                        error_rows.append(f"Row {row_num}: Username {username} already exists")
                        continue
                    
                    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                    
                    users.insert_one({
                        "role": "lecturer",
                        "staff_id": staff_id,
                        "name": name,
                        "designation": designation,
                        "department": department,
                        "category": category,
                        "email": email,
                        "username": username,
                        "password": password_hash,
                        "display_password": password,  # Store for admin display
                        "created_date": datetime.now(),  # Store creation date
                        "assigned_subjects": "",  # Initialize assigned subjects
                    })
                    success_count += 1
                except Exception as e:
                    error_rows.append(f"Row {row_num}: {str(e)}")
            
            message = f"Successfully imported {success_count} record(s)."
            if error_rows:
                message += f" {len(error_rows)} error(s) occurred."
            
            # Redirect back to staff list with status so refresh is safe (no re-upload)
            return redirect(url_for('manage_staff', bulk_success=message))
        except ImportError:
            return render_template('admin/bulk_upload.html', error="openpyxl library not installed. Please install it: pip install openpyxl")
        except Exception as e:
            return render_template('admin/bulk_upload.html', error=f"Error processing file: {str(e)}")
    
    return render_template('admin/bulk_upload.html')

@app.route('/api/timetable/<staff_id>')
@login_required
def get_timetable_metadata(staff_id):
    """Fetch timetable structured data and image URL for a specific staff member."""
    # Find the user by staff_id
    u = users.find_one({"staff_id": {"$regex": f"^{staff_id}$", "$options": "i"}})
    if not u:
        return jsonify({"error": "User not found"}), 404
        
    tt_doc = timetable.find_one({"lecturer_id": str(u["_id"])})
    image_url = None
    if tt_doc and tt_doc.get("image_path"):
        image_path = (tt_doc.get("image_path") or "").replace("\\", "/")
        image_url = url_for("static", filename=image_path)
    
    # Also fallback to check if JSON exists on disk if not in tt_doc
    structured = tt_doc.get("structured") if tt_doc else {}
    if not structured:
        json_path = os.path.join(os.path.dirname(__file__), "static", "json_timetables", f"{staff_id}.json")
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    structured = json.load(f)
            except:
                pass
                
    return jsonify({
        "staff_id": staff_id,
        "name": u.get("name"),
        "image_url": image_url,
        "structured": structured
    })


def _iter_timetable_slots(structured):
    """Yield slot dicts from a structured timetable document."""
    if not isinstance(structured, dict):
        return
    tt = structured.get("timetable") or {}
    for day in tt.get("days") or []:
        slots = day.get("slots") or {}
        if isinstance(slots, dict):
            for slot in slots.values():
                if isinstance(slot, dict):
                    yield slot


def _norm_slot_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in ("null", "none", "-"):
        return ""
    return text


def _collect_timetable_catalog():
    """Aggregate class, section, and subject options from all faculty timetables."""
    classes = set()
    sections = set()
    class_sections = set()
    subjects = set()
    class_subject_map = {}

    def _map_key(cls, sec):
        return f"{cls}|{sec or ''}"

    def _add_class_subject(cls, sec, sub):
        for key in (_map_key(cls, sec), _map_key(cls, "")):
            bucket = class_subject_map.setdefault(key, set())
            bucket.add(sub)

    def absorb_slot(slot):
        cls = _norm_slot_text(slot.get("class"))
        sec = _norm_slot_text(slot.get("section"))
        sub = _norm_slot_text(slot.get("subject"))
        if cls:
            classes.add(cls)
            class_sections.add((cls, sec))
            if sec:
                sections.add(sec)
        if sub:
            subjects.add(sub)
        if cls and sub:
            _add_class_subject(cls, sec, sub)

    for doc in timetable.find({}, {"structured": 1}):
        for slot in _iter_timetable_slots(doc.get("structured")):
            absorb_slot(slot)

    json_dir = os.path.join(os.path.dirname(__file__), "static", "json_timetables")
    if os.path.isdir(json_dir):
        for fname in os.listdir(json_dir):
            if not fname.lower().endswith(".json"):
                continue
            try:
                with open(os.path.join(json_dir, fname), "r", encoding="utf-8") as f:
                    structured = json.load(f)
                for slot in _iter_timetable_slots(structured):
                    absorb_slot(slot)
            except Exception:
                continue

    default_classes = [
        "I BCA", "II BCA", "III BCA",
        "I BBA", "II BBA", "III BBA",
        "I BCOM", "II BCOM", "III BCOM",
    ]
    default_sections = ["A", "B", "C"]
    for cls in default_classes:
        classes.add(cls)
    for sec in default_sections:
        sections.add(sec)

    return {
        "classes": sorted(classes, key=lambda x: x.upper()),
        "sections": sorted(sections, key=lambda x: x.upper()),
        "class_sections": [
            {"class": c, "section": s}
            for c, s in sorted(class_sections, key=lambda x: (x[0].upper(), x[1].upper()))
        ],
        "subjects": sorted(subjects, key=lambda x: x.upper()),
        "class_subject_map": {
            key: sorted(subs, key=lambda x: x.upper())
            for key, subs in class_subject_map.items()
        },
    }


@app.route('/api/timetable/catalog')
@login_required
def timetable_catalog():
    """Classes, sections, and subjects used across management timetables."""
    return jsonify(_collect_timetable_catalog())


def _persist_timetable_structured(lecturer_id, lecturer_name, staff_id, data):
    """Save structured timetable to MongoDB and static/json_timetables/{staff_id}.json."""
    timetable.update_one(
        {"lecturer_id": lecturer_id},
        {
            "$set": {
                "lecturer_id": lecturer_id,
                "lecturer_name": lecturer_name,
                "structured": data,
                "updated_at": datetime.now(),
            }
        },
        upsert=True,
    )
    json_dir = os.path.join(os.path.dirname(__file__), "static", "json_timetables")
    os.makedirs(json_dir, exist_ok=True)
    json_path = os.path.join(json_dir, f"{staff_id}.json")
    with open(json_path, "w", encoding="utf-8") as f_json:
        json.dump(data, f_json, indent=4, ensure_ascii=False)


@app.route('/api/timetable/<staff_id>/save', methods=['POST'])
@login_required
def save_timetable_structured(staff_id):
    """Save timetable grid edits (admin for any faculty, lecturer for own)."""
    staff_id = (staff_id or "").strip().upper()
    u = users.find_one({"staff_id": {"$regex": f"^{staff_id}$", "$options": "i"}})
    if not u:
        return jsonify({"success": False, "error": "User not found"}), 404

    lecturer_id = str(u["_id"])
    if current_user.role == 'admin':
        pass
    elif current_user.role == 'lecturer':
        if current_user.id != lecturer_id:
            return jsonify({"success": False, "error": "Forbidden"}), 403
    else:
        return jsonify({"success": False, "error": "Forbidden"}), 403

    raw = (request.form.get("structured_json") or "").strip()
    if not raw and request.is_json:
        body = request.get_json(silent=True) or {}
        if isinstance(body.get("structured"), dict):
            raw = json.dumps(body["structured"])
    if not raw:
        return jsonify({"success": False, "error": "Timetable data cannot be empty."}), 400

    try:
        data = json.loads(raw)
        if not isinstance(data, dict):
            raise ValueError("Data must be a JSON object.")
    except Exception as exc:
        return jsonify({"success": False, "error": f"Invalid data format: {exc}"}), 400

    try:
        _persist_timetable_structured(
            lecturer_id,
            u.get("name") or "",
            staff_id,
            data,
        )
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500

    socketio.emit(
        'timetable_updated',
        {'staff_id': staff_id},
        room=f'user_{lecturer_id}',
    )
    return jsonify({"success": True})


def get_leave_types():
    types = list(db.leave_types.find().sort("name", 1))
    return types

@app.route('/admin/leaves/api/types', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_leave_types():
    if request.method == 'POST':
        name = request.json.get("name", "").strip()
        days = float(request.json.get("default_days", 10))
        if name:
            db.leave_types.insert_one({"name": name, "default_days": days})
        return jsonify({"success": True})
    
    return jsonify({"types": [{"id": str(t["_id"]), "name": t["name"], "default_days": t["default_days"]} for t in get_leave_types()]})

@app.route('/admin/leaves/api/types/<id>', methods=['DELETE'])
@login_required
@admin_required
def delete_leave_type(id):
    db.leave_types.delete_one({"_id": ObjectId(id)})
    return jsonify({"success": True})


@app.route('/admin/leaves')
@login_required
@admin_required
def admin_leaves():
    # Optional filters: search query and month (YYYY-MM)
    q = (request.args.get("q") or "").strip()
    month = (request.args.get("month") or "").strip()

    # Fetch both standard leaves and time-based permissions
    all_leaves_coll = list(leaves.find())
    all_permissions_coll = list(permissions.find())
    
    # Merge and sort by ID (creation time) descending
    all_leaves = sorted(all_leaves_coll + all_permissions_coll, key=lambda x: x['_id'], reverse=True)

    def matches_filters(doc):
        text_ok = True
        month_ok = True

        if q:
            q_lower = q.lower()
            text_fields = [
                str(doc.get("lecturer_name", "")),
                str(doc.get("type", "")),
                str(doc.get("reason", "")),
                str(doc.get("status", "")),
            ]
            text_ok = any(q_lower in field.lower() for field in text_fields)

        if month:
            # Expect month in format YYYY-MM, match against from_date / to_date strings
            from_date = str(doc.get("from_date", ""))
            to_date = str(doc.get("to_date", ""))
            month_ok = month in from_date or month in to_date

        return text_ok and month_ok

    filtered_leaves = [doc for doc in all_leaves if matches_filters(doc)]

    # For allocations view, show lecturers in the same sorted Staff ID order
    # as the "Add Lecturer / Manage Staff" page (BBHCF001, BBHCF002, ...).
    all_lecturers = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
    
    # Inject detailed leave stats for each lecturer
    for lec in all_lecturers:
        lec['leave_stats'] = get_all_leave_stats(str(lec['_id']))

    return render_template(
        'admin/leave_requests.html',
        leaves=filtered_leaves,
        q=q,
        month=month,
        lecturers=all_lecturers,
        leave_types=get_leave_types(),
    )

@app.route('/admin/all-assignments')
@login_required
@admin_required
def admin_all_assignments():
    """Detailed overview of all class substitution assignments in the system"""
    all_allocs = list(leave_class_allocations.find().sort("created_at", -1))
    return render_template('admin/all_assignments.html', assignments=all_allocs)

@app.route('/admin/api/clear-all-assignments', methods=['POST'])
@login_required
@admin_required
def admin_clear_all_assignments():
    """Wipe all substitution data from the entire system (Assignments and Notifications)"""
    try:
        # 1. Clear all allocations
        leave_class_allocations.delete_many({})
        # 2. Clear all related notifications
        faculty_notifications.delete_many({"type": "class_assignment"})
        
        flash("All assignment data has been cleared from the system.", "success")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/admin/leaves/api/set_allocation/<id>', methods=['POST'])
@login_required
@admin_required
def api_set_leave_allocation(id):
    allocated = request.json.get('leaves_per_month', 1)
    leave_type = request.json.get('leave_type', '').strip()
    if not leave_type:
        return jsonify({"success": False, "message": "Leave type is required"}), 400
        
    try:
        allocated = float(allocated)
    except:
        allocated = 1
        
    # AUTO-REGISTER: If this type doesn't exist globally, add it
    if not leave_types.find_one({"name": leave_type}):
        leave_types.insert_one({"name": leave_type})
        
    user = users.find_one({"_id": ObjectId(id)})
    if user:
        balances = user.get("leave_balances", {})
        balances[leave_type] = allocated
        
        # also set the legacy leaves_per_month if it's the main type so backward compatibility works somewhat
        # and only if it's Casual Leave or the first one
        update_fields = {"leave_balances": balances}
        if leave_type == "Casual Leave" or not user.get("leaves_per_month"):
            update_fields["leaves_per_month"] = allocated
            
        users.update_one({"_id": ObjectId(id)}, {"$set": update_fields})
        
    socketio.emit('leave_types_updated')
    return jsonify({"success": True})
    
@app.route('/admin/leaves/api/get_balances/<id>', methods=['GET'])
@login_required
@admin_required
def api_get_leave_balances(id):
    user = users.find_one({"_id": ObjectId(id)})
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404
        
    stats = get_all_leave_stats(id)
    return jsonify({"success": True, "stats": stats})

@app.route('/admin/leaves/api/delete_allocation/<id>/<type>', methods=['POST'])
@login_required
@admin_required
def api_delete_leave_allocation(id, type):
    user = users.find_one({"_id": ObjectId(id)})
    if user:
        balances = user.get("leave_balances", {})
        if type in balances:
            del balances[type]
            users.update_one({"_id": ObjectId(id)}, {"$set": {"leave_balances": balances}})
            socketio.emit('leave_types_updated')
            return jsonify({"success": True})
    return jsonify({"success": False, "message": "Not found"})

@app.route('/lecturer/api/leave-balance', methods=['GET'])
@login_required
def api_get_lecturer_balance():
    leave_type = request.args.get('type', 'Casual Leave')
    balance = calculate_leaves_left(current_user.id, leave_type)
    
    # Also return the total allocated for this type
    user_doc = users.find_one({"_id": ObjectId(current_user.id)})
    total = 0
    if user_doc:
        if "leave_balances" in user_doc:
            total = user_doc["leave_balances"].get(leave_type, 0)
        else:
            total = user_doc.get("leaves_per_month", 0)
            
    return jsonify({
        "success": True, 
        "balance": balance, 
        "total": total,
        "display": f"{total}/{int(balance) if balance.is_integer() else balance}"
    })

@app.route('/lecturer/api/all-balances', methods=['GET'])
@login_required
def api_get_lecturer_all_balances():
    stats = get_all_leave_stats(current_user.id)
    return jsonify({"success": True, "stats": stats})


@app.route('/admin/leaves/delete-all', methods=['POST'])
@login_required
@admin_required
def admin_leaves_delete_all():
    # 1. Clear both leaves and permissions
    l_res = leaves.delete_many({})
    p_res = permissions.delete_many({})
    
    # 2. Clear all linked data to prevent orphans
    leave_class_allocations.delete_many({})
    # Also clear related notifications
    faculty_notifications.delete_many({"type": {"$in": ["class_assignment", "leave_status"]}})
    
    count = l_res.deleted_count + p_res.deleted_count
    flash(f"Successfully deleted all {count} leave and permission records.", "success")
    return redirect(url_for('admin_leaves'))


@app.route('/admin/leave/delete/<id>', methods=['POST'])
@login_required
@admin_required
def admin_leave_delete(id):
    obj_id = ObjectId(id)
    
    # Try deleting from leaves first
    res = leaves.delete_one({"_id": obj_id})
    
    # If not found in leaves, try permissions
    if res.deleted_count == 0:
        res = permissions.delete_one({"_id": obj_id})
    
    # Cascade delete: clean up allocations and notifications related to this ID
    # Note: leave_id is typically stored as a string in these secondary collections
    leave_class_allocations.delete_many({"leave_id": id})
    faculty_notifications.delete_many({
        "$or": [
            {"leave_id": id},
            {"allocation_id": id}
        ]
    })
    
    return jsonify({"success": True, "deleted": res.deleted_count > 0})


@app.route('/admin/timetables', methods=['GET'])
@login_required
@admin_required
def admin_timetables():
    # Fetch all lecturers
    staff_id_regex = r"^BBHCF\d+$"
    all_lecturers = list(
        users.find(
            {
                "role": "lecturer",
                "staff_id": {"$regex": staff_id_regex, "$options": "i"},
            }
        ).sort("staff_id", 1)
    )
    
    # Path for JSON timetables
    json_dir = os.path.join(os.path.dirname(__file__), "static", "json_timetables")
    os.makedirs(json_dir, exist_ok=True)
    
    # Map staff_id to whether it exists on disk
    existing_files = {f.split('.')[0] for f in os.listdir(json_dir) if f.endswith('.json')}
    
    lecturers_data = []
    uploaded_count = 0
    for lect in all_lecturers:
        staff_id = lect.get("staff_id")
        has_tt = staff_id in existing_files
        if has_tt: uploaded_count += 1
        
        lecturers_data.append({
            "id": str(lect["_id"]),
            "staff_id": staff_id,
            "name": lect.get("name"),
            "department": lect.get("department"),
            "has_timetable": has_tt
        })
    
    return render_template(
        'admin/timetables.html',
        lecturers=lecturers_data,
        total=len(all_lecturers),
        uploaded_count=uploaded_count,
        pending_count=len(all_lecturers) - uploaded_count,
        error=request.args.get('error'),
        message=request.args.get('message')
    )

@app.route('/admin/timetables/upload', methods=['POST'])
@login_required
@admin_required
def admin_timetables_upload():
    file = request.files.get('timetable_pdf')
    if not file or file.filename == '':
        return redirect(url_for('admin_timetables', error="No PDF file selected."))
    
    if not file.filename.lower().endswith('.pdf'):
        return redirect(url_for('admin_timetables', error="Only PDF files are allowed."))

    try:
        pdf_bytes = file.read()
        import threading
        from utils.timetable_processor import process_background_pipeline, stop_events
        
        task_id = "main_worker"
        if task_id in stop_events:
            return redirect(url_for('admin_timetables', error="A bulk process is already running. Please stop it or wait for it to finish."))
            
        # Register the stop event BEFORE starting the thread
        stop_events[task_id] = threading.Event()
        
        threading.Thread(target=process_background_pipeline, args=(
            pdf_bytes, task_id, socketio, db
        )).start()
        
        return redirect(url_for('admin_timetables', live='1', message="PDF upload successful. Processing started..."))
    except Exception as e:
        return redirect(url_for('admin_timetables', error=f"Upload error: {str(e)}"))

@app.route('/admin/timetables/upload-image', methods=['POST'])
@login_required
@admin_required
def admin_timetables_upload_image():
    file = request.files.get('timetable_image')
    if not file or file.filename == '':
        return redirect(url_for('admin_timetables', error="No image file selected."))

    try:
        img_bytes = file.read()
        import threading
        from utils.timetable_processor import extract_from_image, match_and_save, log_event

        def process_image_worker(image_data):
            try:
                log_event("Initializing single image process...", socketio=socketio, progress=10)
                log_event("Sending image to Gemini AI...", socketio=socketio, progress=30)
                data = extract_from_image(image_data)
                if "error" in data:
                    log_event(f"AI Error: {data['error']}", socketio=socketio, status="error", progress=0)
                    return
                
                log_event(f"AI extracted faculty: {data.get('faculty', 'Unknown')}", socketio=socketio, progress=70)
                log_event("Syncing with database...", socketio=socketio, progress=90)
                match_and_save(data, db, socketio)
                log_event(f"Successfully processed: {data.get('faculty', 'Unknown')}", socketio=socketio, progress=100)
                socketio.emit('timetable_progress', {'progress': 100, 'done': True, 'status': "Process complete."})
            except Exception as e:
                log_event(f"Worker Error: {str(e)}", socketio=socketio, status="error")

        threading.Thread(target=process_image_worker, args=(img_bytes,)).start()
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": True, "message": "Processing started"})
        return redirect(url_for('admin_timetables', live='1', message="Image upload successful. Processing started..."))
    except Exception as e:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"success": False, "error": str(e)}), 500
        return redirect(url_for('admin_timetables', error=f"Image upload error: {str(e)}"))

@app.route('/admin/timetables/stop', methods=['POST'])
@login_required
@admin_required
def admin_timetables_stop():
    from utils.timetable_processor import stop_events
    task_id = "main_worker"
    if task_id in stop_events:
        stop_events[task_id].set()
        # Also log it
        from utils.timetable_processor import log_event
        log_event("🛑 Manual stop requested by admin.", socketio=socketio, status="warning")
        return jsonify({"success": True, "message": "Stopping... Please wait for current slice to finish."})
    return jsonify({"success": False, "message": "No active process found."})

@app.route('/admin/timetables/delete/<staff_id>', methods=['POST'])
@login_required
@admin_required
def admin_timetable_delete(staff_id):
    json_path = os.path.join(os.path.dirname(__file__), "static", "json_timetables", f"{staff_id}.json")
    if os.path.exists(json_path):
        os.remove(json_path)
    
    # Also clear from DB
    lecturer = users.find_one({"staff_id": staff_id})
    if lecturer:
        timetable.delete_one({"lecturer_id": str(lecturer["_id"])})
        
    return redirect(url_for('admin_timetables', message=f"Timetable for {staff_id} deleted."))

@app.route('/admin/timetables/bulk-delete', methods=['POST'])
@login_required
@admin_required
def admin_timetables_bulk_delete():
    import shutil
    
    # Directories to clear
    dirs_to_clear = [
        os.path.join(os.path.dirname(__file__), "static", "json_timetables"),
        os.path.join(os.path.dirname(__file__), "static", "timetables_json"),
        os.path.join(os.path.dirname(__file__), "static", "timetable_splits"),
        os.path.join(os.path.dirname(__file__), "static", "timetable_images"),
        os.path.join(os.path.dirname(__file__), "static", "timetables")
    ]
    
    for d in dirs_to_clear:
        if os.path.exists(d):
            shutil.rmtree(d)
        os.makedirs(d, exist_ok=True)
    
    # Also delete tracking file
    track_file = os.path.join(os.path.dirname(__file__), "static", "processed_slices.json")
    if os.path.exists(track_file):
        os.remove(track_file)
    
    # Clear Table from DB
    timetable.delete_many({})
    
    return redirect(url_for('admin_timetables', message="All JSONs, images, tracking files, and database records cleared."))

@app.route('/admin/timetables/edit/<staff_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_timetable_edit(staff_id):
    json_path = os.path.join(os.path.dirname(__file__), "static", "json_timetables", f"{staff_id}.json")
    if not os.path.exists(json_path):
        flash("JSON file not found for this faculty.", "danger")
        return redirect(url_for('admin_timetables'))
    
    # Load JSON
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    json_str = json.dumps(data, indent=4, ensure_ascii=False)
    
    if request.method == 'POST':
        new_json = request.form.get('json_data')
        try:
            data = json.loads(new_json)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            flash("Timetable JSON updated.", "success")
            return redirect(url_for('admin_timetables'))
        except Exception as e:
            flash(f"Invalid JSON: {e}", "danger")
            json_str = new_json

    return render_template('admin/edit_json.html', staff_id=staff_id, json_str=json_str)

@app.route('/admin/leave/<id>/<status>', methods=['GET', 'POST'])
@login_required
@admin_required
def review_leave(id, status):
    if status not in ("Approved", "Rejected", "Pending"):
        flash("Invalid status.", "danger")
        return redirect(request.referrer or url_for('admin_leaves'))

    leave_doc = leaves.find_one({"_id": ObjectId(id)})

    # Update the leave status
    leaves.update_one(
        {"_id": ObjectId(id)},
        {"$set": {"status": status, "reviewed_at": datetime.now()}},
    )
    
    # CASCADE APPROVAL: If Management approves the leave, all allocations become 'approved'
    if status == "Approved":
        leave_class_allocations.update_many(
            {"leave_id": id},
            {"$set": {"status": "approved"}}
        )
        # Automatic Attendance Update for Half Days and Permissions
        if leave_doc.get('half_day') or leave_doc.get('type') == 'Permission' or leave_doc.get('mode') == 'time':
            # Extract times if it's a permission
            t_from = None
            t_to = None
            if leave_doc.get('mode') == 'time':
                try:
                    t_from = leave_doc['from_date'].split(' ')[1]
                    t_to = leave_doc['to_date'].split(' ')[1]
                except: pass
                
            update_attendance_log_on_approval(
                leave_doc['lecturer_id'], 
                leave_doc['from_date'], 
                is_half_day=leave_doc.get('half_day', False), 
                session=leave_doc.get('session'),
                is_permission=(leave_doc.get('type') == 'Permission' or leave_doc.get('mode') == 'time'),
                time_from=t_from,
                time_to=t_to
            )
    
    # CASCADE REJECTION: If Rejected, delete everything
    elif status == "Rejected":
        # Get all allocation IDs first to ensure notification cleanup is absolute
        allocations = list(leave_class_allocations.find({"leave_id": id}))
        alloc_ids = [str(a['_id']) for a in allocations]
        
        # 1. Delete all allocations linked to this leave
        leave_class_allocations.delete_many({"leave_id": id})
        # 2. Delete all related notifications (by leave_id or allocation_id)
        faculty_notifications.delete_many({
            "$or": [
                {"leave_id": id},
                {"allocation_id": {"$in": alloc_ids}}
            ]
        })
        # 3. Notify colleagues via socket so their screens refresh instantly
        for a in allocations:
            socketio.emit('assignment_recalled', {"allocation_id": str(a['_id'])}, room=None)
        
        # 4. CLEAR DRAFT: So they start fresh for the next attempt
        if leave_doc:
            leave_drafts.delete_one({"user_id": str(leave_doc['lecturer_id'])})
    
    if leave_doc:
        leaves_left = calculate_leaves_left(leave_doc['lecturer_id'])
        socketio.emit('leave_status_update', {
            'id': id,
            'status': status,
            'lecturer_id': leave_doc['lecturer_id'],
            'leaves_left': leaves_left
        })
        
    flash(f"Leave {status.lower()} successfully!", "success")
    return redirect(url_for('admin_leaves'))

def get_all_leave_stats(lecturer_id):
    user_doc = users.find_one({"_id": ObjectId(lecturer_id)})
    if not user_doc: return []
    
    balances = user_doc.get("leave_balances", {})
    # Legacy fallback: if NO leave_balances object exists at all, use leaves_per_month for Casual
    if not balances and "leaves_per_month" in user_doc:
        balances = {"Casual Leave": user_doc.get("leaves_per_month", 20)}
    
    # Get all approved leaves for this lecturer once
    approved_leaves = list(leaves.find({"lecturer_id": str(lecturer_id), "status": "Approved"}))
    
    # Get all approved permissions for this lecturer
    approved_permissions = list(permissions.find({"lecturer_id": str(lecturer_id), "status": "Approved"}))
    
    stats = []
    # Fetch standard types from dynamic collection - THIS IS THE SOURCE OF TRUTH
    db_types = [t['name'] for t in leave_types.find().sort("name", 1)]
    
    # We ONLY show types that are in the global leave_types collection.
    # If a type was deleted globally, it will no longer show up here
    # even if it still exists in the user's document (this handles orphaned data nicely).
    all_types = db_types
    
    for lt in all_types:
        # If the type is not in balances, it's 0.0 (unassigned)
        total = float(balances.get(lt, 0))
        used = 0.0
        
        type_leaves = [l for l in approved_leaves if l.get('type') == lt]
        for l in type_leaves:
            # logic from calculate_leaves_left
            if l.get('half_day'): 
                used += 0.5
            elif l.get('mode') == 'time': 
                used += 1.0 # Assuming time-based leaves count as 1 day for now
            else:
                try:
                    days = count_working_leave_days(l['from_date'], l['to_date'])
                    if days > 0:
                        used += float(days)
                except Exception:
                    used += 1.0
        
        stats.append({
            "type": lt,
            "total": total,
            "used": used,
            "left": max(0.0, total - used)
        })
    
    # Add a special entry for Permission Count
    stats.append({
        "type": "Permission",
        "used": len(approved_permissions),
        "is_count_only": True
    })
    
    return stats

@app.route('/admin/leave/api/<id>/<status>', methods=['POST'])
@login_required
@admin_required
def api_review_leave(id, status):
    if status not in ("Approved", "Rejected", "Pending"):
        return jsonify({"success": False, "message": "Invalid status"}), 400

    leave_doc = leaves.find_one({"_id": ObjectId(id)})
    if not leave_doc:
        return jsonify({"success": False, "message": "Not found"}), 404

    leaves.update_one(
        {"_id": ObjectId(id)},
        {"$set": {"status": status, "reviewed_at": datetime.now()}},
    )
    
    # CASCADE APPROVAL: If Management approves the leave, all allocations become 'approved'
    if status == "Approved":
        leave_class_allocations.update_many(
            {"leave_id": id},
            {"$set": {"status": "approved"}}
        )
        # Automatic Attendance Update for Half Days and Permissions
        if leave_doc.get('half_day') or leave_doc.get('type') == 'Permission' or leave_doc.get('mode') == 'time':
            # Extract times if it's a permission
            t_from = None
            t_to = None
            if leave_doc.get('mode') == 'time':
                try:
                    t_from = leave_doc['from_date'].split(' ')[1]
                    t_to = leave_doc['to_date'].split(' ')[1]
                except: pass

            update_attendance_log_on_approval(
                leave_doc['lecturer_id'], 
                leave_doc['from_date'], 
                is_half_day=leave_doc.get('half_day', False), 
                session=leave_doc.get('session'),
                is_permission=(leave_doc.get('type') == 'Permission' or leave_doc.get('mode') == 'time'),
                time_from=t_from,
                time_to=t_to
            )
    
    # CASCADE REJECTION: If Rejected, delete everything
    elif status == "Rejected":
        leave_class_allocations.delete_many({"leave_id": id})
        faculty_notifications.delete_many({"leave_id": id})
        # CLEAR DRAFT: So they start fresh for the next attempt
        leave_drafts.delete_one({"user_id": str(leave_doc['lecturer_id'])})

    leaves_left = calculate_leaves_left(leave_doc['lecturer_id'])
    socketio.emit('leave_status_update', {
        'id': id,
        'status': status,
        'lecturer_id': leave_doc['lecturer_id'],
        'leaves_left': leaves_left
    })
    
    return jsonify({"success": True})

def calculate_leaves_left(lecturer_id, leave_type="Casual Leave"):
    user_doc = users.find_one({"_id": ObjectId(lecturer_id)})
    if not user_doc: return 0
    
    # Calculate across ALL leave types if none specified, or just one
    user_balances = user_doc.get("leave_balances", {})
    if not leave_type:
        # Sum all balances
        total_leaves = sum(float(v) for v in user_balances.values())
        query = {"lecturer_id": lecturer_id, "status": "Approved"}
    else:
        total_leaves = float(user_balances.get(leave_type, 0))
        query = {"lecturer_id": lecturer_id, "status": "Approved", "type": leave_type}
        
    approved_leaves = list(leaves.find(query))
    used_days = 0
    for l in approved_leaves:
        mode = l.get('mode', 'full')
        is_half_day = l.get('half_day', False)
        
        if is_half_day:
            used_days += 0.5
        elif mode == 'time':
            # Permission leave usually doesn't deduct from balance unless specified, 
            # but for "perfect" calculation we follow existing logic or skip it.
            # In this system, Permission Leave is usually separate.
            pass
        else:
            try:
                days = count_working_leave_days(l['from_date'], l['to_date'])
                if days > 0:
                    used_days += days
            except Exception:
                used_days += 1
    return max(0, total_leaves - used_days)

def calculate_lecturer_attendance_stats(staff_id):
    """Calculates attendance percentage based on JSON logs in current month"""
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    if not base_dir or not os.path.isdir(base_dir) or not staff_id:
        return 0
    
    current_month = datetime.now().strftime("%Y-%m")
    present_days = 0
    total_working_days = 0
    
    # Fetch approved leaves/permissions for this staff_id for cross-referencing
    from utils.db import leaves, permissions
    approved_leaves = list(leaves.find({"staff_id": staff_id, "status": "Approved"}))
    approved_permissions = list(permissions.find({"staff_id": staff_id, "status": "Approved"}))
    
    # Also find by lecturer_id if staff_id is an ObjectId string in some docs
    user_doc = users.find_one({"staff_id": staff_id})
    if user_doc:
        uid = str(user_doc['_id'])
        approved_leaves += list(leaves.find({"lecturer_id": uid, "status": "Approved"}))
        approved_permissions += list(permissions.find({"lecturer_id": uid, "status": "Approved"}))

    leave_dates = set()
    for l in approved_leaves:
        try:
            start_dt = datetime.strptime(l['from_date'][:10], "%Y-%m-%d")
            end_dt = datetime.strptime(l['to_date'][:10], "%Y-%m-%d")
            curr = start_dt
            while curr <= end_dt:
                leave_dates.add(curr.strftime("%Y-%m-%d"))
                curr += timedelta(days=1)
        except: pass
    
    permission_dates = {p['date'] for p in approved_permissions if p.get('date')}

    
    try:
        for fname in os.listdir(base_dir):
            if not fname.lower().endswith(".json") or not fname.startswith(current_month):
                continue
            
            total_working_days += 1
            fpath = os.path.join(base_dir, fname)
            try:
                with open(fpath, encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, dict): data = [data]
                    for row in data:
                        if row.get("staff_id") == staff_id:
                            is_on_leave = fname.replace(".json", "") in leave_dates or fname.replace(".json", "") in permission_dates
                            if (row.get("checkin") and row.get("checkout")) or is_on_leave:
                                present_days += 1
                            break
            except: continue
            
        if total_working_days == 0: return 100 # New month, assume 100% until first log
        return round((present_days / total_working_days) * 100)
    except:
        return 0

def calculate_lecturer_monthly_delay(staff_id, user_doc=None, month_str=None):
    """Total delay (H:M:S) for the lecturer in the given month from attendance JSON logs."""
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    if not base_dir or not os.path.isdir(base_dir) or not staff_id:
        return "00:00:00"

    if not month_str:
        month_str = datetime.now().strftime("%Y-%m")

    fid = str(staff_id).strip().upper()
    if fid in EXCLUDED_FACULTY_IDS:
        return "00:00:00"

    if user_doc is None:
        user_doc = users.find_one({"staff_id": staff_id}) or {}

    total_seconds = 0
    try:
        month_files = [
            f for f in os.listdir(base_dir)
            if f.startswith(month_str) and f.endswith(".json")
        ]
        for fname in month_files:
            fpath = os.path.join(base_dir, fname)
            date_obj = datetime.strptime(fname.replace(".json", ""), "%Y-%m-%d")
            with open(fpath, encoding="utf-8") as f:
                try:
                    data = json.load(f)
                    if isinstance(data, dict):
                        data = [data]
                except Exception:
                    f.seek(0)
                    data = [json.loads(line) for line in f if line.strip()]

            rows = [
                r for r in data
                if (r.get("staff_id") or r.get("student_id") or "").strip().upper() == fid
            ]
            if rows:
                delay = compute_daily_delay(rows, date_obj, user_doc)
                if delay and delay != "00:00:00" and ":" in str(delay):
                    parts = str(delay).split(":")
                    h, m, s = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0
                    total_seconds += h * 3600 + m * 60 + s
    except Exception as e:
        print(f"Monthly delay calc error for {staff_id}: {e}")

    return format_to_hhmmss(total_seconds)

def update_attendance_log_on_approval(lecturer_id, date_str, is_half_day=False, session=None, is_permission=False, time_from=None, time_to=None):
    """
    Updates the attendance JSON log for a given date by adding a status note
    for Permissions and Half Days. Checkout time is NOT updated automatically anymore.
    """
    try:
        from utils.db import users
        lecturer = users.find_one({"_id": ObjectId(lecturer_id)})
        if not lecturer: return False
        
        staff_id = lecturer.get('staff_id')
        if not staff_id: return False
        
        base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
        if not base_dir or not os.path.isdir(base_dir):
            return False
            
        # Clean date string to YYYY-MM-DD
        date_iso = date_str.split(' ')[0]
        fpath = os.path.join(base_dir, f"{date_iso}.json")
        
        if not os.path.exists(fpath):
            return False
            
        with open(fpath, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
                if isinstance(data, dict): data = [data]
            except:
                return False
                
        # Generate detailed status note
        if is_permission:
            if time_from and time_to:
                status_note = f"Permission ({time_from} to {time_to})"
            else:
                status_note = "Permission Leave"
        elif is_half_day:
            status_note = f"Half Day Leave ({session or 'Unknown'})"
        else:
            status_note = "Approved Leave/Permission"

        updated = False
        for row in data:
            if row.get("staff_id") == staff_id:
                # Update the status note. We don't touch checkout or delay anymore.
                row["status_note"] = status_note
                updated = True
        
        if updated:
            with open(fpath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            return True
            
    except Exception as e:
        print(f"Error updating attendance log: {e}")
    return False

# ============ TIMETABLE & CLASS ALLOCATION HELPERS ============

def get_day_name_from_date(date_str):
    """Get day name (MONDAY, TUESDAY, etc.) from date string"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%A').upper()
    except:
        return None

def load_faculty_timetable(staff_id):
    """Load timetable from JSON file for a faculty member"""
    json_dir = os.path.join(os.path.dirname(__file__), "static", "json_timetables")
    
    # Try the exact staff_id first
    json_path = os.path.join(json_dir, f"{staff_id}.json")
    
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading timetable for {staff_id}: {e}")
    
    # If not found, try alternative formats (with/without 'N')
    alternative_id = None
    if 'N' in staff_id:
        # If ID has N (e.g., BBHCFN028), try without N (BBHCF028)
        alternative_id = staff_id.replace('N', '')
    elif 'BBHCF' in staff_id:
        # If ID doesn't have N (e.g., BBHCF028), try with N (BBHCFN028)
        alternative_id = staff_id.replace('BBHCF', 'BBHCFN')
    
    if alternative_id:
        alt_path = os.path.join(json_dir, f"{alternative_id}.json")
        if os.path.exists(alt_path):
            try:
                with open(alt_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading alternative timetable for {alternative_id}: {e}")
    
    return None

def get_classes_on_date(staff_id, date_str):
    """Total Scan: Retrieve every single hour that has a class assigned (Deep Search)"""
    classes = []
    try:
        # Normalize date format
        date_obj = None
        for fmt in ('%Y-%m-%d', '%d-%m-%Y'):
            try:
                date_obj = datetime.strptime(date_str, fmt)
                break
            except: continue
        if not date_obj: return []
        
        day_name = date_obj.strftime('%A').upper() 
        iso_date = date_obj.strftime('%Y-%m-%d')
        
        # Load Faculty JSON
        data = load_faculty_timetable(staff_id)
        if not data or not isinstance(data, dict): return []
        
        tt = data.get('timetable', {})
        period_meta = {str(p.get('period', '')): p.get('time', 'TBD') for p in tt.get('periods', [])}
        days = tt.get('days', [])
        
        # Find Day (Support MON vs MONDAY)
        day_block = next((d for d in days if d.get('day', '').upper() == day_name or d.get('day', '').upper().startswith(day_name[:3])), None)
        if not day_block: return []
        
        slots = day_block.get('slots', {})
        for p_key, s_data in slots.items():
            # CAPTURE if either subject OR class exists
            if s_data and (s_data.get('subject') or s_data.get('class')):
                subject_name = s_data.get('subject') or s_data.get('class') or "Class Assignment"
                classes.append({
                    'period': p_key,
                    'time': period_meta.get(str(p_key), 'TBD'),
                    'class': s_data.get('class') or "N/A",
                    'section': s_data.get('section', ''),
                    'subject': subject_name,
                    'room': s_data.get('class') or "N/A",
                    'date': iso_date,
                    'day': day_name,
                    'class_id': f"{staff_id}_{iso_date.replace('-', '')}_{p_key}"
                })
        
        # Sort: 0, I, II, III, IV, V, VI, VII
        sort_map = {"0": 0, "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7}
        classes.sort(key=lambda x: sort_map.get(str(x['period']), 99))
        
    except Exception as e:
        print(f"Deep Scan Error: {e}")
    return classes

def get_classes_for_leave_period(staff_id, from_date, to_date):
    """Hard-Inclusive scanner for all days in selected range"""
    all_rows = []
    try:
        def to_dt(s):
            for f in ('%Y-%m-%d', '%d-%m-%Y'):
                try: return datetime.strptime(s, f).date()
                except: continue
            return None
            
        start = to_dt(from_date)
        end = to_dt(to_date)
        if not start or not end: return []
        
        curr = start
        while curr <= end: # The <= ensures the 'To Date' is always included
            if curr.weekday() != 6:  # Sunday not on timetable
                day_list = get_classes_on_date(staff_id, curr.strftime('%Y-%m-%d'))
                all_rows.extend(day_list)
            curr += timedelta(days=1)
    except Exception as e:
        print(f"Range Scanner Critical Error: {e}")
    return all_rows

def count_working_leave_days(from_date_str, to_date_str, half_day=False):
    """Count leave days in range, excluding Sundays (not on college timetable)."""
    if half_day:
        return 0.5

    def to_date(s):
        if not s:
            return None
        s = str(s).split(' ')[0].strip()
        for fmt in ('%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        return None

    start = to_date(from_date_str)
    end = to_date(to_date_str)
    if not start or not end:
        return 1
    if start > end:
        start, end = end, start
    count = 0
    curr = start
    while curr <= end:
        if curr.weekday() != 6:
            count += 1
        curr += timedelta(days=1)
    return count

def clean_t(t):
    """Helper to clean time strings from timetable"""
    t = (t or '').strip().upper()
    if not t: return None
    try:
        if ':' in t:
            h_str, m_str = t.split(':')
            h = int(h_str)
            # Heuristic: 1-7 are PM, 8-12 are AM
            if 1 <= h <= 7: h += 12
            return f"{str(h).zfill(2)}:{m_str.zfill(2)}"
        else:
            # Handle cases where it's just an hour like "9" or "10"
            h = int(t)
            if 1 <= h <= 7: h += 12
            return f"{str(h).zfill(2)}:00"
    except: 
        return None

def get_faculty_duty_bounds(staff_id):
    """Calculate the earliest start and latest end time for a staff member's duty.
    Includes standard college hours plus any timetable-specific classes.
    """
    try:
        user_doc = users.find_one({"staff_id": staff_id}) if staff_id else None
        staff_type = determine_staff_type(user_doc)
        
        from datetime import date
        test_date = date(2026, 5, 4) # A Monday
        deadline_in, threshold_out = get_thresholds_for(test_date, staff_type, staff_id)
        
        final_start = deadline_in.strftime("%H:%M")
        final_end = threshold_out.strftime("%H:%M")
        
        # 2. Expand bounds based on actual timetable periods if present
        try:
            tb_doc = timetables.find_one({"staff_id": staff_id}) if staff_id else None
            if tb_doc and isinstance(tb_doc, dict) and 'periods' in tb_doc:
                for p in tb_doc.get('periods', []):
                    t_range = p.get('time', '')
                    if '-' in t_range:
                        try:
                            s_part, e_part = t_range.split('-', 1)
                            s_cleaned = clean_t(s_part)
                            e_cleaned = clean_t(e_part)
                            
                            if s_cleaned and s_cleaned < final_start:
                                final_start = s_cleaned
                            if e_cleaned and e_cleaned > final_end:
                                final_end = e_cleaned
                        except Exception:
                            pass
        except Exception:
            pass
        
        # 3. Safety fallbacks
        if final_start < "06:00": final_start = "09:00"
        if final_end < final_start: final_end = "16:30"
        
        return final_start, final_end
    except Exception:
        return "09:15", "17:15"

@app.route('/api/faculty/duty-hours/<staff_id>')
@login_required
def api_get_duty_hours(staff_id):
    start, end = get_faculty_duty_bounds(staff_id)
    return jsonify({"success": True, "start": start, "end": end})


def get_available_faculty_for_slot(date_str, time_slot, exclude_staff_id=None):
    """Get faculty members who are free during a specific time slot"""
    all_faculty = list(users.find({"role": "lecturer"}))
    available = []
    
    day_name = get_day_name_from_date(date_str)
    
    for faculty in all_faculty:
        staff_id = faculty.get('staff_id')
        if not staff_id or staff_id == exclude_staff_id:
            continue
        
        # Check if faculty has any class at this time
        faculty_classes = get_classes_on_date(staff_id, date_str)
        has_conflict = False
        
        for cls in faculty_classes:
            if cls.get('time') == time_slot:
                has_conflict = True
                break
        
        if not has_conflict:
            available.append({
                'staff_id': staff_id,
                'name': faculty.get('name'),
                'user_id': str(faculty.get('_id')),
                'designation': faculty.get('designation', 'Lecturer')
            })
    
    return available

def create_class_assignment_notification(leave_id, assigned_to_id, assigned_by_id, class_details):
    """Create notification for faculty about class assignment"""
    notification = {
        'leave_id': leave_id,
        'type': 'class_assignment',
        'assigned_to': assigned_to_id,
        'assigned_by': assigned_by_id,
        'class_details': class_details,
        'status': 'pending',  # pending, accepted, rejected
        'created_at': datetime.now(),
        'read': False
    }
    return faculty_notifications.insert_one(notification)

def save_timetable_backup(staff_id, original_data, reason="leave_assignment"):
    """Save original timetable before making changes"""
    backup = {
        'staff_id': staff_id,
        'original_data': original_data,
        'backup_date': datetime.now(),
        'reason': reason
    }
    return timetable_history.insert_one(backup)

# Lecturer Routes
@app.route('/lecturer/dashboard')
@login_required
@lecturer_required
def lecturer_dashboard():
    # Load recent leaves and permissions
    my_leaves = list(leaves.find({"lecturer_id": current_user.id}).sort("_id", -1).limit(5))
    my_permissions = list(permissions.find({"lecturer_id": current_user.id}).sort("_id", -1).limit(5))
    
    # Combine and sort
    combined = sorted(my_leaves + my_permissions, key=lambda x: x.get('_id'), reverse=True)[:5]

    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    timetable_image_url = None
    has_timetable = False
    if tt_doc and tt_doc.get("image_path"):
        image_path = (tt_doc.get("image_path") or "").replace("\\", "/")
        timetable_image_url = url_for("static", filename=image_path)
        has_timetable = True

    # Calculate real-time stats
    staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
    staff_id = staff_doc.get("staff_id") if staff_doc else None
    
    leaves_left = calculate_leaves_left(current_user.id, leave_type=None) # Sum all types
    attendance_percent = calculate_lecturer_attendance_stats(staff_id)
    month_delay = calculate_lecturer_monthly_delay(staff_id, staff_doc) if staff_id else "00:00:00"
    current_month_label = datetime.now().strftime("%B %Y")
    
    # Get unread notifications count for the badge
    notif_count = faculty_notifications.count_documents({
        "recipient_id": str(current_user.id),
        "status": "unread"
    })
    # Re-calculate profile_locked based on actual file presence
    p_pic = staff_doc.get('profile_pic')
    is_locked = False
    
    if p_pic:
        full_path = os.path.join(app.root_path, 'static', p_pic)
        if os.path.exists(full_path):
            # Only lock if file exists and (it was locked by admin or auto-locked)
            is_locked = True if (staff_doc.get('profile_locked') or p_pic) else False

    return render_template(
        'lecturer/dashboard.html',
        leaves=combined,
        has_timetable=has_timetable,
        timetable_image_url=timetable_image_url,
        leaves_left=leaves_left,
        attendance_percent=attendance_percent,
        month_delay=month_delay,
        current_month_label=current_month_label,
        notif_count=notif_count,
        profile_pic=p_pic,
        profile_locked=is_locked
    )

@app.route('/lecturer/update-email', methods=['POST'])
@lecturer_required
def lecturer_update_email():
    new_email = request.json.get('email')
    if not new_email:
        return jsonify({"success": False, "message": "Email is required"}), 400
    
    users.update_one({"_id": ObjectId(current_user.id)}, {"$set": {"email": new_email}})
    return jsonify({"success": True, "message": "Email updated successfully"})

@app.route('/lecturer/change-password', methods=['POST'])
@lecturer_required
def lecturer_change_password():
    data = request.json
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    
    if not old_password or not new_password:
        return jsonify({"success": False, "message": "All fields are required"}), 400
    
    user_data = users.find_one({"_id": ObjectId(current_user.id)})
    if not bcrypt.check_password_hash(user_data['password'], old_password):
        return jsonify({"success": False, "message": "Incorrect old password"}), 401
    
    hashed_password = bcrypt.generate_password_hash(new_password).decode('utf-8')
    users.update_one({"_id": ObjectId(current_user.id)}, {"$set": {"password": hashed_password, "display_password": new_password}})
    return jsonify({"success": True, "message": "Password updated successfully"})

@app.route('/lecturer/upload-profile-pic', methods=['POST'])
@login_required
@lecturer_required
def upload_profile_pic():
    if 'photo' not in request.files:
        return jsonify({"success": False, "message": "No file"}), 400
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400
        
    if not file.content_type.startswith('image/'):
        return jsonify({"success": False, "message": "File must be an image"}), 400
        
    staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
    staff_id = staff_doc.get('staff_id', 'unknown')
    
    # Simple filename as requested: BBHCF048.png
    filename = f"{staff_id}.png"
    save_dir = os.path.join(app.root_path, 'static', 'img', 'profiles')
    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)
        
    save_path = os.path.join(save_dir, filename)
    file.save(save_path)
    
    relative_path = f"img/profiles/{filename}"
    users.update_one(
        {"_id": ObjectId(current_user.id)},
        {"$set": {"profile_pic": relative_path, "profile_locked": True}} # Auto-lock after upload
    )
    socketio.emit('profile_lock_updated', {"userId": str(current_user.id), "locked": True})
    return jsonify({"success": True, "path": url_for('static', filename=relative_path)})

@app.route('/admin/api/toggle-profile-lock/<id>', methods=['POST'])
@login_required
@admin_required
def toggle_profile_lock(id):
    user = users.find_one({"_id": ObjectId(id)})
    if not user:
        return jsonify({"success": False, "message": "User not found"}), 404
    
    new_status = not user.get('profile_locked', False)
    users.update_one({"_id": ObjectId(id)}, {"$set": {"profile_locked": new_status}})
    return jsonify({"success": True, "locked": new_status})


@app.route('/admin/api-keys', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_api_keys():
    api_keys_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "api_keys.txt")
    if request.method == 'POST':
        new_key = request.form.get('new_key', '').strip()
        if new_key:
            with open(api_keys_path, 'a') as f:
                f.write(f"\n{new_key}")
            flash('API Key added successfully!', 'success')
        return redirect(url_for('admin_timetables'))
        
    keys = []
    if os.path.exists(api_keys_path):
        with open(api_keys_path, 'r') as f:
            keys = [line.strip() for line in f if line.strip() and not line.startswith("#")]
    return jsonify(keys)


@app.route('/admin/view-logs')
@login_required
@admin_required
def admin_view_logs():
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reconstruction_log.txt")
    if os.path.exists(log_path):
        with open(log_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content
    return "No logs found."

@app.route('/admin/clear-logs', methods=['POST'])
@login_required
@admin_required
def admin_clear_logs():
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reconstruction_log.txt")
    try:
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Logs cleared by admin.\n")
        flash('Logs cleared successfully!', 'success')
    except Exception as e:
        flash(f'Error clearing logs: {e}', 'danger')
    return redirect(url_for('admin_timetables'))

def _parse_date_loose(text: str):
    """
    Parse a date from common user formats to YYYY-MM-DD.
    Accepts:
    - YYYY-MM-DD
    - DD-MM-YYYY / D-M-YYYY
    - DD/MM/YYYY / D/M/YYYY
    Returns iso_date string or None.
    """
    raw = (text or "").strip()
    if not raw:
        return None

    # ISO first
    m = re.fullmatch(r"(?P<y>\d{4})-(?P<m>\d{1,2})-(?P<d>\d{1,2})", raw)
    if m:
        try:
            dt = datetime(int(m["y"]), int(m["m"]), int(m["d"]))
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None

    m = re.fullmatch(r"(?P<d>\d{1,2})[\/\-](?P<m>\d{1,2})[\/\-](?P<y>\d{4})", raw)
    if m:
        try:
            dt = datetime(int(m["y"]), int(m["m"]), int(m["d"]))
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None

    return None

def _extract_date_range(message: str):
    """
    Best-effort date range extraction from a message.
    Returns (from_iso, to_iso) or (None, None).
    """
    msg = (message or "").strip()
    if not msg:
        return (None, None)

    # Find all date-like tokens in the message
    tokens = re.findall(r"\b(\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})\b", msg)
    parsed = [_parse_date_loose(t) for t in tokens]
    parsed = [p for p in parsed if p]
    if len(parsed) >= 2:
        return (parsed[0], parsed[1])

    return (None, None)

CHATBOT_INTENT_PHRASES = {
    # Base seed phrases (always available)
    "apply_leave": [
        "apply leave", "apply for leave", "leave apply", "need leave", "want leave", "take leave",
        "i need leave", "i want leave", "i want to apply leave", "i need to apply leave",
        "leave", "apply leave please",
    ],
    "leave_balance": [
        "leave balance", "leaves left", "leave left", "how many leaves", "how many leave", "leave remaining",
        "remaining leave", "my leave balance",
    ],
    "attendance": [
        "attendance", "my attendance", "attendance report", "present days", "absent days",
    ],
    "timetable": [
        "timetable", "my timetable", "time table", "schedule", "my schedule",
    ],
    "dashboard": [
        "dashboard", "home", "main page",
    ],
    "greeting": [
        "hi", "hello", "hey", "good morning", "good evening", "good afternoon",
    ],
}

def _load_excel_phrase_booster():
    """
    Use the Excel dataset (ID, Sentence) to boost phrase matching.
    This is NOT supervised training (no labels in the file). We auto-assign sentences
    into intents by keyword rules and add them as extra phrases.
    """
    try:
        xlsx_path = Path(os.path.dirname(__file__)) / "static" / "chatbot_data_trani" / "chatbot_100k_sentences_dataset.xlsx"
        if not xlsx_path.exists():
            return

        # Allow disabling via env (startup time / memory)
        if (os.getenv("CHATBOT_EXCEL_BOOST", "1") or "1").strip() in {"0", "false", "False", "no", "NO"}:
            return

        import pandas as pd

        df = pd.read_excel(str(xlsx_path))
        if "Sentence" not in df.columns:
            return

        # keyword buckets for auto intent assignment
        buckets = {
            "apply_leave": ["apply leave", "leave apply", "need leave", "want leave", "take leave", "apply for leave"],
            "leave_balance": ["leave balance", "leaves left", "leave remaining", "remaining leave"],
            "attendance": ["attendance", "present", "absent", "checkin", "checkout"],
            "timetable": ["timetable", "time table", "schedule", "class", "period"],
            "dashboard": ["dashboard", "home"],
            "greeting": ["hi", "hello", "hey", "good morning", "good evening", "good afternoon"],
        }

        # cap per bucket to avoid huge memory growth
        cap = int((os.getenv("CHATBOT_EXCEL_BOOST_CAP", "3000") or "3000").strip() or 3000)
        added = {k: 0 for k in buckets.keys()}

        for s in df["Sentence"].astype(str).fillna("").tolist():
            s_norm = s.strip().lower()
            if not s_norm:
                continue
            # Keep phrase length reasonable (avoid huge paragraphs)
            if len(s_norm) > 140:
                continue

            for intent, keys in buckets.items():
                if added[intent] >= cap:
                    continue
                if any(k in s_norm for k in keys):
                    CHATBOT_INTENT_PHRASES.setdefault(intent, []).append(s_norm)
                    added[intent] += 1
                    break
    except Exception:
        # Never fail app startup due to phrase booster
        return

# Load phrase booster once on import/startup
_load_excel_phrase_booster()

def _chatbot_make_reply(text: str, *, actions=None, cards=None, ok: bool = True, message: str | None = None):
    return jsonify({
        "ok": ok,
        "text": text,
        "message": message,
        "actions": actions or [],
        "cards": cards or [],
        "ts": datetime.now().isoformat(),
    })

@app.route('/lecturer/api/chat', methods=['POST'])
def lecturer_chat_api():
    import traceback
    """
    Lightweight lecturer chatbot endpoint (session-based state).
    The UI sends either:
    - { "message": "..." } for normal chat
    - { "action": "...", "payload": {...} } for button clicks
    """
    body = request.get_json(silent=True) or {}
    message = (body.get("message") or "").strip()
    action = (body.get("action") or "").strip()
    payload = body.get("payload") or {}

    # API-friendly auth (do not redirect with HTML)
    if not current_user.is_authenticated:
        return _chatbot_make_reply(
            "Session expired. Please login again.",
            ok=False,
            message="unauthorized",
            actions=[{"type": "navigate", "label": "Login", "url": url_for("login", next=request.path)}],
        ), 401
    if getattr(current_user, "role", None) != "lecturer":
        return _chatbot_make_reply(
            "Lecturer access required.",
            ok=False,
            message="forbidden",
            actions=[{"type": "navigate", "label": "Go Home", "url": url_for("index")}],
        ), 403

    state = session.get("lecturer_chat_state") or {}
    pending = state.get("pending") or {}

    def reset_pending():
        state["pending"] = {}
        session["lecturer_chat_state"] = state

    def set_pending(p):
        state["pending"] = p
        session["lecturer_chat_state"] = state

    # Handle explicit actions from UI buttons
    if action == "navigate":
        url = payload.get("url") or url_for("lecturer_dashboard")
        return _chatbot_make_reply("Opening…", actions=[{"type": "navigate", "url": url}])

    if action == "cancel_flow":
        reset_pending()
        return _chatbot_make_reply("Okay, cancelled. What would you like to do next?")

    # IMPORTANT: handle slot-setting actions before flow continuation
    if action == "set_leave_type":
        if pending.get("intent") != "apply_leave":
            pending = {"intent": "apply_leave"}
        pending["leave_type"] = (payload.get("value") or payload.get("leave_type") or "").strip()
        set_pending(pending)
        return _chatbot_make_reply("Okay. Now tell me the leave dates (example: 03-01-2026 to 06-01-2026)." if (not pending.get("from_date") or not pending.get("to_date")) else "Okay. What’s the reason for your leave?")

    if action == "confirm_leave":
        if pending.get("intent") != "apply_leave":
            reset_pending()
            return _chatbot_make_reply("I don’t have a leave request ready to submit. Try: “Apply leave from 03-01-2026 to 06-01-2026”.")

        leave_type = (pending.get("leave_type") or "").strip()
        from_date = (pending.get("from_date") or "").strip()
        to_date = (pending.get("to_date") or "").strip()
        reason = (pending.get("reason") or "").strip()
        description = (pending.get("description") or "").strip()

        if not (leave_type and from_date and to_date and reason):
            return _chatbot_make_reply("Some fields are missing. Please continue the leave flow.")

        # Validate date ordering
        try:
            fdt = datetime.strptime(from_date, "%Y-%m-%d")
            tdt = datetime.strptime(to_date, "%Y-%m-%d")
            if fdt > tdt:
                return _chatbot_make_reply("From date cannot be after To date. Please re-enter the dates.")
        except Exception:
            return _chatbot_make_reply("I couldn’t validate your dates. Please provide them as DD-MM-YYYY or YYYY-MM-DD.")

        leave_data = {
            "lecturer_id": current_user.id,
            "lecturer_name": current_user.name,
            "type": leave_type,
            "from_date": from_date,
            "to_date": to_date,
            "reason": reason,
            "description": description,
            "status": "Pending",
            "created_at": datetime.now(),
            "mode": "full",
        }
        res = leaves.insert_one(leave_data)
        socketio.emit('new_leave_request', {
            "id": str(res.inserted_id),
            "lecturer_name": current_user.name,
            "type": leave_type,
            "from_date": from_date,
            "to_date": to_date,
            "status": "Pending"
        })

        reset_pending()
        return _chatbot_make_reply(
            f"Submitted. Your leave request is Pending (ID: {str(res.inserted_id)}).",
            actions=[
                {"type": "navigate", "label": "Open Dashboard", "url": url_for("lecturer_dashboard")},
                {"type": "navigate", "label": "Apply another leave", "url": url_for("apply_leave", mode="full")},
            ],
        )

    # If user is in the middle of a flow, continue slot-filling
    if pending.get("intent") == "apply_leave":
        # allow user to restart with new dates
        if message:
            from_d, to_d = _extract_date_range(message)
            if from_d and to_d:
                pending["from_date"] = from_d
                pending["to_date"] = to_d

            if not pending.get("leave_type"):
                # Dynamically match against known types from global list
                active_types = {t['name'].lower(): t['name'] for t in leave_types.find()}
                for k, v in active_types.items():
                    if k in msg_lower:
                        pending["leave_type"] = v
                        break

            elif not pending.get("reason"):
                pending["reason"] = message
            elif not pending.get("description"):
                pending["description"] = message

            set_pending(pending)

        missing = []
        if not pending.get("from_date") or not pending.get("to_date"):
            missing.append("dates")
        if not pending.get("leave_type"):
            missing.append("leave type")
        if not pending.get("reason"):
            missing.append("reason")
        if not pending.get("description"):
            missing.append("description")

        if missing:
            if missing[0] == "dates":
                return _chatbot_make_reply("Tell me the leave dates (example: 03-01-2026 to 06-01-2026).", actions=[{"type": "cancel_flow", "label": "Cancel"}])
            if missing[0] == "leave type":
                dynamic_actions = [
                    {"type": "set_leave_type", "label": name, "value": name} 
                    for name in [t['name'] for t in leave_types.find().sort("name", 1)]
                ]
                dynamic_actions.append({"type": "cancel_flow", "label": "Cancel"})
                
                return _chatbot_make_reply(
                    "Choose a leave type.",
                    actions=dynamic_actions,
                )
            if missing[0] == "reason":
                return _chatbot_make_reply("What’s the reason for your leave?")
            if missing[0] == "description":
                return _chatbot_make_reply("Add a short description (1 line).")

        # All fields present → show confirmation card
        leaves_left = calculate_leaves_left(current_user.id)
        card = {
            "type": "leave_confirm",
            "title": "Confirm leave application",
            "fields": [
                {"label": "From", "value": pending.get("from_date")},
                {"label": "To", "value": pending.get("to_date")},
                {"label": "Type", "value": pending.get("leave_type")},
                {"label": "Reason", "value": pending.get("reason")},
                {"label": "Description", "value": pending.get("description")},
                {"label": "Leaves left (approx.)", "value": str(leaves_left)},
            ],
        }
        return _chatbot_make_reply(
            "Please confirm.",
            cards=[card],
            actions=[
                {"type": "confirm_leave", "label": "Confirm & Submit"},
                {"type": "cancel_flow", "label": "Cancel"},
            ],
        )

    # No active flow: detect intent from message
    msg_lower = message.lower()
    if not message:
        return _chatbot_make_reply(
            "Examples you can type:\n- Apply leave from 03-01-2026 to 06-01-2026\n- My timetable\n- My attendance\n- Leave balance",
            actions=[
                {"type": "navigate", "label": "Dashboard", "url": url_for("lecturer_dashboard")},
                {"type": "navigate", "label": "Timetable", "url": url_for("lecturer_timetable")},
                {"type": "navigate", "label": "Attendance", "url": url_for("lecturer_attendance")},
                {"type": "navigate", "label": "Apply Leave", "url": url_for("apply_leave", mode="full")},
            ],
        )

    # Quick navigation intents
    def _match_any(intent_key: str) -> bool:
        phrases = CHATBOT_INTENT_PHRASES.get(intent_key) or []
        return any(p in msg_lower for p in phrases)

    if "timetable" in msg_lower or _match_any("timetable"):
        return _chatbot_make_reply("Opening your timetable.", actions=[{"type": "navigate", "url": url_for("lecturer_timetable")}])
    if "attendance" in msg_lower or _match_any("attendance"):
        return _chatbot_make_reply("Opening your attendance.", actions=[{"type": "navigate", "url": url_for("lecturer_attendance")}])
    if "dashboard" in msg_lower or "home" in msg_lower or _match_any("dashboard"):
        return _chatbot_make_reply("Opening dashboard.", actions=[{"type": "navigate", "url": url_for("lecturer_dashboard")}])

    # Leaves left
    if ("leave" in msg_lower and ("left" in msg_lower or "balance" in msg_lower)) or _match_any("leave_balance"):
        leaves_left = calculate_leaves_left(current_user.id)
        return _chatbot_make_reply(f"You have {leaves_left} leave day(s) left (based on approved leaves).")

    # Apply leave intent
    is_leave_apply = (
        ("leave" in msg_lower and any(k in msg_lower for k in ("apply", "need", "want", "take")))
        or msg_lower.startswith("leave")
        or _match_any("apply_leave")
        or msg_lower.strip() in {"leave", "apply leave"}
    )
    if is_leave_apply:
        from_d, to_d = _extract_date_range(message)
        p = {"intent": "apply_leave"}
        if from_d and to_d:
            p["from_date"] = from_d
            p["to_date"] = to_d
        set_pending(p)
        if not (from_d and to_d):
            return _chatbot_make_reply(
                "Sure — first tell me the date range (example: 03-01-2026 to 06-01-2026).",
                actions=[
                    {"type": "navigate", "label": "Open Leave Form", "url": url_for("apply_leave", mode="full")},
                    {"type": "cancel_flow", "label": "Cancel"},
                ],
            )
        dynamic_actions = [
            {"type": "set_leave_type", "label": name, "value": name} 
            for name in [t['name'] for t in leave_types.find().sort("name", 1)]
        ]
        dynamic_actions.append({"type": "cancel_flow", "label": "Cancel"})

        return _chatbot_make_reply(
            "Got the dates. Now choose a leave type.",
            actions=dynamic_actions,
        )

    # Default fallback
    return _chatbot_make_reply(
        "I can help with Leave, Timetable, Attendance, and Leave balance. Try: “Apply leave from 03-01-2026 to 06-01-2026”.",
        actions=[
            {"type": "navigate", "label": "Apply Leave", "url": url_for("apply_leave", mode="full")},
            {"type": "navigate", "label": "Timetable", "url": url_for("lecturer_timetable")},
            {"type": "navigate", "label": "Attendance", "url": url_for("lecturer_attendance")},
        ],
    )


# --- Attendance Delay Calculation Helpers ---

def parse_ts(ts_str):
    if not ts_str: return None
    try:
        # Handle formats like 2026-04-25T20:53:03 or 2026-04-25 20:53:03
        return datetime.fromisoformat(ts_str.replace(" ", "T"))
    except:
        return None

def determine_staff_type(user_doc):
    """Classifies staff into categories like 'teaching', 'admin', etc."""
    if not user_doc: return 'teaching'
    category = (user_doc.get('category') or '').strip().lower()
    designation = (user_doc.get('designation') or '').strip().lower()
    
    if 'sanitary worker' in designation: return 'sanitary'
    if 'attender' in designation: return 'attender'
    if 'security guard' in designation: return 'security'
    if 'computer programmer' in designation: return 'programmer'
    
    if category == 'teaching faculty' or 'teaching' in category:
        return 'teaching'
    if 'non-teaching' in category or 'admin' in category:
        return 'admin' 
    return 'teaching'

def get_thresholds_for(date_obj, staff_type, faculty_id=None):
    """Defines the exact Check-in and Check-out deadlines."""
    weekday = date_obj.weekday() # 0=Mon, 5=Sat
    is_saturday = (weekday == 5)
    fid_upper = (faculty_id or '').strip().upper()
    
    if fid_upper == 'BBHCFN020':
        return time(9, 0), (time(13, 30) if is_saturday else time(17, 30))

    if staff_type == 'teaching': # BBHCF
        checkin_deadline = time(9, 25)
        checkout_after = time(13, 0) if is_saturday else time(16, 30)
    elif staff_type == 'admin': # BBHCFN
        checkin_deadline = time(9, 15)
        checkout_after = time(13, 30) if is_saturday else time(17, 15)
    elif staff_type == 'sanitary':
        checkin_deadline = time(8, 45)
        checkout_after = time(17, 15)
    else:
        checkin_deadline = time(9, 30)
        checkout_after = time(16, 30)
        
    return checkin_deadline, checkout_after

def load_json_attendance_file(source):
    """Safely loads attendance JSON files, supporting standard JSON,
    duplicate key JSON objects (which are split into distinct session records),
    and JSONL line-by-line files.
    """
    def parse_with_dup_keys(pairs):
        keys = [k for k, v in pairs]
        if keys.count('checkin') > 1 or keys.count('checkout') > 1:
            records = []
            base_meta = {}
            curr_rec = {}
            for k, v in pairs:
                if k in ('checkin', 'checkout'):
                    if k in curr_rec:
                        records.append(curr_rec)
                        curr_rec = {}
                    curr_rec[k] = v
                else:
                    base_meta[k] = v
            if curr_rec:
                records.append(curr_rec)
            for r in records:
                r.update(base_meta)
            return records
        return dict(pairs)

    raw_data = None
    try:
        if isinstance(source, (str, os.PathLike)) and os.path.exists(str(source)):
            with open(source, 'r', encoding='utf-8') as f:
                content = f.read().strip()
        elif hasattr(source, 'read'):
            content = source.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            content = content.strip()
        elif isinstance(source, str):
            content = source.strip()
        else:
            return []

        if not content:
            return []

        try:
            raw_data = json.loads(content, object_pairs_hook=parse_with_dup_keys)
        except Exception:
            raw_data = []
            for line in content.splitlines():
                line_str = line.strip()
                if line_str:
                    try:
                        parsed = json.loads(line_str, object_pairs_hook=parse_with_dup_keys)
                        if isinstance(parsed, list): raw_data.extend(parsed)
                        elif isinstance(parsed, dict): raw_data.append(parsed)
                    except Exception:
                        pass
    except Exception:
        return []

    flattened = []
    if isinstance(raw_data, list):
        for item in raw_data:
            if isinstance(item, list):
                flattened.extend(item)
            elif isinstance(item, dict):
                flattened.append(item)
        return flattened
    elif isinstance(raw_data, dict):
        return [raw_data]
    return []

def format_to_hhmmss(seconds):
    if seconds <= 0: return "00:00:00"
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{int(h):02}:{int(m):02}:{int(s):02}"

def compute_daily_delay(records, date_obj, user_doc):
    """Calculates exactly how much time was missed during required duty hours,
    subtracting any approved permissions or leaves.
    """
    checkins = [parse_ts(r.get('checkin')) for r in records if r.get('checkin')]
    checkouts = [parse_ts(r.get('checkout')) for r in records if r.get('checkout')]
    
    if not checkins: return 'Absent'
    
    # Extract status note if present
    status_note = ""
    for r in records:
        if r.get('status_note'):
            status_note = r.get('status_note')
            break

    # If no checkout punches exist and no leave/permission note, return 'Absent' for delay
    if not checkouts and not status_note:
        return 'Absent'

    # 1. Define the Required Duty Window for the day
    staff_type = determine_staff_type(user_doc)
    faculty_id = user_doc.get('staff_id') if isinstance(user_doc, dict) else (user_doc if isinstance(user_doc, str) else None)
    deadline_in, threshold_out = get_thresholds_for(date_obj, staff_type, faculty_id)
    
    ref_date = min(checkins).date()
    req_start = datetime.combine(ref_date, deadline_in)
    req_end = datetime.combine(ref_date, threshold_out)
    
    # 2. Collect all "Covered" intervals (Actual Attendance + Permissions/Leaves)
    covered_intervals = []
    
    # Add actual physical attendance
    for r in records:
        ci = parse_ts(r.get('checkin'))
        co = parse_ts(r.get('checkout'))
        if ci and co:
            covered_intervals.append((ci, co))
    
    # Add Permissions/Leaves from status_note
    status_note = ""
    for r in records:
        if r.get('status_note'):
            status_note = r.get('status_note')
            break
            
    if status_note:
        if "Half Day Leave (Morning)" in status_note:
            # Exempt until 12:30 PM
            covered_intervals.append((datetime.combine(ref_date, time(0,0)), datetime.combine(ref_date, time(12,30))))
        elif "Half Day Leave (Afternoon)" in status_note:
            # Exempt from 12:30 PM onwards
            covered_intervals.append((datetime.combine(ref_date, time(12,30)), datetime.combine(ref_date, time(23,59))))
        elif "Permission" in status_note:
            import re
            match = re.search(r"\((\d{1,2}:\d{2})\s+to\s+(\d{1,2}:\d{2})\)", status_note)
            if match:
                try:
                    p_s_str, p_e_str = match.groups()
                    p_start_t = datetime.strptime(p_s_str, "%H:%M").time()
                    p_end_t = datetime.strptime(p_e_str, "%H:%M").time()
                    covered_intervals.append((datetime.combine(ref_date, p_start_t), datetime.combine(ref_date, p_end_t)))
                except: pass

    # 3. Merge overlapping intervals to get total covered time
    covered_intervals.sort()
    merged = []
    if covered_intervals:
        curr_start, curr_end = covered_intervals[0]
        for next_start, next_end in covered_intervals[1:]:
            if next_start <= curr_end:
                curr_end = max(curr_end, next_end)
            else:
                merged.append((curr_start, curr_end))
                curr_start, curr_end = next_start, next_end
        merged.append((curr_start, curr_end))
    
    # 4. Calculate Gaps within the Required Duty Window
    total_delay_seconds = 0
    last_pos = req_start
    
    for m_start, m_end in merged:
        if m_start > req_end: break
        if m_end < req_start: continue
        
        # Any gap between last_pos and the start of this covered interval is a delay
        actual_covered_start = max(req_start, m_start)
        if actual_covered_start > last_pos:
            total_delay_seconds += int((actual_covered_start - last_pos).total_seconds())
        
        # Advance last_pos to the end of this covered interval
        last_pos = max(last_pos, m_end)
        
    # Final gap at the end of the day
    if last_pos < req_end:
        total_delay_seconds += int((req_end - last_pos).total_seconds())
        
    return format_to_hhmmss(total_delay_seconds)

# --- Attendance Report Constants & PDF Helpers ---
EXCLUDED_FACULTY_IDS = {'BBHCF010', 'BBHCF017', 'BBHCF018', 'BBHCF044', 'BBHCFN029'}

def create_pdf_header():
    """Identical PDF Header Design from the Report Tool"""
    styles = getSampleStyleSheet()
    header_title = ParagraphStyle('HeaderTitle', parent=styles['Title'], alignment=0, fontSize=16, leading=19)
    header_sub = ParagraphStyle('HeaderSub', parent=styles['Normal'], alignment=0, fontSize=10, leading=12)
    
    # Logo setup - adjusted for main project path
    try:
        logo_path = os.path.join(app.root_path, "static", "img", "logo-removebg-preview.png")
        if os.path.exists(logo_path):
            logo_img = Image(logo_path)
            logo_img._restrictSize(26*mm, 26*mm)
        else:
            logo_img = ''
    except Exception:
        logo_img = ''
    
    header_text = [
        Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
        Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
    ]
    
    header_table = Table([[logo_img, header_text]], colWidths=[26*mm, (A4[0] - (18*mm + 18*mm) - 26*mm)])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    return header_table

@app.route('/lecturer/attendance')
@login_required
@lecturer_required
def lecturer_attendance():
    """
    Attendance view for the logged-in lecturer.
    Reads JSON attendance files from ATTENDANCE_DIR and filters by this lecturer's staff ID.
    """
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    from datetime import datetime

    # Filters
    selected_month = (request.args.get("month") or "").strip()
    search_q = (request.args.get("q") or "").strip().lower()

    if not selected_month:
        selected_month = datetime.now().strftime("%Y-%m")

    # Find staff_id for current lecturer
    staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
    staff_id = staff_doc.get("staff_id") if staff_doc else None

    records = []
    debug_info = {
        "base_dir": base_dir,
        "dir_exists": os.path.isdir(base_dir) if base_dir else False,
        "staff_id": staff_id,
        "json_files": [],
        "total_rows_all_files": 0,
        "rows_for_staff_before_filters": 0,
    }
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    # Fetch approved leaves and permissions for this lecturer to cross-reference
    approved_leaves = list(leaves.find({
        "lecturer_id": str(current_user.id),
        "status": "Approved"
    }))
    approved_permissions = list(permissions.find({
        "lecturer_id": str(current_user.id),
        "status": "Approved"
    }))
    
    leave_dates = {}
    for l in approved_leaves:
        # Assuming from_date and to_date are strings like "YYYY-MM-DD" or similar
        # We need to map every date in the range to the leave type
        try:
            start_dt = datetime.strptime(l['from_date'][:10], "%Y-%m-%d")
            end_dt = datetime.strptime(l['to_date'][:10], "%Y-%m-%d")
            curr = start_dt
            while curr <= end_dt:
                leave_dates[curr.strftime("%Y-%m-%d")] = l['type']
                curr += timedelta(days=1)
        except: pass
        
    permission_dates = {p['date']: p for p in approved_permissions if p.get('date')}

    if base_dir and debug_info["dir_exists"] and staff_id:
        for fname in os.listdir(base_dir):
            if not fname.lower().endswith(".json") or not fname.startswith(selected_month):
                continue
            
            fpath = os.path.join(base_dir, fname)
            iso_date = fname.replace(".json", "")
            display_date = ""
            try:
                d_obj = datetime.strptime(iso_date, "%Y-%m-%d")
                display_date = d_obj.strftime("%d-%m-%Y")
            except:
                display_date = iso_date

            try:
                data = load_json_attendance_file(fpath)

                user_rows = [r for r in data if str(r.get("staff_id") or "").strip().upper() == staff_id.upper()]
                if user_rows:
                    perm = permission_dates.get(iso_date)
                    if perm and isinstance(perm, dict):
                        t_from = perm.get('time_from') or perm.get('start_time') or ""
                        t_to = perm.get('time_to') or perm.get('end_time') or ""
                        p_note = f"Permission ({t_from} to {t_to})" if (t_from and t_to) else "Permission"
                        for r in user_rows:
                            if not r.get('status_note'):
                                r['status_note'] = p_note

                    checkins = [parse_ts(r.get("checkin")) for r in user_rows if r.get("checkin")]
                    checkouts = [parse_ts(r.get("checkout")) for r in user_rows if r.get("checkout")]
                    min_ci = min(checkins) if checkins else None
                    max_co = max(checkouts) if checkouts else None
                    time_in = min_ci.strftime("%H:%M") if min_ci else ""
                    time_out = max_co.strftime("%H:%M") if max_co else ""

                    dt_obj = datetime.fromisoformat(iso_date)
                    staff_type = determine_staff_type(staff_doc)
                    deadline_in, threshold_out = get_thresholds_for(dt_obj, staff_type, staff_id)
                    req_end_dt = datetime.combine(dt_obj.date(), threshold_out)

                    if min_ci and max_co:
                        status = "Present"
                    elif min_ci:
                        if iso_date == today_str and datetime.now() < req_end_dt:
                            status = "Checked-in"
                        else:
                            status = "Absent"
                    else:
                        status = "Absent"

                    try:
                        delay_val = compute_daily_delay(user_rows, dt_obj, staff_doc)
                    except:
                        delay_val = "00:00:00"

                    # Override status if on leave/permission
                    l_type = leave_dates.get(iso_date)
                    p_type = permission_dates.get(iso_date)
                    if l_type: status = l_type
                    elif p_type: status = "Permission"

                    records.append({
                        "date": iso_date,
                        "display_date": display_date,
                        "time_in": time_in,
                        "time_out": time_out,
                        "status": status,
                        "delay": delay_val
                    })
                else:
                    # Override status if on leave/permission even if no record in file
                    l_type = leave_dates.get(iso_date)
                    p_type = permission_dates.get(iso_date)
                    
                    status = "Absent"
                    if l_type: status = l_type
                    elif p_type: status = "Permission"
                    
                    records.append({
                        "date": iso_date,
                        "display_date": display_date,
                        "time_in": "--:--",
                        "time_out": "--:--",
                        "status": status,
                        "delay": "Absent"
                    })
            except:
                continue

    # Sort by date+time descending
    def sort_key(rec):
        return (rec.get("date") or "", rec.get("time") or "")

    records.sort(key=sort_key, reverse=True)

    today_records = [r for r in records if r.get("date") == today_str]

    return render_template(
        "lecturer/attendance.html",
        records=records,
        today_records=today_records,
        month=selected_month,
        q=search_q,
        debug_info=debug_info,
    )

@app.route('/admin/faculty-attendance')
@login_required

@admin_required
def admin_faculty_attendance():
    """
    Admin view for all faculty attendance.
    Reads all JSON files in ATTENDANCE_DIR and summarizes them.
    Also cross-references with approved leaves and permissions.
    """
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    
    # Filters
    selected_date = (request.args.get("date") or "").strip()
    staff_filter = (request.args.get("staff_id") or "").strip().upper()
    
    view_date = selected_date or datetime.now().strftime('%Y-%m-%d')
    records = []
    
    # 1. Fetch all faculty to ensure we cover everyone
    all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
    staff_lookup = {s['staff_id']: s for s in all_staff}
    
    # 2. Fetch approved leaves and permissions for the view date
    # Leaves: from_date <= view_date <= to_date
    approved_leaves = list(leaves.find({
        "status": "Approved",
        "from_date": {"$lte": f"{view_date} 23:59:59"},
        "to_date": {"$gte": f"{view_date} 00:00:00"}
    }))
    
    # Permissions: date == view_date
    approved_permissions = list(permissions.find({
        "status": "Approved",
        "date": view_date
    }))
    
    leave_map = {l['lecturer_id']: l['type'] for l in approved_leaves}
    # Permission map - some might use staff_id or lecturer_id (ObjectId string)
    permission_map = {}
    for p in approved_permissions:
        pid = p.get('lecturer_id') or p.get('staff_id')
        permission_map[pid] = "Permission"

    attendance_data_map = {}
    
    if base_dir and os.path.isdir(base_dir):
        files_to_read = []
        if selected_date:
            fname = f"{selected_date}.json"
            if os.path.exists(os.path.join(base_dir, fname)):
                files_to_read = [fname]
        else:
            files_to_read = sorted([f for f in os.listdir(base_dir) if f.lower().endswith(".json")], reverse=True)[:30]
            
        for fname in files_to_read:
            fpath = os.path.join(base_dir, fname)
            date_part = fname.replace(".json", "")
            
            try:
                data = load_json_attendance_file(fpath)
                
                grouped = {}
                for row in data:
                    s_id = str(row.get("staff_id") or "").strip().upper()
                    if s_id:
                        if s_id not in grouped: grouped[s_id] = []
                        grouped[s_id].append(row)

                for s_id, s_rows in grouped.items():
                    if staff_filter and s_id != staff_filter:
                        continue
                    
                    if date_part == view_date:
                        attendance_data_map[s_id] = s_rows[0]

                    u_doc = staff_lookup.get(s_id)
                    try:
                        dt_obj = datetime.fromisoformat(date_part)
                        staff_type = determine_staff_type(u_doc)
                        deadline_in, threshold_out = get_thresholds_for(dt_obj, staff_type, s_id)
                        req_end_dt = datetime.combine(dt_obj.date(), threshold_out)
                    except:
                        req_end_dt = datetime.now()
                        dt_obj = datetime.now()

                    checkins = [parse_ts(r.get("checkin")) for r in s_rows if r.get("checkin")]
                    checkouts = [parse_ts(r.get("checkout")) for r in s_rows if r.get("checkout")]
                    min_ci = min(checkins) if checkins else None
                    max_co = max(checkouts) if checkouts else None

                    if min_ci and max_co:
                        status = "Present"
                    elif min_ci:
                        if date_part == datetime.now().strftime('%Y-%m-%d') and datetime.now() < req_end_dt:
                            status = "Checked-in"
                        else:
                            status = "Absent"
                    else:
                        status = "Absent"

                    try:
                        delay_val = compute_daily_delay(s_rows, dt_obj, u_doc)
                    except:
                        delay_val = "00:00:00"
                    
                    if date_part == view_date:
                        l_type = leave_map.get(str(u_doc['_id']) if u_doc else "")
                        p_type = permission_map.get(str(u_doc['_id']) if u_doc else "") or permission_map.get(s_id)
                        if l_type: status = l_type
                        elif p_type: status = "Permission"

                    records.append({
                        "date": date_part,
                        "staff_id": s_id,
                        "name": s_rows[0].get("name") or (u_doc['name'] if u_doc else "Unknown"),
                        "checkin": min_ci.isoformat() if min_ci else "",
                        "checkout": max_co.isoformat() if max_co else "",
                        "status": status,
                        "delay": delay_val
                    })
            except:
                continue

    # Add staff who are NOT in the attendance file but might be on leave or absent
    if selected_date or view_date == datetime.now().strftime('%Y-%m-%d'):
        present_ids = set(attendance_data_map.keys())
        for staff in all_staff:
            s_id = staff['staff_id']
            if staff_filter and s_id != staff_filter:
                continue
                
            if s_id not in present_ids:
                l_type = leave_map.get(str(staff['_id']))
                p_type = permission_map.get(str(staff['_id'])) or permission_map.get(s_id)
                
                status = "Absent"
                if l_type: status = l_type
                elif p_type: status = "Permission"
                
                records.append({
                    "date": view_date,
                    "staff_id": s_id,
                    "name": staff.get('name', 'Unknown'),
                    "checkin": "",
                    "checkout": "",
                    "status": status,
                    "delay": "Absent"
                })

    # Sort records by date descending, then name
    records.sort(key=lambda x: (x['date'], x['name']), reverse=True)
    
    # Calculate stats for the current/selected view
    total_staff_count = len(all_staff)
    present_count = 0
    late_count = 0
    checked_in_count = 0
    on_leave_count = 0
    
    view_records = [r for r in records if r['date'] == view_date]
    for r in view_records:
        if r['status'] == 'Present':
            present_count += 1
        elif r['status'] == 'Checked-in':
            checked_in_count += 1
        elif r['status'] not in ['Absent', 'Unknown']:
            on_leave_count += 1
            
        if r['delay'] != '00:00:00':
            late_count += 1
            
    stats = {
        "total": total_staff_count,
        "present": present_count,
        "late": late_count,
        "checked_in": checked_in_count,
        "on_leave": on_leave_count,
        "absent": max(0, total_staff_count - (present_count + checked_in_count + on_leave_count)),
        "view_date": view_date
    }
    
    return render_template(
        "admin/faculty_attendance.html",
        records=records,
        all_staff=all_staff,
        selected_date=selected_date,
        staff_filter=staff_filter,
        stats=stats
    )

@app.route('/api/admin/attendance/log', methods=['POST'])
@login_required
@admin_required
def api_log_attendance():
    """
    Log attendance manually or from device.
    Writes to JSON file and emits socket event for live updates.
    """
    data = request.json
    s_id = data.get('staff_id', '').upper()
    checkin = data.get('checkin')
    checkout = data.get('checkout')
    date_str = data.get('date') or datetime.now().strftime('%Y-%m-%d')
    
    if not s_id:
        return jsonify({"success": False, "message": "Staff ID required"}), 400
        
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    if not base_dir or not os.path.isdir(base_dir):
        return jsonify({"success": False, "message": "Attendance directory not configured"}), 500
        
    fpath = os.path.join(base_dir, f"{date_str}.json")
    
    try:
        records = []
        if os.path.exists(fpath):
            with open(fpath, 'r', encoding='utf-8') as f:
                try:
                    records = json.load(f)
                    if isinstance(records, dict): records = [records]
                except:
                    f.seek(0)
                    records = [json.loads(line) for line in f if line.strip()]
        
        # Find existing record for this staff on this day
        found = False
        for r in records:
            if str(r.get('staff_id', '')).upper() == s_id:
                if checkin: r['checkin'] = checkin
                if checkout: r['checkout'] = checkout
                found = True
                break
        
        if not found:
            u_doc = users.find_one({"staff_id": s_id})
            records.append({
                "staff_id": s_id,
                "name": u_doc.get('name') if u_doc else "Unknown",
                "checkin": checkin or "",
                "checkout": checkout or ""
            })
            
        with open(fpath, 'w', encoding='utf-8') as f:
            json.dump(records, f, indent=4)
            
        # Emit SocketIO event for live update
        socketio.emit('attendance_update', {
            "date": date_str,
            "staff_id": s_id,
            "checkin": checkin,
            "checkout": checkout
        })
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/admin/attendance-report/daily/pdf')
@login_required
@admin_required
def admin_daily_attendance_report_pdf():
    """Export daily attendance report as PDF, matching the original design."""
    date_str = request.args.get('date')
    if not date_str:
        return "Date is required", 400
    
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    fpath = os.path.join(base_dir, f"{date_str}.json")
    if not os.path.exists(fpath):
        return f"No attendance data found for {date_str}", 404
        
    try:
        with open(fpath, encoding="utf-8") as f:
            try:
                data = json.load(f)
                if isinstance(data, dict): data = [data]
            except:
                f.seek(0)
                data = [json.loads(line) for line in f if line.strip()]
        
        # Group by ID and calculate delays
        records_by_id = {}
        for row in data:
            fid = (row.get('staff_id') or row.get('student_id') or '').strip().upper()
            if fid in EXCLUDED_FACULTY_IDS: continue
            records_by_id.setdefault(fid, []).append(row)
            
        dt_obj = datetime.strptime(date_str, "%Y-%m-%d")
        
        # Get all faculty from DB
        all_lecturers = list(users.find({"role": "lecturer"}))
        table_data = [['Faculty ID', 'Name', 'Check-in', 'Check-out', 'Status', 'Delay']]
        
        # Natural sort for IDs
        def natural_sort_key(u):
            fid = u.get('staff_id', '').upper()
            match = re.search(r'(\d+)', fid)
            return (fid[:match.start()] if match else fid, int(match.group(1)) if match else 0, fid)
        
        all_lecturers.sort(key=natural_sort_key)
        
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=10, leading=12)
        
        for u in all_lecturers:
            fid = u.get('staff_id', '').upper()
            if fid in EXCLUDED_FACULTY_IDS: continue
            
            rows = records_by_id.get(fid, [])
            if rows:
                row = rows[0] # Simplification
                ci = row.get('checkin') or ""
                co = row.get('checkout') or ""
                status = "Present" if (ci and co) else ("Checked-in" if ci else "Unknown")
                delay = compute_daily_delay(rows, dt_obj, u)
                
                ci_time = parse_ts(ci).strftime('%H:%M:%S') if parse_ts(ci) else "--:--"
                co_time = parse_ts(co).strftime('%H:%M:%S') if parse_ts(co) else "--:--"
            else:
                ci_time, co_time, status, delay = "Absent", "Absent", "Absent", "00:00:00"
            
            table_data.append([
                fid,
                Paragraph(u.get('name', 'Unknown'), cell_style),
                ci_time,
                co_time,
                status,
                delay
            ])
            
        # PDF Generation
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=18)
        story = [create_pdf_header(), Spacer(1, 10)]
        
        report_title_style = ParagraphStyle('ReportTitle', parent=styles['Heading2'], fontSize=14, spaceAfter=5, alignment=1, fontName='Helvetica-Bold')
        story.append(Paragraph(f"Faculty Attendance Report - {date_str}", report_title_style))
        story.append(Spacer(1, 15))
        
        table = Table(table_data, colWidths=[80, 180, 65, 65, 60, 60], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ]))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"Daily_Attendance_{date_str}.pdf", mimetype='application/pdf')
    except Exception as e:
        return f"Error: {str(e)}", 500

@app.route('/admin/attendance-report/monthly/pdf')
@login_required
@admin_required
def admin_monthly_attendance_report_pdf():
    """Export monthly delay summary as PDF."""
    month_str = request.args.get('month') # YYYY-MM
    if not month_str:
        return "Month is required", 400
        
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    all_lecturers = list(users.find({"role": "lecturer"}))
    
    # Natural sort
    def natural_sort_key(u):
        fid = u.get('staff_id', '').upper()
        match = re.search(r'(\d+)', fid)
        return (fid[:match.start()] if match else fid, int(match.group(1)) if match else 0, fid)
    all_lecturers.sort(key=natural_sort_key)
    
    summary = []
    
    # Iterate through all files for that month
    month_files = [f for f in os.listdir(base_dir) if f.startswith(month_str) and f.endswith(".json")]
    
    for u in all_lecturers:
        fid = u.get('staff_id', '').upper()
        if fid in EXCLUDED_FACULTY_IDS: continue
        
        total_seconds = 0
        for fname in month_files:
            fpath = os.path.join(base_dir, fname)
            date_obj = datetime.strptime(fname.replace(".json", ""), "%Y-%m-%d")
            
            with open(fpath, encoding="utf-8") as f:
                try:
                    data = json.load(f)
                    if isinstance(data, dict): data = [data]
                except:
                    f.seek(0)
                    data = [json.loads(line) for line in f if line.strip()]
            
            rows = [r for r in data if (r.get('staff_id') or r.get('student_id') or '').strip().upper() == fid]
            if rows:
                delay = compute_daily_delay(rows, date_obj, u)
                if delay and ":" in delay:
                    h, m, s = map(int, delay.split(":"))
                    total_seconds += h * 3600 + m * 60 + s
        
        summary.append({
            'fid': fid,
            'name': u.get('name', 'Unknown'),
            'total_delay': format_to_hhmmss(total_seconds)
        })
        
    # PDF Generation
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=40, bottomMargin=18)
    styles = getSampleStyleSheet()
    story = [create_pdf_header(), Spacer(1, 20)]
    
    story.append(Paragraph(f"Monthly Delay Summary - {month_str}", ParagraphStyle('Title', parent=styles['Heading2'], alignment=1, spaceAfter=20)))
    
    table_data = [['Faculty ID', 'Name', 'Total Delay (H:M:S)']]
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=11)
    
    for item in summary:
        table_data.append([
            item['fid'],
            Paragraph(item['name'], cell_style),
            item['total_delay']
        ])
        
    table = Table(table_data, colWidths=[100, 250, 100], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Monthly_Delay_Report_{month_str}.pdf", mimetype='application/pdf')

@app.route('/admin/attendance-report/detailed/pdf')
@login_required
@admin_required
def admin_detailed_faculty_report_pdf():
    """Export detailed attendance for one faculty for a month."""
    staff_id = request.args.get('staff_id')
    month_str = request.args.get('month') # YYYY-MM
    
    if not staff_id or not month_str:
        return "Staff ID and Month are required", 400
        
    u_doc = users.find_one({"staff_id": staff_id.upper()})
    if not u_doc:
        return f"Faculty {staff_id} not found", 404
        
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    month_files = sorted([f for f in os.listdir(base_dir) if f.startswith(month_str) and f.endswith(".json")])
    
    story = [create_pdf_header(), Spacer(1, 20)]
    styles = getSampleStyleSheet()
    
    story.append(Paragraph("Detailed Faculty Attendance Report", ParagraphStyle('Title', parent=styles['Heading2'], alignment=1)))
    story.append(Paragraph(f"Faculty: {u_doc.get('name')} ({staff_id})", ParagraphStyle('Sub', parent=styles['Normal'], alignment=1)))
    story.append(Paragraph(f"Month: {month_str}", ParagraphStyle('Sub', parent=styles['Normal'], alignment=1, spaceAfter=20)))
    
    table_data = [['Date', 'Check-in', 'Check-out', 'Status', 'Delay']]
    
    for fname in month_files:
        date_part = fname.replace(".json", "")
        date_obj = datetime.strptime(date_part, "%Y-%m-%d")
        fpath = os.path.join(base_dir, fname)
        
        with open(fpath, encoding="utf-8") as f:
            try:
                data = json.load(f)
                if isinstance(data, dict): data = [data]
            except:
                f.seek(0)
                data = [json.loads(line) for line in f if line.strip()]
                
        rows = [r for r in data if (r.get('staff_id') or r.get('student_id') or '').strip().upper() == staff_id.upper()]
        if rows:
            row = rows[0]
            ci = row.get('checkin') or ""
            co = row.get('checkout') or ""
            status = "Present" if (ci and co) else ("Checked-in" if ci else "Unknown")
            delay = compute_daily_delay(rows, date_obj, u_doc)
            ci_time = parse_ts(ci).strftime('%H:%M:%S') if parse_ts(ci) else "--:--"
            co_time = parse_ts(co).strftime('%H:%M:%S') if parse_ts(co) else "--:--"
        else:
            ci_time, co_time, status, delay = "--", "--", "Absent", "00:00:00"
            
        table_data.append([date_part, ci_time, co_time, status, delay])
        
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=18)
    table = Table(table_data, colWidths=[90, 80, 80, 100, 80], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Detailed_Attendance_{staff_id}_{month_str}.pdf", mimetype='application/pdf')

@app.route('/admin/attendance-report/daily/excel')
@login_required
@admin_required
def admin_daily_attendance_report_excel():
    """Export daily attendance report as Excel."""
    date_str = request.args.get('date')
    if not date_str:
        return "Date is required", 400
        
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    fpath = os.path.join(base_dir, f"{date_str}.json")
    if not os.path.exists(fpath):
        return f"No data found for {date_str}", 404
        
    with open(fpath, encoding="utf-8") as f:
        try:
            data = json.load(f)
            if isinstance(data, dict): data = [data]
        except:
            f.seek(0)
            data = [json.loads(line) for line in f if line.strip()]
            
    records_by_id = { (r.get('staff_id') or r.get('student_id') or '').strip().upper(): r for r in data }
    all_lecturers = list(users.find({"role": "lecturer"}))
    
    excel_data = []
    dt_obj = datetime.strptime(date_str, "%Y-%m-%d")
    
    for u in all_lecturers:
        fid = u.get('staff_id', '').upper()
        row = records_by_id.get(fid, {})
        ci = row.get('checkin') or ""
        co = row.get('checkout') or ""
        
        excel_data.append({
            'Staff ID': fid,
            'Name': u.get('name'),
            'Check-in': parse_ts(ci).strftime('%H:%M:%S') if parse_ts(ci) else "Absent",
            'Check-out': parse_ts(co).strftime('%H:%M:%S') if parse_ts(co) else "Absent",
            'Status': "Present" if (ci and co) else ("Checked-in" if ci else "Absent"),
            'Delay': compute_daily_delay([row] if row else [], dt_obj, u)
        })
        
    df = pd.DataFrame(excel_data)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance')
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"Daily_Attendance_{date_str}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/admin/attendance/notifications/delays')
@login_required
@admin_required
def api_admin_attendance_notifications_delays():
    """Get monthly delay analysis data."""
    month = request.args.get('month') # MM
    year = request.args.get('year') # YYYY
    if not month or not year:
        return jsonify({"success": False, "error": "Month and year required"})
        
    month_str = f"{year}-{month.zfill(2)}"
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    all_lecturers = list(users.find({"role": "lecturer"}))
    month_files = [f for f in os.listdir(base_dir) if f.startswith(month_str) and f.endswith(".json")]
    
    results = []
    for u in all_lecturers:
        fid = u.get('staff_id', '').upper()
        if fid in EXCLUDED_FACULTY_IDS: continue
        
        total_seconds = 0
        delay_count = 0
        for fname in month_files:
            fpath = os.path.join(base_dir, fname)
            date_obj = datetime.strptime(fname.replace(".json", ""), "%Y-%m-%d")
            with open(fpath, encoding="utf-8") as f:
                try: data = json.load(f)
                except: continue
                if isinstance(data, dict): data = [data]
            
            rows = [r for r in data if (r.get('staff_id') or r.get('student_id') or '').strip().upper() == fid]
            if rows:
                delay = compute_daily_delay(rows, date_obj, u)
                if delay and delay != "00:00:00":
                    h, m, s = map(int, delay.split(":"))
                    total_seconds += h * 3600 + m * 60 + s
                    delay_count += 1
                    
        if delay_count > 0:
            results.append({
                'faculty_id': fid,
                'faculty_name': u.get('name'),
                'total_delay': format_to_hhmmss(total_seconds),
                'delay_count': delay_count
            })
            
    results.sort(key=lambda x: x['delay_count'], reverse=True)
    return jsonify({"success": True, "data": results})

@app.route('/api/admin/attendance/notifications/absences')
@login_required
@admin_required
def api_admin_attendance_notifications_absences():
    """Get monthly absence analysis data."""
    month = request.args.get('month')
    year = request.args.get('year')
    if not month or not year:
        return jsonify({"success": False, "error": "Month and year required"})
        
    month_str = f"{year}-{month.zfill(2)}"
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    all_lecturers = list(users.find({"role": "lecturer"}))
    month_files = [f for f in os.listdir(base_dir) if f.startswith(month_str) and f.endswith(".json")]
    
    results = []
    for u in all_lecturers:
        fid = u.get('staff_id', '').upper()
        if fid in EXCLUDED_FACULTY_IDS: continue
        
        absent_count = 0
        present_count = 0
        max_consecutive = 0
        current_consecutive = 0
        
        # Check every day of the month files
        for fname in sorted(month_files):
            fpath = os.path.join(base_dir, fname)
            with open(fpath, encoding="utf-8") as f:
                try: data = json.load(f)
                except: continue
                if isinstance(data, dict): data = [data]
            
            rows = [r for r in data if (r.get('staff_id') or r.get('student_id') or '').strip().upper() == fid]
            if rows:
                present_count += 1
                current_consecutive = 0
            else:
                absent_count += 1
                current_consecutive += 1
                max_consecutive = max(max_consecutive, current_consecutive)
                
        if absent_count > 0:
            results.append({
                'faculty_id': fid,
                'faculty_name': u.get('name'),
                'absent_count': absent_count,
                'present_count': present_count,
                'max_continuous_absent': max_consecutive
            })
            
    results.sort(key=lambda x: x['absent_count'], reverse=True)
    return jsonify({"success": True, "data": results})

@app.route('/lecturer/apply-leave', methods=['GET', 'POST'])
@login_required
@lecturer_required
def apply_leave():
    mode = request.args.get('mode', 'full')
    if request.method == 'POST':
        leave_mode = request.form.get('mode', 'full')
        
        if leave_mode == 'time':
            # Use user-selected date as base
            base_date = request.form.get('from_date') or datetime.now().strftime('%Y-%m-%d')
            time_from = request.form.get('time_from', '')
            time_to = request.form.get('time_to', '')
            from_date = f"{base_date} {time_from}"
            to_date = f"{base_date} {time_to}"
            
            # Validation: Permission must be within Duty Hours
            lecturer = users.find_one({"_id": ObjectId(current_user.id)})
            staff_id = lecturer.get('staff_id')
            duty_start, duty_end = get_faculty_duty_bounds(staff_id)
            
            # Convert to comparable integers (minutes from midnight)
            def to_mins(t):
                try:
                    h, m = map(int, t.split(':'))
                    return h * 60 + m
                except: return 0
            
            req_start = to_mins(time_from)
            req_end = to_mins(time_to)
            d_start = to_mins(duty_start)
            d_end = to_mins(duty_end)
            
            if req_start < d_start or req_end > d_end:
                flash(f"Invalid Permission Time! Your duty hours are {duty_start} to {duty_end}. Please apply within these hours.", "danger")
                return redirect(url_for('apply_leave', mode='time'))
        else:
            from_date = request.form.get('from_date')
            to_date = request.form.get('to_date')

        today_str = datetime.now().strftime('%Y-%m-%d')
        now_mins = datetime.now().hour * 60 + datetime.now().minute

        if leave_mode == 'time':
            if base_date < today_str:
                flash("Permission Date cannot be in the past.", "danger")
                return redirect(url_for('apply_leave', mode='time'))
            if base_date == today_str and req_start < now_mins:
                flash("Permission 'From' time cannot be in the past.", "danger")
                return redirect(url_for('apply_leave', mode='time'))
        else:
            if from_date < today_str:
                flash("From Date cannot be in the past.", "danger")
                return redirect(url_for('apply_leave', mode=leave_mode))
            if to_date and to_date < today_str:
                flash("To Date cannot be in the past.", "danger")
                return redirect(url_for('apply_leave', mode=leave_mode))
            if from_date == today_str:
                if now_mins >= 15 * 60:
                    flash("It's past 3:00 PM. Current day leave is no longer available.", "danger")
                    return redirect(url_for('apply_leave', mode=leave_mode))
                elif now_mins >= 9 * 60 + 30:
                    half_day_flag = request.form.get('half_day') == 'on'
                    session_val = request.form.get('session')
                    if not half_day_flag or session_val != 'afternoon':
                        flash("It's past 9:30 AM. Today you can only apply for 'Afternoon Half-Day'.", "danger")
                        return redirect(url_for('apply_leave', mode=leave_mode))

        # CLEANUP: Delete only STALE drafts (Pending/Rejected) before submitting.
        # We must PRESERVE 'accepted' drafts so they can be linked to this leave.
        old_stale_drafts = list(leave_class_allocations.find({
            "assigned_by": str(current_user.id),
            "is_draft": True,
            "status": {"$in": ["Pending", "rejected"]}
        }))
        old_stale_ids = [str(d['_id']) for d in old_stale_drafts]
        
        if old_stale_ids:
            leave_class_allocations.delete_many({"_id": {"$in": [ObjectId(id) for id in old_stale_ids]}})
            faculty_notifications.delete_many({"allocation_id": {"$in": old_stale_ids}})

        if leave_mode == 'time':
            permission_data = {
                "lecturer_id": current_user.id,
                "lecturer_name": current_user.name,
                "type": "Permission",
                "from_date": from_date,
                "to_date": to_date,
                "reason": request.form.get('reason'),
                "status": "Pending",
                "created_at": datetime.now(),
                "mode": "time",
                "half_day": False
            }
            res = permissions.insert_one(permission_data)
            leave_id = str(res.inserted_id)
        else:
            no_class_dates_raw = request.form.get('no_class_dates_json')
            no_class_dates = []
            if no_class_dates_raw:
                try:
                    import json
                    no_class_dates = json.loads(no_class_dates_raw) or []
                except Exception:
                    no_class_dates = []

            half_day_flag = request.form.get('half_day') == 'on'
            leave_data = {
                "lecturer_id": current_user.id,
                "lecturer_name": current_user.name,
                "type": request.form.get('type'),
                "from_date": from_date,
                "to_date": to_date,
                "reason": request.form.get('reason'),
                "status": "Pending",
                "created_at": datetime.now(),
                "mode": leave_mode,
                "half_day": half_day_flag,
                "session": request.form.get('session') if half_day_flag else None,
                "no_class_dates": no_class_dates,
                "working_days": count_working_leave_days(from_date, to_date, half_day=half_day_flag)
            }
            res = leaves.insert_one(leave_data)
            leave_id = str(res.inserted_id)
        
        # Link HOD request to this leave
        hod_requests.find_one_and_update(
            {"requester_id": str(current_user.id), "status": "Approved", "leave_id": {"$exists": False}},
            {"$set": {"leave_id": leave_id}},
            sort=[("created_at", -1)]
        )
        
        # Process assignments if provided (integrated flow)
        assignments_raw = request.form.get('assignments_json')
        if assignments_raw:
            try:
                import json
                assignments_data = json.loads(assignments_raw)
                for entry in assignments_data:
                    assigned_to_id = entry.get('assigned_to_id')
                    class_details = entry.get('class_details')
                    if assigned_to_id and class_details:
                        # DUPLICATE SHIELD: Check if this class is already assigned and accepted/approved
                        existing = leave_class_allocations.find_one({
                            "assigned_by": str(current_user.id),
                            "assigned_to": assigned_to_id,
                            "class_details.date": class_details.get('date'),
                            "class_details.time": class_details.get('time'),
                            "class_details.subject": class_details.get('subject'),
                            "status": {"$in": ["accepted", "approved", "finalized"]}
                        })
                        if existing:
                            # Already accepted by colleague! Link it to this leave and mark as permanent.
                            leave_class_allocations.update_one(
                                {"_id": existing['_id']}, 
                                {"$set": {"leave_id": leave_id, "is_draft": False}}
                            )
                            continue

                        target_faculty = users.find_one({"_id": ObjectId(assigned_to_id)})
                        # 1. Save Allocation Record (Permanent, not a draft)
                        alloc_res = leave_class_allocations.insert_one({
                            "leave_id": leave_id,
                            "assigned_by": str(current_user.id),
                            "assigned_by_name": current_user.name,
                            "assigned_to": assigned_to_id,
                            "assigned_to_name": target_faculty.get('name', 'Unknown') if target_faculty else 'Unknown',
                            "class_details": class_details,
                            "status": "Pending",
                            "created_at": datetime.now(),
                            "is_draft": False
                        })
                        alloc_id = str(alloc_res.inserted_id)

                        # 2. Create Notification for the assigned faculty
                        faculty_notifications.insert_one({
                            "recipient_id": assigned_to_id,
                            "sender_id": str(current_user.id),
                            "sender_name": current_user.name,
                            "message": f"{current_user.name} requested you for substitution: {class_details.get('subject')} on {class_details.get('date')} at {class_details.get('time')}.",
                            "type": "class_assignment",
                            "allocation_id": alloc_id,
                            "leave_id": leave_id,  # LINKED FOR CLEANUP
                            "class_details": class_details,
                            "status": "unread",
                            "created_at": datetime.now()
                        })

                        # 3. Real-time Socket Notification
                        socketio.emit('new_class_assignment', {
                            "recipient_id": assigned_to_id,
                            "message": f"New class substitution request from {current_user.name}",
                            "allocation_id": alloc_id
                        })
            except Exception as e:
                print(f"DEBUG: Assignment Error: {e}")

        socketio.emit('new_leave_request', {
            "id": leave_id,
            "lecturer_name": current_user.name,
            "type": "Permission" if leave_mode == 'time' else request.form.get('type'),
            "from_date": from_date,
            "to_date": to_date,
            "status": "Pending",
            "mode": leave_mode
        })
        
        # CLEANUP: Clear the draft after successful submission
        leave_drafts.delete_one({"user_id": str(current_user.id)})

        flash("Leave application submitted successfully with class assignments!", "success")
        return redirect(url_for('lecturer_dashboard'))
    
    staff_id_val = str(getattr(current_user, 'staff_id', '') or '').strip().upper()
    user_doc = users.find_one({"_id": ObjectId(current_user.id)}) if (current_user and hasattr(current_user, 'id')) else {}
    cat_val = str((user_doc or {}).get('category') or getattr(current_user, 'category', '') or '').strip().lower()
    stype_val = str((user_doc or {}).get('staff_type') or getattr(current_user, 'staff_type', '') or '').strip().lower()
    is_non_faculty = staff_id_val.startswith('BBHCFN') or 'non' in cat_val or 'non' in stype_val

    return render_template(
        'lecturer/apply_leave.html',
        mode=mode,
        current_user_staff_id=staff_id_val,
        is_non_faculty=is_non_faculty
    )


@app.route('/lecturer/leave/<id>/cancel', methods=['POST'])
@login_required
@lecturer_required
def cancel_leave(id):
    """
    Allow a lecturer to cancel one of their own pending leave requests.
    """
    leave_doc = leaves.find_one({"_id": ObjectId(id), "lecturer_id": current_user.id})
    if not leave_doc:
        flash("Leave request not found.", "danger")
        return redirect(url_for('lecturer_dashboard'))

    if leave_doc.get("status") != "Pending":
        flash("Only pending leave requests can be cancelled.", "warning")
        return redirect(url_for('lecturer_dashboard'))

    leaves.update_one(
        {"_id": leave_doc["_id"]},
        {"$set": {"status": "Cancelled", "cancelled_at": datetime.now()}},
    )
    
    socketio.emit('leave_cancelled', {'id': id, 'lecturer_id': current_user.id})
    flash("Leave request cancelled.", "success")
    return redirect(url_for('lecturer_dashboard'))

@app.route('/lecturer/leave/api/<id>/cancel', methods=['POST'])
@login_required
@lecturer_required
def api_cancel_leave(id):
    leave_doc = leaves.find_one({"_id": ObjectId(id), "lecturer_id": current_user.id})
    if not leave_doc:
        return jsonify({"success": False, "message": "Not found"}), 404
        
    if leave_doc.get("status") != "Pending":
        return jsonify({"success": False, "message": "Not pending"}), 400
        
    leaves.update_one(
        {"_id": leave_doc["_id"]},
        {"$set": {"status": "Cancelled", "cancelled_at": datetime.now()}},
    )
    
    socketio.emit('leave_cancelled', {'id': id, 'lecturer_id': current_user.id})
    return jsonify({"success": True})

@app.route('/lecturer/salary')
@login_required
@lecturer_required
def view_salary():
    # Show only uploaded/published slips to lecturers
    my_salaries = list(salaries.find({"lecturer_id": current_user.id, "published": True}).sort("month_year", -1))
    return render_template('lecturer/salary.html', salaries=my_salaries)

@app.route('/lecturer/salary/<salary_id>')
@login_required
@lecturer_required
def lecturer_salary_view(salary_id):
    doc = salaries.find_one({"_id": ObjectId(salary_id), "lecturer_id": current_user.id})
    if not doc:
        flash("Salary slip not found.", "danger")
        return redirect(url_for("view_salary"))
    payload = doc.get("payload") or {}
    return render_template("lecturer/salary_view.html", salary=doc, payload=payload)


def _num(v):
    try:
        if v is None:
            return 0.0
        s = str(v).strip()
        if s == "":
            return 0.0
        return float(s)
    except Exception:
        return 0.0


@app.route('/lecturer/salary/<salary_id>/pdf')
@login_required
@lecturer_required
def lecturer_salary_pdf(salary_id):
    doc = salaries.find_one({"_id": ObjectId(salary_id), "lecturer_id": current_user.id})
    if not doc:
        flash("Salary slip not found.", "danger")
        return redirect(url_for("view_salary"))

    pdf_bytes, filename = build_salary_pdf_bytes(doc.get("payload") or {})
    buffer = BytesIO(pdf_bytes)
    inline = request.args.get("inline") == "1"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=not inline,
        download_name=filename,
    )


@app.route('/lecturer/timetable')
@login_required
@lecturer_required
def lecturer_timetable():
    """Show the logged-in lecturer's own timetable image and structured data."""
    if current_user.staff_id and current_user.staff_id.startswith('BBHCFN'):
        flash('Timetable is not available for your category.', 'info')
        return redirect(url_for('lecturer_dashboard'))
    # 1. Try DB first
    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    
    # 2. Try disk-based JSON if DB is missing or structured is empty
    structured = tt_doc.get("structured") if tt_doc else {}
    if not structured:
        # Try finding by staff_id or username
        staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
        staff_id = staff_doc.get("staff_id") if staff_doc else current_user.username.upper()
        
        json_path = os.path.join(os.path.dirname(__file__), "static", "json_timetables", f"{staff_id}.json")
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    structured = json.load(f)
            except:
                pass

    # 3. Default empty structure if still nothing (allows manual creation)
    if not structured:
        structured = {
            "faculty": current_user.name,
            "timetable": {
                "periods": [
                    {"period": "0", "time": ""},
                    {"period": "I", "time": "9.45-10.35"},
                    {"period": "II", "time": "10.40-11.30"},
                    {"period": "III", "time": "11.35-12.25"},
                    {"period": "IV", "time": "1.05-1.55"},
                    {"period": "V", "time": "2.00-2.50"},
                    {"period": "VI", "time": "2.55-3.45"},
                    {"period": "VII", "time": ""}
                ],
                "days": [{"day": d, "slots": {}} for d in ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]]
            }
        }

    image_url = None
    if tt_doc and tt_doc.get("image_path"):
        image_path = (tt_doc.get("image_path") or "").replace("\\", "/")
        image_url = url_for("static", filename=image_path)
        
    accepted_assignments_cursor = leave_class_allocations.find({
        "assigned_to": str(current_user.id),
        "status": {"$in": ["accepted", "approved"]}
    })
    
    from datetime import datetime
    now_str = datetime.now().strftime('%Y-%m-%d')
    now_str2 = datetime.now().strftime('%d-%m-%Y')
    now_str3 = datetime.now().strftime('%d/%m/%Y')
    
    accepted_assignments = []
    for a in accepted_assignments_cursor:
        date_str = a.get('class_details', {}).get('date', '')
        if date_str in [now_str, now_str2, now_str3]:
            a['_id'] = str(a['_id'])
            accepted_assignments.append(a)
    
    return render_template(
        'lecturer/timetable.html',
        has_timetable=image_url is not None,
        timetable_image_url=image_url,
        structured=structured,
        accepted_assignments=accepted_assignments
    )


@app.route('/lecturer/timetable/edit', methods=['GET', 'POST'])
@login_required
@lecturer_required
def edit_lecturer_timetable():
    """
    Saves or updates the structured timetable data for the current lecturer.
    """
    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    
    if request.method == 'POST':
        raw = request.form.get("structured_json", "").strip()
        if not raw:
            flash("Timetable data cannot be empty.", "danger")
            return redirect(url_for('lecturer_timetable'))
        try:
            data = json.loads(raw)
            if not isinstance(data, dict):
                raise ValueError("Data must be a JSON object.")
        except Exception as exc:
            flash(f"Invalid data format: {exc}", "danger")
            return redirect(url_for('lecturer_timetable'))

        staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
        staff_id = (staff_doc.get("staff_id") if staff_doc else current_user.username).upper()
        try:
            _persist_timetable_structured(
                current_user.id,
                current_user.name,
                staff_id,
                data,
            )
        except Exception as exc:
            flash(f"Failed to save timetable: {exc}", "danger")
            return redirect(url_for('lecturer_timetable'))

        socketio.emit(
            'timetable_updated',
            {'staff_id': staff_id},
            room=f'user_{current_user.id}',
        )
        flash("Timetable synchronized successfully.", "success")
        return redirect(url_for('lecturer_timetable'))

    # GET request - should ideally not reach here if using modal edit on the main timetable page
    # but kept for backward compatibility if template uses a separate page.
    structured = tt_doc.get("structured", {}) if tt_doc else {}
    return render_template(
        'lecturer/edit_timetable.html',
        structured_json=json.dumps(structured, indent=2, ensure_ascii=False),
    )


# ============ LEAVE CLASS ALLOCATION ROUTES ============

@app.route('/lecturer/leave/<leave_id>/allocate-classes')
@login_required
@lecturer_required
def allocate_leave_classes(leave_id):
    """Page to allocate classes to other faculty members"""
    leave_doc = leaves.find_one({"_id": ObjectId(leave_id), "lecturer_id": current_user.id})
    if not leave_doc:
        flash("Leave request not found.", "danger")
        return redirect(url_for('lecturer_dashboard'))
    
    if leave_doc.get('status') != 'Pending':
        flash("Can only allocate classes for pending leave requests.", "warning")
        return redirect(url_for('lecturer_dashboard'))
    
    # Get staff_id for current lecturer
    staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
    staff_id = staff_doc.get('staff_id') if staff_doc else None
    
    # Get classes during leave period
    from_date = leave_doc.get('from_date', '').split(' ')[0] if ' ' in leave_doc.get('from_date', '') else leave_doc.get('from_date', '')
    to_date = leave_doc.get('to_date', '').split(' ')[0] if ' ' in leave_doc.get('to_date', '') else leave_doc.get('to_date', '')
    
    if leave_doc.get('mode') == 'time':
        # For time-based leave, just use today
        today = datetime.now().strftime('%Y-%m-%d')
        classes = get_classes_on_date(staff_id, today) if staff_id else []
        # Filter by time if specified
        time_from = leave_doc.get('from_date', '').split(' ')[1] if ' ' in leave_doc.get('from_date', '') else None
        time_to = leave_doc.get('to_date', '').split(' ')[1] if ' ' in leave_doc.get('to_date', '') else None
        if time_from and time_to:
            classes = [c for c in classes if time_from <= c.get('time', '') <= time_to]
    else:
        classes = get_classes_for_leave_period(staff_id, from_date, to_date) if staff_id else []
    
    # Get all faculty except current user for assignment
    all_faculty_raw = list(users.find({
        "role": "lecturer",
        "_id": {"$ne": ObjectId(current_user.id)}
    }).sort("name", 1))
    
    # Convert faculty data to JSON-serializable format
    all_faculty = []
    for f in all_faculty_raw:
        faculty_data = {
            'id': str(f['_id']),
            'name': f.get('name', ''),
            'staff_id': f.get('staff_id', ''),
            'email': f.get('email', ''),
            'department': f.get('department', '')
        }
        all_faculty.append(faculty_data)
    
    # Get existing allocations
    existing_allocations_raw = list(leave_class_allocations.find({"leave_id": leave_id}))
    existing_allocations = []
    for a in existing_allocations_raw:
        alloc_data = {
            'id': str(a['_id']),
            'leave_id': a.get('leave_id'),
            'assigned_by': a.get('assigned_by'),
            'assigned_to': a.get('assigned_to'),
            'assigned_to_name': a.get('assigned_to_name'),
            'class_details': a.get('class_details', {}),
            'status': a.get('status'),
            'created_at': a.get('created_at').isoformat() if a.get('created_at') else None
        }
        existing_allocations.append(alloc_data)
    
    allocated_class_ids = [a['class_details'].get('class_id') for a in existing_allocations if a['class_details']]
    
    return render_template(
        'lecturer/allocate_classes.html',
        leave=leave_doc,
        classes=classes,
        faculty=all_faculty,
        allocated_class_ids=allocated_class_ids,
        existing_allocations=existing_allocations
    )


@app.route('/lecturer/leave/<leave_id>/assign-class', methods=['POST'])
@login_required
@lecturer_required
def assign_leave_class(leave_id):
    """Assign a specific class to another faculty member"""
    leave_doc = leaves.find_one({"_id": ObjectId(leave_id), "lecturer_id": current_user.id})
    if not leave_doc:
        return jsonify({"success": False, "message": "Leave not found"}), 404
    
    class_data = request.json
    assigned_to_id = class_data.get('assigned_to_id')
    class_details = class_data.get('class_details', {})
    
    # Create class allocation record
    allocation = {
        'leave_id': leave_id,
        'assigned_by': current_user.id,
        'assigned_to': assigned_to_id,
        'class_details': class_details,
        'status': 'pending_faculty',  # pending_faculty, accepted, rejected, approved
        'created_at': datetime.now()
    }
    
    result = leave_class_allocations.insert_one(allocation)
    allocation_id = str(result.inserted_id)
    
    # Create notification for assigned faculty
    notification = {
        'type': 'class_assignment',
        'allocation_id': allocation_id,
        'leave_id': leave_id,
        'recipient_id': assigned_to_id,
        'sender_id': current_user.id,
        'sender_name': current_user.name,
        'message': f"{current_user.name} has requested you to take their class on {class_details.get('date')} at {class_details.get('time')} ({class_details.get('subject')})",
        'class_details': class_details,
        'status': 'unread',
        'created_at': datetime.now()
    }
    faculty_notifications.insert_one(notification)
    
    # Emit socket event for real-time notification
    socketio.emit('new_class_assignment', {
        'recipient_id': assigned_to_id,
        'message': notification['message'],
        'allocation_id': allocation_id
    })
    
    return jsonify({
        "success": True,
        "allocation_id": allocation_id,
        "message": "Class assigned successfully. Waiting for faculty approval."
    })


@app.route('/lecturer/api/request-substitution', methods=['POST'])
@login_required
@lecturer_required
def api_request_substitution():
    data = request.json
    faculty_id = data.get('faculty_id')
    class_details = data.get('class_details')
    
    if not faculty_id or not class_details:
        return jsonify({"success": False, "message": "Missing data"}), 400
        
    target_faculty = users.find_one({"_id": ObjectId(faculty_id)})
    if not target_faculty:
        return jsonify({"success": False, "message": "Faculty not found"}), 404
        
    # Create allocation record (without leave_id yet, using 'Draft' or 'Pre-Leave')
    alloc_data = {
        "assigned_by": str(current_user.id),
        "assigned_by_name": current_user.name,
        "assigned_to": faculty_id,
        "assigned_to_name": target_faculty.get('name', 'Unknown'),
        "class_details": class_details,
        "status": "Pending",
        "created_at": datetime.now(),
        "is_draft": True
    }
    # CLEANUP: Delete only STALE draft allocations/notifications for this exact class
    # PRESERVE 'accepted' ones so the user doesn't have to ask again if they refresh
    old_stale = list(leave_class_allocations.find({
        "assigned_by": str(current_user.id),
        "class_details.date": class_details.get('date'),
        "class_details.time": class_details.get('time'),
        "is_draft": True,
        "status": {"$in": ["Pending", "rejected"]}
    }))
    old_stale_ids = [str(a['_id']) for a in old_stale]
    if old_stale_ids:
        leave_class_allocations.delete_many({"_id": {"$in": [ObjectId(i) for i in old_stale_ids]}})
        faculty_notifications.delete_many({"allocation_id": {"$in": old_stale_ids}})

    res = leave_class_allocations.insert_one(alloc_data)
    alloc_id = str(res.inserted_id)
    
    # Send Notification
    faculty_notifications.insert_one({
        "recipient_id": faculty_id,
        "sender_id": str(current_user.id),
        "sender_name": current_user.name,
        "message": f"{current_user.name} requested substitution for: {class_details.get('subject')} on {class_details.get('date')}.",
        "type": "class_assignment",
        "allocation_id": alloc_id,
        "class_details": class_details,
        "status": "unread",
        "created_at": datetime.now()
    })
    
    socketio.emit('new_class_assignment', {
        "recipient_id": faculty_id,
        "message": f"New substitution request from {current_user.name}",
        "allocation_id": alloc_id
    })
    
    return jsonify({"success": True, "allocation_id": alloc_id})

@app.route('/lecturer/api/save-draft', methods=['POST'])
@login_required
@lecturer_required
def api_save_leave_draft():
    data = request.json
    if not data:
        return jsonify({"success": False, "message": "No data provided"}), 400
    
    # Update or insert draft for current user
    leave_drafts.update_one(
        {"user_id": str(current_user.id)},
        {"$set": {
            "draft_data": data,
            "updated_at": datetime.now()
        }},
        upsert=True
    )
    return jsonify({"success": True})

@app.route('/lecturer/api/clear-draft', methods=['POST'])
@login_required
@lecturer_required
def api_clear_leave_draft():
    leave_drafts.delete_one({"user_id": str(current_user.id)})
    return jsonify({"success": True})

@app.route('/lecturer/api/get-draft', methods=['GET'])
@login_required
@lecturer_required
def api_get_leave_draft():
    draft = leave_drafts.find_one({"user_id": str(current_user.id)})
    if draft:
        return jsonify({"success": True, "draft": draft['draft_data']})
    return jsonify({"success": False, "message": "No draft found"})

@app.route('/lecturer/api/preview-classes')
@login_required
@lecturer_required
def api_preview_classes():
    """Preview classes that occur during a proposed leave period"""
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    mode = request.args.get('mode', 'full')
    time_from = request.args.get('time_from')
    time_to = request.args.get('time_to')
    half_day = request.args.get('half_day') == 'true'
    session_type = request.args.get('session')  # 'morning' or 'afternoon'
    
    # Get staff_id
    staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
    staff_id = staff_doc.get('staff_id') if staff_doc else None
    
    if not staff_id:
        return jsonify({"success": False, "message": "Staff ID not found", "classes": []})
        
    classes = []
    if mode == 'time':
        today = datetime.now().strftime('%Y-%m-%d')
        classes = get_classes_on_date(staff_id, today)
        if time_from and time_to:
            def time_to_min(t_str):
                """Converts time string (9.45, 1:05, 13:10) to minutes from midnight"""
                try:
                    if not t_str: return 0
                    # Clean and split
                    t_str = t_str.strip().replace('.', ':')
                    # Remove any characters that aren't digits or colons
                    t_str = "".join(c for c in t_str if c.isdigit() or c == ':')
                    if not t_str: return 0
                    
                    parts = t_str.split(':')
                    h = int(parts[0])
                    m = int(parts[1]) if (len(parts) > 1 and parts[1]) else 0
                    
                    # Robust AM/PM heuristic for college hours:
                    # 1-7 are PM (1:00 PM to 7:00 PM)
                    # 8-12 are AM (8:00 AM to 12:00 PM)
                    if 1 <= h <= 7:
                        h += 12
                    return h * 60 + m
                except Exception as e:
                    return 0

            u_start = time_to_min(time_from)
            u_end = time_to_min(time_to)
            
            filtered = []
            for c in classes:
                c_time = (c.get('time') or '').strip()
                if not c_time or c_time.upper() == 'TBD': continue
                
                # Split by '-' or ' TO ' or ' - '
                time_parts = re.split(r'[-–]| TO ', c_time, flags=re.IGNORECASE)
                
                if len(time_parts) >= 2:
                    try:
                        c_start = time_to_min(time_parts[0])
                        c_end = time_to_min(time_parts[1])
                        
                        # Strict overlap check: max(starts) < min(ends)
                        # We also ensure the class has a valid duration (c_start < c_end)
                        if c_start < c_end and max(u_start, c_start) < min(u_end, c_end):
                            filtered.append(c)
                    except:
                        continue
                else:
                    # Single time point comparison
                    point = time_to_min(c_time)
                    if u_start <= point < u_end:
                        filtered.append(c)
            classes = filtered
    else:
        if from_date and to_date:
            classes = get_classes_for_leave_period(staff_id, from_date, to_date)
            
    # Filter by session if half-day
    if half_day and session_type:
        morning_periods = ["0", "I", "II", "III"]
        afternoon_periods = ["IV", "V", "VI", "VII"]
        
        if session_type == 'morning':
            classes = [c for c in classes if str(c.get('period')) in morning_periods]
        elif session_type == 'afternoon':
            classes = [c for c in classes if str(c.get('period')) in afternoon_periods]
            
    # Get faculty list for assignments
    faculty = list(users.find({
        "role": "lecturer",
        "_id": {"$ne": ObjectId(current_user.id)}
    }, {"name": 1, "staff_id": 1, "phone": 1}).sort("name", 1))
    
    for f in faculty:
        f['_id'] = str(f['_id'])
        
    live_allocs = list(leave_class_allocations.find({
        "assigned_by": str(current_user.id),
        "status": {"$in": ["Pending", "pending", "accepted", "approved", "finalized", "pending_faculty", "requested"]}
    }))
    
    # Map allocations to classes by a unique key (subject+date+time) for easy frontend syncing
    alloc_map = {}
    for a in live_allocs:
        c = a.get('class_details') or {}
        # Normalize: Trim spaces and ignore case for robust matching
        sub = str(c.get('subject', '')).strip().upper()
        raw_key = f"{c.get('date')}_{c.get('time')}_{sub}"
        key = re.sub(r"\s+", "_", raw_key)
        alloc_map[key] = {
            "status": a.get('status'),
            "faculty_id": a.get('assigned_to'),
            "allocation_id": str(a.get('_id'))
        }

    # Get HOD status for current user (Only requests not yet linked to a leave)
    hod_req = hod_requests.find_one({
        "requester_id": str(current_user.id),
        "status": {"$in": ["Pending", "Approved", "Rejected"]},
        "leave_id": {"$exists": False}
    }, sort=[("created_at", -1)])
    
    hod_status = hod_req.get('status', 'Not Requested') if hod_req else 'Not Requested'

    # Get Assigned HOD details for UI (New Dynamic Logic)
    lecturer = users.find_one({"_id": ObjectId(current_user.id)})
    dept = lecturer.get('department') if lecturer else None
    hod_info = {"name": "HOD Not Assigned", "dept": dept or "Unknown Department"}
    if dept:
        hod_assignment = department_hods.find_one({"department": dept})
        if hod_assignment and hod_assignment.get('hod_id'):
            hod_user = users.find_one({"_id": ObjectId(hod_assignment['hod_id'])})
            if hod_user:
                hod_info["name"] = hod_user.get('name')
                hod_info["dept"] = f"HOD, {dept}"

    return jsonify({
        "success": True,
        "classes": classes,
        "faculty": faculty,
        "live_allocations": alloc_map,
        "hod_status": hod_status,
        "hod_info": hod_info
    })

@app.route('/lecturer/api/request-hod-permission', methods=['POST'])
@login_required
@lecturer_required
def api_request_hod_permission():
    # Get HOD based on department (Dynamic Logic)
    lecturer = users.find_one({"_id": ObjectId(current_user.id)})
    dept = lecturer.get('department')
    
    if not dept:
        return jsonify({"success": False, "message": "Your department is not set in your profile. Please contact Admin."}), 400
        
    hod_assignment = department_hods.find_one({"department": dept})
    if not hod_assignment:
        return jsonify({"success": False, "message": f"HOD for the {dept} department has not been assigned yet."}), 404
        
    hod_user = users.find_one({"_id": ObjectId(hod_assignment['hod_id'])})
    if not hod_user:
        return jsonify({"success": False, "message": "Assigned HOD user record not found."}), 404
        
    hod_id = str(hod_user['_id'])
    hod_name = hod_user.get('name', 'HOD')
    
    # Check if a request already exists for this user (draft/pending and not yet linked)
    existing = hod_requests.find_one({
        "requester_id": str(current_user.id),
        "status": "Pending",
        "leave_id": {"$exists": False}
    })
    if existing:
        return jsonify({"success": True, "message": "Request already pending.", "request_id": str(existing['_id'])})
        
    # Get extra details from request
    data = request.json or {}
    
    # Create request
    req_data = {
        "requester_id": str(current_user.id),
        "requester_name": current_user.name,
        "hod_id": hod_id,
        "hod_name": hod_name,
        "status": "Pending",
        "leave_type": data.get('type'),
        "from_date": data.get('from_date'),
        "to_date": data.get('to_date'),
        "reason": data.get('reason'),
        "half_day": data.get('half_day', False),
        "session": data.get('session'),
        "mode": data.get('mode', 'full'),
        "created_at": datetime.now()
    }
    res = hod_requests.insert_one(req_data)
    req_id = str(res.inserted_id)
    
    # Send Notification to HOD
    faculty_notifications.insert_one({
        "recipient_id": str(hod_user['_id']),
        "sender_id": str(current_user.id),
        "sender_name": current_user.name,
        "message": f"{current_user.name} has requested HOD permission for leave.",
        "type": "hod_permission",
        "request_id": req_id,
        "status": "unread",
        "created_at": datetime.now()
    })
    
    socketio.emit('new_hod_request', {
        "recipient_id": str(hod_user['_id']),
        "message": f"New HOD permission request from {current_user.name}",
        "request_id": req_id
    })
    
    return jsonify({"success": True, "request_id": req_id})

@app.route('/lecturer/api/respond-hod-permission/<request_id>/<action>', methods=['POST'])
@login_required
@lecturer_required
def respond_hod_permission(request_id, action):
    if action not in ['approve', 'reject']:
        return jsonify({"success": False, "message": "Invalid action"}), 400
        
    req = hod_requests.find_one({"_id": ObjectId(request_id)})
    if not req:
        return jsonify({"success": False, "message": "Request not found"}), 404
        
    # Verify this request is for the current user (HOD)
    if req.get('hod_id') != str(current_user.id):
        return jsonify({"success": False, "message": "Not authorized"}), 403
        
    new_status = 'Approved' if action == 'approve' else 'Rejected'
    hod_requests.update_one(
        {"_id": ObjectId(request_id)},
        {"$set": {"status": new_status, "responded_at": datetime.now()}}
    )
    
    # Notify the requester
    emit_to_user('hod_permission_response', req.get('requester_id'), {
        'request_id': request_id,
        'status': new_status,
        'message': f"HOD has {action}ed your permission request"
    })
    
    # MARK NOTIFICATION AS READ: To fix the persistent badge count
    faculty_notifications.update_many(
        {"request_id": request_id, "recipient_id": str(current_user.id)},
        {"$set": {"status": "read"}}
    )
    
    return jsonify({"success": True, "message": f"Request {action}ed successfully"})


@app.route('/lecturer/api/clear-hod-permission', methods=['POST'])
@login_required
@lecturer_required
def api_clear_hod_permission():
    """Clear HOD permission if form changes"""
    hod_requests.delete_many({
        "requester_id": str(current_user.id),
        "status": {"$in": ["Pending", "Approved", "Rejected"]},
        "leave_id": {"$exists": False}
    })
    return jsonify({"success": True})


@app.route('/lecturer/hod-permission/<request_id>/preview-sheet')
@login_required
@lecturer_required
def hod_preview_sheet(request_id):
    """Show a preview of the class allocations for an HOD permission request"""
    req = hod_requests.find_one({"_id": ObjectId(request_id)})
    if not req:
        flash("Request not found.", "danger")
        return redirect(url_for('my_class_assignments'))
    
    # Verify authorization (Only HOD or the Requester can see)
    if req.get('hod_id') != str(current_user.id) and req.get('requester_id') != str(current_user.id):
        flash("Not authorized.", "danger")
        return redirect(url_for('lecturer_dashboard'))
    
    # Fallback to requester's latest draft if request is missing metadata (dates/reason)
    if not req.get('from_date'):
        draft_doc = leave_drafts.find_one({"user_id": req.get('requester_id')})
        if draft_doc and 'draft_data' in draft_doc:
            d = draft_doc['draft_data']
            # Only update local 'req' object for display, don't mutate DB
            req['from_date'] = d.get('from_date')
            req['to_date'] = d.get('to_date')
            req['leave_type'] = d.get('type')
            req['reason'] = d.get('reason')
            req['half_day'] = d.get('half_day')
            req['session'] = d.get('session')

    # Lecturer details for the sheet
    lecturer = users.find_one({"_id": ObjectId(req.get('requester_id'))})
    staff_id = lecturer.get('staff_id') if lecturer else None
    
    # Calculate Total Days / Duration
    total_days = 0
    duration_str = None
    f_val = req.get('from_date', '')
    t_val = req.get('to_date', '')
    
    try:
        from datetime import datetime
        def parse_dt(dt_str):
            if not dt_str: return None
            dt_str = str(dt_str).strip()
            # Try full datetime formats
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M', '%Y/%m/%d %H:%M'):
                try: return datetime.strptime(dt_str, fmt)
                except: continue
            # Try date only
            for fmt in ('%Y-%m-%d', '%d-%m-%Y'):
                try: return datetime.strptime(dt_str, fmt)
                except: continue
            return None

        f_dt = parse_dt(f_val)
        t_dt = parse_dt(t_val)
        
        if f_dt and t_dt:
            # For time-based, calculate hours/minutes
            if ' ' in str(f_val) and ' ' in str(t_val):
                diff = t_dt - f_dt
                total_seconds = int(diff.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, _ = divmod(remainder, 60)
                
                if hours > 0:
                    duration_str = f"{hours}h {minutes}m"
                else:
                    duration_str = f"{minutes}m"
                total_days = total_seconds // 60
            else:
                total_days = count_working_leave_days(f_val, t_val, half_day=req.get('half_day', False))
        elif req.get('mode') == 'time':
            # Fallback for time mode if parsing failed but it's clearly a permission
            total_days = 60 # Default to 60 mins if unknown
            duration_str = "1h (Est.)"
        else:
            total_days = "N/A"
    except Exception as e:
        print(f"DEBUG: HOD Preview calc failed: {e}")
        total_days = 60 if req.get('mode') == 'time' else "N/A"
        duration_str = "1h (Est.)" if req.get('mode') == 'time' else None

    # For query filtering, extract just the date part
    f_str = f_val.split(' ')[0] if f_val else ''
    t_str = t_val.split(' ')[0] if t_val else ''
    
    # Get all allocations requested by this person that are currently active
    # AND fall within the requested leave period
    query = {
        "assigned_by": req.get('requester_id'),
        "status": {"$in": ["Pending", "pending", "pending_faculty", "accepted", "approved", "finalized"]}
    }
    
    if f_str and t_str:
        query["class_details.date"] = {"$gte": f_str, "$lte": t_str}
    
    allocations_raw = list(leave_class_allocations.find(query))
    
    # Original classes context
    original_classes = get_classes_for_leave_period(staff_id, f_str, t_str) if staff_id and f_str and t_str else []
    
    # Filter for Half Day Session
    if req.get('half_day'):
        session = str(req.get('session', '')).lower()
        if session == 'morning':
            # Morning = Periods 0, I, II, III
            original_classes = [c for c in original_classes if str(c.get('period')) in ['0', 'I', 'II', 'III']]
        elif session == 'afternoon':
            # Afternoon = Periods IV, V, VI, VII
            original_classes = [c for c in original_classes if str(c.get('period')) in ['IV', 'V', 'VI', 'VII']]

    # Unified list construction
    alloc_map = {}
    for alloc in allocations_raw:
        c = alloc.get('class_details', {})
        # Create a ROBUST key: Date + Period (Reliable across minor subject/class naming variations)
        key = f"{c.get('date')}_{c.get('period')}"
        alloc_map[key] = alloc
    
    display_list = []
    if not original_classes and allocations_raw:
        for alloc in allocations_raw:
            assigned_to = users.find_one({"_id": ObjectId(alloc.get('assigned_to', ''))})
            display_list.append({
                'class_details': alloc.get('class_details', {}),
                'status': alloc.get('status'),
                'assigned_to_name': assigned_to.get('name') if assigned_to else 'Unknown'
            })
    else:
        for c in original_classes:
            # Use the same ROBUST key for matching
            key = f"{c.get('date')}_{c.get('period')}"
            alloc = alloc_map.get(key)
            if alloc:
                assigned_to = users.find_one({"_id": ObjectId(alloc.get('assigned_to', ''))})
                display_list.append({
                    'class_details': c,
                    'status': alloc.get('status'),
                    'assigned_to_name': assigned_to.get('name') if assigned_to else 'Unknown'
                })
            else:
                display_list.append({'class_details': c, 'status': 'Pending', 'assigned_to_name': None})
            
    # Group by date
    dates_seen = {}
    for item in display_list:
        d = item['class_details'].get('date')
        if d not in dates_seen:
            dates_seen[d] = {'date': d, 'day': item['class_details'].get('day'), 'group_classes': []}
        dates_seen[d]['group_classes'].append(item)
    
    # Sort dates chronologically
    grouped_items = sorted(dates_seen.values(), key=lambda x: x['date'])

    # Update session text for better clarity
    if req.get('half_day'):
        s = str(req.get('session', '')).lower()
        if s == 'morning':
            req['session_display'] = 'Morning (Period 0 - III)'
        elif s == 'afternoon':
            req['session_display'] = 'Afternoon (Period IV - VII)'

    return render_template(
        'lecturer/hod_preview_sheet.html',
        hod_req=req,
        lecturer=lecturer,
        total_days=total_days,
        duration_str=duration_str,
        grouped_items=grouped_items,
        total_classes=len(display_list)
    )


@app.route('/lecturer/api/faculty-timetable/<staff_id>')
@login_required
def api_faculty_timetable(staff_id):
    """Get faculty timetable for a specific date"""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({"success": False, "message": "Date required"}), 400
    
    classes = get_classes_on_date(staff_id, date_str)
    faculty = users.find_one({"staff_id": staff_id})
    
    return jsonify({
        "success": True,
        "faculty_name": faculty.get('name') if faculty else 'Unknown',
        "classes": classes,
        "date": date_str,
        "day": get_day_name_from_date(date_str)
    })


@app.route('/lecturer/my-assignments')
@login_required
@lecturer_required
def my_class_assignments():
    if current_user.staff_id and current_user.staff_id.startswith('BBHCFN'):
        flash('Assignments are not applicable for your category.', 'info')
        return redirect(url_for('lecturer_dashboard'))
    """Show pending class assignments for the current faculty member"""
    # Get unread/pending notifications
    raw = list(faculty_notifications.find({
        "recipient_id": str(current_user.id),
        "status": {"$in": ["unread", "pending"]}
    }).sort("created_at", -1))
    
    # HOD CHECK: If user is assigned as HOD for ANY department (Dynamic Logic)
    is_hod = department_hods.find_one({"hod_id": str(current_user.id)}) is not None
    
    hod_pending = []
    if is_hod:
        hod_pending = list(hod_requests.find({
            "hod_id": str(current_user.id),
            "status": "Pending"
        }).sort("created_at", -1))
        for h in hod_pending: h['id'] = str(h['_id'])

    # Get accepted/approved assignments history
    accepted_assignments = list(leave_class_allocations.find({
        "assigned_to": str(current_user.id),
        "status": {"$in": ["accepted", "approved"]}
    }).sort("created_at", -1))
    
    # CLASS-BASED PURGE: Create a set of classes the user has already accepted/approved
    confirmed_set = set()
    for a in accepted_assignments:
        c = a.get('class_details', {})
        sub = str(c.get('subject', '')).strip().upper()
        conf_key = f"{c.get('date')}_{c.get('time')}_{sub}"
        confirmed_set.add(conf_key)

    # GHOST HUNTER: Verify each notification has a real allocation record and isn't a duplicate
    notifications_raw = []
    for n in raw:
        # 0. HOD SYNC: If it's an HOD permission, check if it's still pending
        if n.get('type') == 'hod_permission':
            req_id = n.get('request_id')
            if req_id:
                req_doc = hod_requests.find_one({"_id": ObjectId(req_id)})
                if not req_doc or req_doc.get('status') != 'Pending':
                    faculty_notifications.delete_one({"_id": n['_id']})
                continue
            continue

        alloc_id = n.get('allocation_id')
        c = n.get('class_details', {})
        sub = str(c.get('subject', '')).strip().upper()
        notif_key = f"{c.get('date')}_{c.get('time')}_{sub}"
        
        # 1. Check if they already have an official assignment for this exact class
        if notif_key in confirmed_set:
            faculty_notifications.delete_one({"_id": n['_id']})
            continue

        if alloc_id:
            exists = leave_class_allocations.find_one({"_id": ObjectId(alloc_id)})
            if exists:
                # 2. AUTO-CLOSURE: If this specific assignment is already processed (Accepted/Rejected/Approved/Finalized)
                if exists.get('status') in ['approved', 'finalized', 'accepted', 'rejected']:
                    faculty_notifications.delete_one({"_id": n['_id']})
                else:
                    notifications_raw.append(n)
            else:
                # 3. BREAKING THE GHOST: Assignment is gone
                faculty_notifications.delete_one({"_id": n['_id']})
        else:
            notifications_raw.append(n)


    # Fetch user timetable to check for existing subjects in requested slots
    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    structured = tt_doc.get("structured") if tt_doc else {}
    
    # Convert notifications to JSON-serializable format
    notifications = []
    for n in notifications_raw:
        class_details = n.get('class_details', {})
        existing_subject = "Free Slot"
        
        if structured and class_details:
            req_day = class_details.get('day', '').upper()
            req_time = class_details.get('time', '')
            
            period_key = None
            if structured.get('timetable') and structured['timetable'].get('periods'):
                for p in structured['timetable']['periods']:
                    if p.get('time') == req_time:
                        period_key = p.get('period')
                        break
                        
            if period_key and structured['timetable'].get('days'):
                for d in structured['timetable']['days']:
                    if d.get('day', '').upper() == req_day:
                        slot = d.get('slots', {}).get(period_key, {})
                        if slot and slot.get('subject'):
                            existing_subject = slot.get('subject')
                        break

        notifications.append({
            'id': str(n['_id']),
            'allocation_id': n.get('allocation_id'),
            'leave_id': n.get('leave_id'),
            'recipient_id': n.get('recipient_id'),
            'sender_id': n.get('sender_id'),
            'sender_name': n.get('sender_name'),
            'message': n.get('message'),
            'class_details': class_details,
            'status': n.get('status'),
            'created_at': n.get('created_at'),
            'existing_subject': existing_subject
        })
    # Get all related allocations
    allocation_ids = [n.get('allocation_id') for n in notifications if n.get('allocation_id')]
    allocations = []
    for alloc_id in allocation_ids:
        alloc = leave_class_allocations.find_one({"_id": ObjectId(alloc_id)})
        if alloc:
            allocations.append({
                'id': str(alloc['_id']),
                'leave_id': alloc.get('leave_id'),
                'assigned_by': alloc.get('assigned_by'),
                'assigned_to': alloc.get('assigned_to'),
                'class_details': alloc.get('class_details', {}),
                'status': alloc.get('status'),
                'created_at': alloc.get('created_at')
            })
    from datetime import datetime
    
    upcoming_assignments = []
    completed_assignments = []
    now = datetime.now()
    
    for a in accepted_assignments:
        c = a.get('class_details', {})
        date_str = c.get('date', '')
        time_str = c.get('time', '')
        
        is_completed = False
        try:
            if date_str:
                end_time_str = time_str.split('-')[-1].strip().replace('.', ':')
                time_parts = end_time_str.split(':')
                hour = int(time_parts[0])
                minute = int(time_parts[1]) if len(time_parts) > 1 else 0
                if hour < 8:
                    hour += 12
                    
                if '-' in date_str and len(date_str) == 10:
                    y, m, d = map(int, date_str.split('-'))
                elif '-' in date_str:
                    parts = date_str.split('-')
                    y, m, d = (int(parts[0]), int(parts[1]), int(parts[2])) if len(parts[0]) == 4 else (int(parts[2]), int(parts[1]), int(parts[0]))
                else:
                    y, m, d = map(int, date_str.split('/'))
                    
                class_end_dt = datetime(y, m, d, hour, minute)
                a['sort_dt'] = class_end_dt
                if class_end_dt < now:
                    is_completed = True
        except Exception:
            try:
                if date_str and date_str < now.strftime('%Y-%m-%d'):
                    is_completed = True
                a['sort_dt'] = now
            except:
                a['sort_dt'] = now
                
        a['is_completed'] = is_completed
        if is_completed:
            completed_assignments.append(a)
        else:
            upcoming_assignments.append(a)
            
    upcoming_assignments.sort(key=lambda x: x.get('sort_dt', now))
    completed_assignments.sort(key=lambda x: x.get('sort_dt', now), reverse=True)

    return render_template(
        'lecturer/my_assignments.html',
        notifications=notifications,
        allocations=allocations,
        accepted_assignments=accepted_assignments,
        upcoming_assignments=upcoming_assignments,
        completed_assignments=completed_assignments,
        hod_pending=hod_pending,
        is_hod=is_hod
    )

@app.route('/lecturer/assignment/<allocation_id>/<action>', methods=['POST'])
@login_required
@lecturer_required
def respond_to_assignment(allocation_id, action):
    """Accept or reject a class assignment"""
    if action not in ['accept', 'reject']:
        return jsonify({"success": False, "message": "Invalid action"}), 400
    
    allocation = leave_class_allocations.find_one({"_id": ObjectId(allocation_id)})
    if not allocation:
        return jsonify({"success": False, "message": "Assignment not found"}), 404
    
    # Verify this assignment is for the current user
    if allocation.get('assigned_to') != current_user.id:
        return jsonify({"success": False, "message": "Not authorized"}), 403
    
    new_status = 'accepted' if action == 'accept' else 'rejected'
    leave_class_allocations.update_one(
        {"_id": ObjectId(allocation_id)},
        {"$set": {"status": new_status, "responded_at": datetime.now()}}
    )
    
    # PERMANENT REMOVAL: Delete the notification so it never reappears on refresh
    faculty_notifications.delete_many(
        {"allocation_id": allocation_id, "recipient_id": current_user.id}
    )
    
    # GHOST PURGE: Also delete any other notifications for the SAME class and recipient
    # (In case there were duplicate requests from draft resets)
    c = allocation.get('class_details', {})
    faculty_notifications.delete_many({
        "recipient_id": current_user.id,
        "class_details.date": c.get('date'),
        "class_details.time": c.get('time'),
        "class_details.subject": c.get('subject')
    })
    
    # Notify the leave applicant
    recipient_id = None
    leave_id = allocation.get('leave_id')
    if leave_id:
        leave_doc = leaves.find_one({"_id": ObjectId(leave_id)})
        if leave_doc:
            recipient_id = leave_doc.get('lecturer_id')
    
    # If no leave_id (draft mode), use assigned_by
    if not recipient_id:
        recipient_id = allocation.get('assigned_by')

    if recipient_id:
        emit_to_user('assignment_response', recipient_id, {
            'allocation_id': allocation_id,
            'action': action,
            'message': f"{current_user.name} has {action}ed your class assignment"
        })
    
    # If rejected, update leave to indicate class allocation failed
    if action == 'reject':
        leaves.update_one(
            {"_id": ObjectId(allocation.get('leave_id'))},
            {"$set": {"class_allocation_status": "needs_reassignment"}}
        )
    
    return jsonify({
        "success": True,
        "message": f"Assignment {action}ed successfully"
    })


# Admin routes for managing class allocations
@app.route('/admin/leave/<leave_id>/class-allocation-sheet')
@login_required
def admin_class_allocation_sheet(leave_id):
    """Show detailed formal sheet view of class allocations for a leave request"""
    leave_doc = leaves.find_one({"_id": ObjectId(leave_id)})
    if not leave_doc:
        # Check permissions collection if not in leaves
        leave_doc = permissions.find_one({"_id": ObjectId(leave_id)})
        if not leave_doc:
            flash("Leave or Permission request not found.", "danger")
            return redirect(url_for('admin_dashboard'))
    
    # Check if user is admin OR the lecturer who applied
    if current_user.role != 'admin' and str(leave_doc.get('lecturer_id')) != str(current_user.id):
        flash("Not authorized to view this sheet.", "danger")
        return redirect(url_for('lecturer_dashboard'))
    
    # Get all allocations for this leave
    allocations_raw = list(leave_class_allocations.find({
        "leave_id": leave_id,
        "status": {"$in": ["Pending", "pending", "pending_faculty", "accepted", "approved", "finalized"]}
    }))
    
    # Fresh Lecturer Details (for Dept/Designation)
    lecturer = users.find_one({"_id": ObjectId(leave_doc.get('lecturer_id', ''))})
    staff_id = lecturer.get('staff_id') if lecturer else None
    
    # Calculate Total Days
    total_days = 0
    duration_str = None
    try:
        from datetime import datetime
        f_val = leave_doc.get('from_date', '')
        t_val = leave_doc.get('to_date', '')
        
        f_str = f_val.split(' ')[0]
        t_str = t_val.split(' ')[0]
        total_days = count_working_leave_days(f_str, t_str, half_day=leave_doc.get('half_day', False))

        if leave_doc.get('mode') == 'time':
            try:
                def parse_dt(dt_str):
                    if not dt_str: return None
                    dt_str = str(dt_str).strip()
                    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M', '%Y/%m/%d %H:%M'):
                        try: return datetime.strptime(dt_str, fmt)
                        except: continue
                    # Try to split by space and take first two parts
                    try:
                        parts = dt_str.split(' ')
                        if len(parts) >= 2:
                            d_part = parts[0]
                            t_part = parts[1]
                            # Try date + time
                            for d_fmt in ('%Y-%m-%d', '%d-%m-%Y'):
                                for t_fmt in ('%H:%M:%S', '%H:%M'):
                                    try: return datetime.strptime(f"{d_part} {t_part}", f"{d_fmt} {t_fmt}")
                                    except: continue
                    except: pass
                    return None
                
                f_dt_full = parse_dt(f_val)
                t_dt_full = parse_dt(t_val)
                
                if f_dt_full and t_dt_full:
                    diff = t_dt_full - f_dt_full
                    total_seconds = int(diff.total_seconds())
                    hours, remainder = divmod(total_seconds, 3600)
                    minutes, _ = divmod(remainder, 60)
                    duration_str = f"{hours}h {minutes}m"
                    total_days = 1 
            except: pass
        else:
            print(f"DEBUG: Sheet Mode is '{leave_doc.get('mode')}' for leave {leave_id}")
    except Exception as e:
        print(f"DEBUG: Main total_days calculation failed: {e}")
        total_days = "N/A"

    # Get Leave Stats for the Credits section (CL, EL, RH)
    all_stats = get_all_leave_stats(str(leave_doc.get('lecturer_id')))
    leave_stats_map = {
        'CL': next((s['left'] for s in all_stats if s['type'] == 'Casual Leave'), 0),
        'EL': next((s['left'] for s in all_stats if s['type'] == 'Earned Leave'), 0),
        'RH': next((s['left'] for s in all_stats if s['type'] == 'Restricted Holiday'), 0)
    }
    
    original_classes = get_classes_for_leave_period(staff_id, f_str, t_str) if staff_id else []

    # Filter for Half Day Session if applicable
    if leave_doc.get('half_day'):
        session = str(leave_doc.get('session', '')).lower()
        if session == 'morning':
            # Morning = Periods 0, I, II, III
            original_classes = [c for c in original_classes if str(c.get('period')) in ['0', 'I', 'II', 'III']]
        elif session == 'afternoon':
            # Afternoon = Periods IV, V, VI, VII
            original_classes = [c for c in original_classes if str(c.get('period')) in ['IV', 'V', 'VI', 'VII']]

    # Construct a unified list of items to display
    # We start with the FULL set of classes for the period (Original Timetable)
    # and then 'overlay' any existing allocations on top of them.
    
    # 1. Map existing allocations for easy lookup
    alloc_map = {}
    for alloc in allocations_raw:
        c = alloc.get('class_details', {})
        # Create a ROBUST key: Date + Period (Reliable across minor subject/class naming variations)
        key = f"{c.get('date')}_{c.get('period')}"
        alloc_map[key] = alloc
    
    # 2. Build the display list based on the full period's classes
    display_list = []
    
    # Fallback: if we have no original classes (e.g. error), just show allocations
    if not original_classes and allocations_raw:
        for alloc in allocations_raw:
            display_list.append({
                'class_details': alloc.get('class_details', {}),
                'status': alloc.get('status'),
                'assigned_to_name': alloc.get('assigned_to_name', 'Unknown')
            })
    else:
        for c in original_classes:
            # Use the same ROBUST key for matching
            key = f"{c.get('date')}_{c.get('period')}"
            alloc = alloc_map.get(key)
            
            if alloc:
                assigned_to = users.find_one({"_id": ObjectId(alloc.get('assigned_to', ''))}) if alloc.get('assigned_to') else None
                display_list.append({
                    'class_details': c,
                    'status': alloc.get('status'),
                    'assigned_to_name': assigned_to.get('name') if assigned_to else 'Unknown'
                })
            else:
                display_list.append({
                    'class_details': c,
                    'status': 'Pending',
                    'assigned_to_name': None
                })
            
    # Group by date
    dates_seen = {}
    for item in display_list:
        date = item['class_details'].get('date')
        if date not in dates_seen:
            dates_seen[date] = []
        dates_seen[date].append(item)
    
    # Maintain chronological order
    unique_dates = []
    for item in display_list:
        date = item['class_details'].get('date')
        if date not in unique_dates:
            unique_dates.append(date)
            
    grouped_items = []
    sort_map = {"0": 0, "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6, "VII": 7}

    for date in unique_dates:
        items_for_date = dates_seen[date]
        items_for_date.sort(key=lambda x: sort_map.get(str(x['class_details'].get('period')), 99))
        
        grouped_items.append({
            'date': date,
            'day': items_for_date[0]['class_details'].get('day'),
            'group_classes': items_for_date
        })
    
    # Calculate total classes for row filling logic in template
    total_classes = sum(len(g['group_classes']) for g in grouped_items)
    
    # Load original timetable backup
    original_timetable = None
    backup_doc = timetable_history.find_one({
        "staff_id": staff_id,
        "reason": "leave_assignment"
    })
    if backup_doc:
        original_timetable = backup_doc.get('original_data')
    
    # Get HOD approval status for this specific leave
    hod_req = hod_requests.find_one({
        "leave_id": leave_id,
        "status": "Approved"
    })
    hod_approved = hod_req is not None
    hod_signature = None
    if hod_req:
        hod_user = users.find_one({"_id": ObjectId(hod_req['hod_id'])})
        if hod_user:
            hod_signature = hod_user.get('signature_path')

    return render_template(
        'admin/leave_application_sheet.html',
        leave=leave_doc,
        leave_id=leave_id,
        lecturer=lecturer,
        grouped_items=grouped_items,
        total_classes=total_classes,
        leave_stats=leave_stats_map,
        total_days=total_days,
        duration_str=duration_str,
        original_timetable=original_timetable,
        applicant_name=leave_doc.get('lecturer_name'),
        hod_approved=hod_approved,
        hod_signature=hod_signature,
        request_is_time=(
            leave_doc.get('mode') == 'time'
            or (leave_doc.get('type') or '').strip().lower() == 'permission'
        ),
    )


@app.route('/admin/leave/<leave_id>/finalize-allocation', methods=['POST'])
@login_required
@admin_required
def admin_finalize_allocation(leave_id):
    """Finalize class allocations and update timetables"""
    leave_doc = leaves.find_one({"_id": ObjectId(leave_id)})
    if not leave_doc:
        return jsonify({"success": False, "message": "Leave not found"}), 404
    
    # Get all accepted allocations
    allocations = list(leave_class_allocations.find({
        "leave_id": leave_id,
        "status": "accepted"
    }))
    
    if not allocations:
        return jsonify({"success": False, "message": "No accepted class allocations found"}), 400
    
    # Get applicant's staff_id
    applicant = users.find_one({"_id": ObjectId(leave_doc.get('lecturer_id'))})
    applicant_staff_id = applicant.get('staff_id') if applicant else None
    
    # Save original timetable backup before making changes
    if applicant_staff_id:
        original_timetable = load_faculty_timetable(applicant_staff_id)
        if original_timetable:
            save_timetable_backup(applicant_staff_id, original_timetable, "leave_assignment")
    
    # Update timetables for each allocation
    updated_count = 0
    for alloc in allocations:
        assigned_to_id = alloc.get('assigned_to')
        class_details = alloc.get('class_details', {})
        
        assigned_faculty = users.find_one({"_id": ObjectId(assigned_to_id)})
        assigned_staff_id = assigned_faculty.get('staff_id') if assigned_faculty else None
        
        if assigned_staff_id:
            # Add class to assigned faculty's timetable
            success = add_class_to_timetable(assigned_staff_id, class_details)
            if success:
                updated_count += 1
                # Remove class from applicant's timetable
                if applicant_staff_id:
                    remove_class_from_timetable(applicant_staff_id, class_details)
    
    # Update leave status and allocation status
    leaves.update_one(
        {"_id": ObjectId(leave_id)},
        {"$set": {
            "class_allocation_status": "completed",
            "class_allocation_finalized_at": datetime.now(),
            "status": "Approved"  # Auto-approve after allocation
        }}
    )
    
    # Update all allocations to approved
    leave_class_allocations.update_many(
        {"leave_id": leave_id, "status": "accepted"},
        {"$set": {"status": "approved", "approved_at": datetime.now()}}
    )
    
    return jsonify({
        "success": True,
        "message": f"Class allocation finalized. {updated_count} classes assigned.",
        "updated_count": updated_count
    })


def add_class_to_timetable(staff_id, class_details):
    """Add a class to a faculty's timetable"""
    try:
        json_dir = os.path.join(os.path.dirname(__file__), "static", "json_timetables")
        json_path = os.path.join(json_dir, f"{staff_id}.json")
        
        if not os.path.exists(json_path):
            return False
        
        with open(json_path, 'r', encoding='utf-8') as f:
            timetable_data = json.load(f)
        
        day_name = class_details.get('day', '').upper()
        
        # Find the day and add the slot
        if isinstance(timetable_data, dict):
            for day_data in timetable_data.get('timetable', []):
                if day_data.get('day', '').upper() == day_name:
                    new_slot = {
                        'time': class_details.get('time'),
                        'subject': class_details.get('subject'),
                        'room': class_details.get('room', ''),
                        'is_assigned': True,
                        'original_faculty': class_details.get('original_faculty')
                    }
                    if 'slots' not in day_data:
                        day_data['slots'] = []
                    day_data['slots'].append(new_slot)
                    break
        elif isinstance(timetable_data, list):
            for day_data in timetable_data:
                if day_data.get('day', '').upper() == day_name:
                    new_slot = {
                        'time': class_details.get('time'),
                        'subject': class_details.get('subject'),
                        'room': class_details.get('room', ''),
                        'is_assigned': True,
                        'original_faculty': class_details.get('original_faculty')
                    }
                    if 'slots' not in day_data:
                        day_data['slots'] = []
                    day_data['slots'].append(new_slot)
                    break
        
        # Save updated timetable
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(timetable_data, f, indent=2, ensure_ascii=False)
        
        return True
    except Exception as e:
        print(f"Error adding class to timetable: {e}")
        return False


def remove_class_from_timetable(staff_id, class_details):
    """Remove a class from a faculty's timetable (mark as on-leave)"""
    try:
        json_dir = os.path.join(os.path.dirname(__file__), "static", "json_timetables")
        json_path = os.path.join(json_dir, f"{staff_id}.json")
        
        if not os.path.exists(json_path):
            return False
        
        with open(json_path, 'r', encoding='utf-8') as f:
            timetable_data = json.load(f)
        
        day_name = class_details.get('day', '').upper()
        time_slot = class_details.get('time')
        
        # Find and mark the slot as on-leave
        if isinstance(timetable_data, dict):
            for day_data in timetable_data.get('timetable', []):
                if day_data.get('day', '').upper() == day_name:
                    for slot in day_data.get('slots', []):
                        if slot.get('time') == time_slot:
                            slot['on_leave'] = True
                            slot['assigned_to'] = class_details.get('assigned_to_name')
                            break
                    break
        elif isinstance(timetable_data, list):
            for day_data in timetable_data:
                if day_data.get('day', '').upper() == day_name:
                    for slot in day_data.get('slots', []):
                        if slot.get('time') == time_slot:
                            slot['on_leave'] = True
                            slot['assigned_to'] = class_details.get('assigned_to_name')
                            break
                    break
        
        # Save updated timetable
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(timetable_data, f, indent=2, ensure_ascii=False)
        
        return True
    except Exception as e:
        print(f"Error removing class from timetable: {e}")
        return False


@app.route('/lecturer/api/notifications')
@login_required
@lecturer_required
def api_get_notifications():
    """Get unread notifications for current user with auto-cleanup of ghosts"""
    raw = list(faculty_notifications.find({
        "recipient_id": str(current_user.id),
        "status": "unread"
    }).sort("created_at", -1))
    
    valid_notifications = []
    
    for n in raw:
        # 1. HOD Sync check
        if n.get('type') == 'hod_permission':
            req_id = n.get('request_id')
            if req_id:
                req_doc = hod_requests.find_one({"_id": ObjectId(req_id)})
                if not req_doc or req_doc.get('status') != 'Pending':
                    faculty_notifications.delete_one({"_id": n['_id']})
                    continue
            valid_notifications.append(n)
            continue
            
        # 2. Class Assignment check
        alloc_id = n.get('allocation_id')
        if alloc_id:
            exists = leave_class_allocations.find_one({"_id": ObjectId(alloc_id)})
            if not exists or exists.get('status') in ['approved', 'finalized', 'accepted', 'rejected']:
                faculty_notifications.delete_one({"_id": n['_id']})
                continue
        
        valid_notifications.append(n)

    # Convert ObjectIds to strings for JSON serialization
    for n in valid_notifications:
        n['_id'] = str(n['_id'])
        if 'leave_id' in n:
            n['leave_id'] = str(n['leave_id'])
    
    return jsonify({
        "success": True,
        "notifications": valid_notifications,
        "count": len(valid_notifications)
    })


@app.route('/admin/permissions')
@login_required
@admin_required
def admin_permissions():
    """Manage Permission Leave requests (separate from formal leaves)"""
    all_permissions = list(permissions.find().sort("created_at", -1))
    # Add display dates/times
    for p in all_permissions:
        p['_id'] = str(p['_id'])
        # Duration string for display
        try:
            from datetime import datetime
            f_val = p.get('from_date', '')
            t_val = p.get('to_date', '')
            def parse_dt(dt_str):
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M'):
                    try: return datetime.strptime(dt_str, fmt)
                    except: continue
                return None
            f_dt = parse_dt(f_val)
            t_dt = parse_dt(t_val)
            if f_dt and t_dt:
                diff = t_dt - f_dt
                hours, remainder = divmod(diff.total_seconds(), 3600)
                minutes, _ = divmod(remainder, 60)
                p['duration_display'] = f"{int(hours)}h {int(minutes)}m"
        except:
            p['duration_display'] = "N/A"
            
    return render_template('admin/permissions.html', permissions=all_permissions)


@app.route('/admin/permission/api/<pid>/<status>', methods=['POST'])
@login_required
@admin_required
def admin_api_review_permission(pid, status):
    """API endpoint for reviewing permission requests from the dashboard"""
    permission_doc = permissions.find_one({"_id": ObjectId(pid)})
    if not permission_doc:
        return jsonify({"success": False, "message": "Not found"}), 404
        
    permissions.update_one({"_id": ObjectId(pid)}, {"$set": {"status": status}})
    
    if status == "Approved":
        # Extract times for permission
        t_from = None
        t_to = None
        try:
            t_from = permission_doc['from_date'].split(' ')[1]
            t_to = permission_doc['to_date'].split(' ')[1]
        except: pass

        update_attendance_log_on_approval(
            permission_doc['lecturer_id'],
            permission_doc['from_date'],
            is_permission=True,
            time_from=t_from,
            time_to=t_to
        )

    # Notify via socket for real-time dashboard updates
    leaves_left = calculate_leaves_left(permission_doc['lecturer_id'])
    socketio.emit('leave_status_update', {
        'id': pid,
        'status': status,
        'lecturer_id': permission_doc['lecturer_id'],
        'leaves_left': leaves_left
    })
    
    return jsonify({"success": True})

@app.route('/admin/permission/<pid>/update-status', methods=['POST'])
@login_required
@admin_required
def admin_update_permission_status(pid):
    """Update status for a permission request (Form-based)"""
    status = request.form.get('status')
    
    # NEW: Fetch the permission doc to get lecturer_id and date
    permission_doc = permissions.find_one({"_id": ObjectId(pid)})
    
    permissions.update_one({"_id": ObjectId(pid)}, {"$set": {"status": status}})
    
    # NEW: Update attendance log if approved
    if status == "Approved" and permission_doc:
        # Extract times for permission
        t_from = None
        t_to = None
        try:
            t_from = permission_doc['from_date'].split(' ')[1]
            t_to = permission_doc['to_date'].split(' ')[1]
        except: pass

        update_attendance_log_on_approval(
            permission_doc['lecturer_id'],
            permission_doc['from_date'],
            is_permission=True,
            time_from=t_from,
            time_to=t_to
        )
        
    flash(f"Permission status updated to {status}.", "success")
    return redirect(url_for('admin_permissions'))

@app.route('/admin/permissions/delete-all', methods=['POST'])
@login_required
@admin_required
def admin_permissions_delete_all():
    """Wipe all permission leave records from the system"""
    result = permissions.delete_many({})
    count = result.deleted_count
    flash(f"Deleted {count} permission record(s).", "success")
    return redirect(url_for('admin_permissions'))




@app.route('/lecturer/api/cancel-substitution/<allocation_id>', methods=['POST'])
@login_required
@lecturer_required
def cancel_substitution(allocation_id):
    """Cancel a pending substitution request sent by the lecturer"""
    try:
        from bson import ObjectId
        
        # Find the allocation first to know who to notify
        alloc = leave_class_allocations.find_one({
            "_id": ObjectId(allocation_id),
            "assigned_by": str(current_user.id)
        })
        
        if alloc:
            recipient_id = alloc.get('assigned_to')
            # Delete the allocation record
            leave_class_allocations.delete_one({"_id": ObjectId(allocation_id)})
            # Delete the notification
            faculty_notifications.delete_one({"allocation_id": allocation_id})
            
            # Emit socket event to the recipient to remove it from their UI immediately
            emit_to_user('assignment_recalled', recipient_id, {
                "allocation_id": allocation_id,
            })
            
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route('/api/chat', methods=['POST'])
@login_required
def api_chat():
    try:
        import traceback
        data = request.get_json() or {}
        message = data.get('message', '')
        page_name = data.get('page', 'Unknown')
        chat_history = data.get('history') or []
        
        # 1. Gather Basic Context
        context = {
            "user_name": current_user.name,
            "user_role": current_user.role,
            "current_page": page_name,
            "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        
        # 2. Gather Role-Specific Live Data
        if current_user.role == 'admin':
            context.update({
                "pending_leaves_count": leaves.count_documents({"status": "Pending"}),
                "pending_permissions_count": permissions.count_documents({"status": "Pending"}),
                "total_staff_count": users.count_documents({"role": "lecturer"}),
            })
        else:
            # For faculty, get their specific data
            staff_id = current_user.staff_id
            context.update({
                "my_staff_id": staff_id,
                "my_pending_leaves": leaves.count_documents({"lecturer_id": str(current_user.id), "status": "Pending"}),
            })
            
            # 1. Fetch Schedules using robust helpers
            today_dt = datetime.now()
            tomorrow_dt = today_dt + timedelta(days=1)
            
            def get_readable_day(dt):
                cls_list = get_classes_on_date(staff_id, dt.strftime('%Y-%m-%d'))
                if not cls_list: return "No classes"
                return " | ".join([f"Period {c['period']} ({c['time']}): {c['subject']} for {c['class']}" for c in cls_list])

            context["my_schedule_today"] = get_readable_day(today_dt)
            context["my_schedule_tomorrow"] = get_readable_day(tomorrow_dt)
            
            weekly_summary = []
            for i in range(7):
                d = today_dt + timedelta(days=i)
                d_name = d.strftime('%A').upper()
                cls = get_classes_on_date(staff_id, d.strftime('%Y-%m-%d'))
                if cls:
                    weekly_summary.append(f"{d_name}: {len(cls)} classes")
            
            context["my_weekly_summary"] = ", ".join(weekly_summary)
            context["today_day_name"] = today_dt.strftime('%A').upper()
            context["today_full_date"] = today_dt.strftime('%d %B %Y')
            
            # 2. Fetch Attendance Stats
            att_percent = calculate_lecturer_attendance_stats(staff_id)
            context["my_attendance_percentage"] = f"{att_percent}%"
            
            # 3. Fetch Leave Balances
            user_doc = users.find_one({"_id": ObjectId(current_user.id)})
            if user_doc and "leave_balances" in user_doc:
                bal_str = ", ".join([f"{k}: {v}" for k, v in user_doc["leave_balances"].items()])
                context["my_leave_balances"] = bal_str
            
        # 3. Get AI Response Stream

        def generate():
            try:
                for chunk in get_hrms_response_stream(message, context, chat_history=chat_history):
                    yield f"data: {json.dumps({'text': chunk})}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'error': str(e)})}\n\n"

        return Response(generate(), mimetype='text/event-stream')
    except Exception as e:
        with open("chatbot_errors.log", "a") as f:
            f.write(f"\n--- Error at {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        return jsonify({"success": False, "message": str(e)})


# --- Staff WhatsApp-style messaging (separate from AI chatbot) ---
STAFF_CHAT_UPLOAD_DIR = os.path.join(os.getcwd(), 'static', 'uploads', 'chat')
STAFF_CHAT_ALLOWED_EXT = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.pdf', '.doc', '.docx', '.txt', '.xls', '.xlsx'}
STAFF_CHAT_MAX_BYTES = 10 * 1024 * 1024


def _staff_chat_conversation_id(uid_a, uid_b):
    return '_'.join(sorted([str(uid_a), str(uid_b)]))


def _staff_chat_user_avatar(user_doc):
    staff_id = (user_doc or {}).get('staff_id', '')
    if staff_id:
        rel = os.path.join('static', 'img', 'profiles', f'{staff_id}.png')
        if os.path.exists(rel):
            return f'/static/img/profiles/{staff_id}.png'
    return None


def _staff_chat_user_payload(user_doc):
    return {
        'id': str(user_doc['_id']),
        'name': user_doc.get('name') or user_doc.get('username', 'User'),
        'role': user_doc.get('role', ''),
        'department': user_doc.get('department', ''),
        'staff_id': user_doc.get('staff_id', ''),
        'avatar': _staff_chat_user_avatar(user_doc),
    }


def _staff_chat_other_participant(conv, my_id):
    for pid in conv.get('participants', []):
        if pid != str(my_id):
            return pid
    return None


def _staff_chat_socket_id(user_id):
    """Return active Socket.IO socket id for a given user."""
    if not user_id:
        return None
    doc = staff_socket_sessions.find_one({'user_id': str(user_id)})
    socket_id = (doc or {}).get('socket_id')
    return socket_id if socket_id else None


@app.route('/staff-chat')
@login_required
def staff_chat_page():
    return render_template('staff_chat.html')


@app.route('/api/staff-chat/contacts')
@login_required
def staff_chat_contacts():
    my_id = str(current_user.id)
    query = {'_id': {'$ne': ObjectId(my_id)}}
    if current_user.role == 'lecturer':
        query['role'] = {'$in': ['lecturer', 'admin']}
    else:
        query['role'] = {'$in': ['lecturer', 'admin']}
    docs = list(users.find(query).sort('name', 1))
    return jsonify([_staff_chat_user_payload(u) for u in docs])


@app.route('/api/staff-chat/conversations')
@login_required
def staff_chat_conversations():
    my_id = str(current_user.id)
    convs = list(staff_conversations.find({'participants': my_id}).sort('updated_at', -1))
    result = []
    for conv in convs:
        if my_id in (conv.get('hidden_for') or []):
            continue
        other_id = _staff_chat_other_participant(conv, my_id)
        if not other_id:
            continue
        other_doc = users.find_one({'_id': ObjectId(other_id)})
        if not other_doc:
            continue
        unread = (conv.get('unread') or {}).get(my_id, 0)
        updated = conv.get('updated_at')
        result.append({
            'conversation_id': conv.get('conversation_id'),
            'other_user': _staff_chat_user_payload(other_doc),
            'last_message': conv.get('last_message', ''),
            'last_sender_name': conv.get('last_sender_name', ''),
            'updated_at': updated.isoformat() if updated else None,
            'unread': unread,
        })
    return jsonify(result)


@app.route('/api/staff-chat/messages/<other_user_id>')
@login_required
def staff_chat_messages(other_user_id):
    my_id = str(current_user.id)
    conv_id = _staff_chat_conversation_id(my_id, other_user_id)
    conv = staff_conversations.find_one({'conversation_id': conv_id}) or {}
    cleared_at = (conv.get('cleared_at') or {}).get(my_id)
    query = {'conversation_id': conv_id}
    if cleared_at:
        query['created_at'] = {'$gt': cleared_at}
    msgs = list(staff_messages.find(query).sort('created_at', 1).limit(200))
    staff_conversations.update_one(
        {'conversation_id': conv_id},
        {'$set': {f'unread.{my_id}': 0}},
    )
    out = []
    for m in msgs:
        if my_id in (m.get('deleted_for') or []):
            continue
        att = m.get('attachment')
        out.append({
            'id': str(m['_id']),
            'sender_id': m.get('sender_id'),
            'sender_name': m.get('sender_name', ''),
            'text': m.get('text', ''),
            'attachment': att,
            'deleted_for_all': bool(m.get('deleted_for_all')),
            'created_at': m['created_at'].isoformat() if m.get('created_at') else None,
            'is_mine': m.get('sender_id') == my_id,
        })
    return jsonify({'conversation_id': conv_id, 'messages': out})


@app.route('/api/staff-chat/send', methods=['POST'])
@login_required
def staff_chat_send():
    my_id = str(current_user.id)
    payload = request.get_json(silent=True) or {}
    other_id = request.form.get('recipient_id') or payload.get('recipient_id')
    text = (request.form.get('text') or payload.get('text') or '').strip()
    if not other_id:
        return jsonify({'success': False, 'message': 'Recipient is required'}), 400
    if other_id == my_id:
        return jsonify({'success': False, 'message': 'Cannot chat with yourself'}), 400

    recipient_doc = users.find_one({'_id': ObjectId(other_id)})
    if not recipient_doc:
        return jsonify({'success': False, 'message': 'Recipient not found'}), 404

    uploaded_file = request.files.get('file')
    has_uploaded_file = bool(uploaded_file and uploaded_file.filename)
    if not text and not has_uploaded_file:
        return jsonify({'success': False, 'message': 'Message or file is required'}), 400

    attachment = None
    if has_uploaded_file:
        ext = os.path.splitext(uploaded_file.filename)[1].lower()
        if ext not in STAFF_CHAT_ALLOWED_EXT:
            return jsonify({'success': False, 'message': 'File type not allowed'}), 400
        uploaded_file.seek(0, os.SEEK_END)
        size = uploaded_file.tell()
        uploaded_file.seek(0)
        if size > STAFF_CHAT_MAX_BYTES:
            return jsonify({'success': False, 'message': 'File too large (max 10MB)'}), 400
        os.makedirs(STAFF_CHAT_UPLOAD_DIR, exist_ok=True)
        safe_name = re.sub(r'[^\w.\-]', '_', uploaded_file.filename)[:120]
        filename = f"{int(datetime.now().timestamp())}_{my_id[:6]}_{safe_name}"
        save_path = os.path.join(STAFF_CHAT_UPLOAD_DIR, filename)
        uploaded_file.save(save_path)
        is_image = ext in {'.png', '.jpg', '.jpeg', '.gif', '.webp'}
        attachment = {
            'url': f'/static/uploads/chat/{filename}',
            'name': uploaded_file.filename,
            'type': 'image' if is_image else 'file',
        }

    conv_id = _staff_chat_conversation_id(my_id, other_id)
    now = datetime.now()
    preview = text or (f"📎 {attachment['name']}" if attachment else '')
    msg_doc = {
        'conversation_id': conv_id,
        'sender_id': my_id,
        'sender_name': current_user.name or current_user.username,
        'text': text,
        'attachment': attachment,
        'created_at': now,
    }
    inserted = staff_messages.insert_one(msg_doc)
    staff_conversations.update_one(
        {'conversation_id': conv_id},
        {
            '$set': {
                'conversation_id': conv_id,
                'participants': sorted([my_id, str(other_id)]),
                'last_message': preview[:200],
                'last_sender_name': current_user.name or current_user.username,
                'updated_at': now,
            },
            '$inc': {f'unread.{other_id}': 1},
            '$pull': {'hidden_for': {'$in': [my_id, str(other_id)]}},
        },
        upsert=True,
    )
    payload = {
        'id': str(inserted.inserted_id),
        'conversation_id': conv_id,
        'sender_id': my_id,
        'sender_name': current_user.name or current_user.username,
        'text': text,
        'attachment': attachment,
        'created_at': now.isoformat(),
        'is_mine': True,
        'other_user_id': other_id,
    }
    # Emit to the sender and recipient so outgoing/incoming both appear instantly.
    sender_sid = _staff_chat_socket_id(my_id)
    recipient_sid = _staff_chat_socket_id(other_id)

    if sender_sid:
        socketio.emit('staff_chat_message', payload, to=sender_sid)
    else:
        socketio.emit('staff_chat_message', payload, room=f'user_{my_id}')

    if recipient_sid:
        # On recipient screen this should render as "theirs".
        payload_for_recipient = dict(payload)
        payload_for_recipient['is_mine'] = False
        socketio.emit('staff_chat_message', payload_for_recipient, to=recipient_sid)
    else:
        socketio.emit('staff_chat_message', dict(payload, is_mine=False), room=f'user_{other_id}')
    return jsonify({'success': True, 'message': payload})


@app.route('/api/staff-chat/message/<msg_id>/delete', methods=['POST'])
@login_required
def staff_chat_delete_message(msg_id):
    my_id = str(current_user.id)
    body = request.get_json(silent=True) or {}
    scope = body.get('scope', 'me')
    other_user_id = body.get('other_user_id')

    if not msg_id or str(msg_id).lower() in {'null', 'none', ''}:
        return jsonify({'success': False, 'message': 'Invalid message id'}), 400
    try:
        msg_obj_id = ObjectId(msg_id)
    except Exception:
        return jsonify({'success': False, 'message': 'Invalid message id'}), 400

    msg = staff_messages.find_one({'_id': msg_obj_id})
    if not msg:
        return jsonify({'success': False, 'message': 'Message not found'}), 404

    conv_id = msg.get('conversation_id')
    if not conv_id:
        return jsonify({'success': False, 'message': 'Invalid message'}), 400

    conv = staff_conversations.find_one({'conversation_id': conv_id}) or {}
    if my_id not in (conv.get('participants') or []):
        return jsonify({'success': False, 'message': 'Not allowed'}), 403

    if scope == 'all':
        if msg.get('sender_id') != my_id:
            return jsonify({'success': False, 'message': 'Delete for all allowed only for your messages'}), 403
        now = datetime.now()
        staff_messages.update_one(
            {'_id': ObjectId(msg_id)},
            {
                '$set': {
                    'deleted_for_all': True,
                    'text': '',
                    'attachment': None,
                    'deleted_at': now,
                    'deleted_by': my_id,
                }
            },
        )
        if other_user_id:
            recipient_sid = _staff_chat_socket_id(other_user_id)
            if recipient_sid:
                socketio.emit(
                    'staff_chat_message_deleted',
                    {'message_id': msg_id, 'conversation_id': conv_id, 'scope': 'all'},
                    to=recipient_sid,
                )
            else:
                socketio.emit(
                    'staff_chat_message_deleted',
                    {'message_id': msg_id, 'conversation_id': conv_id, 'scope': 'all'},
                    room=f'user_{other_user_id}',
                )
        return jsonify({'success': True})

    staff_messages.update_one(
        {'_id': ObjectId(msg_id)},
        {'$addToSet': {'deleted_for': my_id}},
    )
    return jsonify({'success': True})


@app.route('/api/staff-chat/conversation/<other_user_id>/delete', methods=['POST'])
@login_required
def staff_chat_delete_conversation(other_user_id):
    my_id = str(current_user.id)
    conv_id = _staff_chat_conversation_id(my_id, other_user_id)
    now = datetime.now()
    staff_conversations.update_one(
        {'conversation_id': conv_id},
        {
            '$addToSet': {'hidden_for': my_id},
            '$set': {f'cleared_at.{my_id}': now, f'unread.{my_id}': 0},
        },
        upsert=True,
    )
    return jsonify({'success': True})


@app.route('/api/staff-chat/unread-count')
@login_required
def staff_chat_unread_count():
    my_id = str(current_user.id)
    total = 0
    for conv in staff_conversations.find({'participants': my_id}, {'unread': 1}):
        total += (conv.get('unread') or {}).get(my_id, 0)
    return jsonify({'count': total})


@socketio.on('connect')
def staff_chat_socket_connect():
    # Try to join the user's room immediately.
    # If `current_user` is unavailable, `register_socket` (from the client) will handle it.
    try:
        if current_user.is_authenticated:
            sid = request.sid
            user_id = str(current_user.id)
            staff_socket_sessions.update_one(
                {'user_id': user_id},
                {'$set': {'socket_id': sid, 'updated_at': datetime.now()}},
                upsert=True,
            )
            join_room(f'user_{user_id}')
    except Exception:
        # Never block connect due to DB issues.
        pass


@socketio.on('disconnect')
def staff_chat_socket_disconnect():
    try:
        sid = request.sid
        staff_socket_sessions.update_one(
            {'socket_id': sid},
            {'$set': {'socket_id': None, 'updated_at': datetime.now()}},
        )
    except Exception:
        # Never block disconnect due to DB issues.
        pass


@socketio.on('register_socket')
def staff_chat_register_socket(data):
    """
    Store browser Socket.IO socket id in Mongo so we can emit directly to the recipient.
    """
    data = data or {}
    uid = str(data.get('user_id', '')).strip()
    if not uid:
        return False
    try:
        ObjectId(uid)
    except Exception:
        return False
    # Validate user exists (prevents storing arbitrary socket ids).
    if not users.find_one({'_id': ObjectId(uid)}):
        return False
    sid = request.sid
    staff_socket_sessions.update_one(
        {'user_id': uid},
        {'$set': {'socket_id': sid, 'updated_at': datetime.now()}},
        upsert=True,
    )
    join_room(f'user_{uid}')
    return True


if __name__ == '__main__':
    init_db()
    # On some Windows setups (especially with newer Python), the watchdog reloader can throw WinError 10038.
    # Disabling the reloader keeps dev runs stable; restart the server manually after code changes.
    socketio.run(app, host='0.0.0.0', debug=True, port=8000, use_reloader=False, allow_unsafe_werkzeug=True)

